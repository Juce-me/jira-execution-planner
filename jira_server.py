#!/usr/bin/env python3

from flask import abort, has_request_context, jsonify, redirect, request, send_file, send_from_directory, session
import requests
import argparse
import base64
import copy
import csv
import logging
import os
import re
import json
import hashlib
import threading
import time
import uuid
from contextlib import nullcontext
from datetime import datetime, timedelta, date, timezone
try:
    from zoneinfo import ZoneInfo
except Exception:  # pragma: no cover - Python < 3.9 fallback
    ZoneInfo = None
from urllib.parse import parse_qs, urlparse
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import io
from requests import Session
from concurrent.futures import ThreadPoolExecutor, as_completed, TimeoutError as FuturesTimeoutError
from backend.epm import config as epm_config
from backend.epm import home as epm_home
from backend.epm import aggregate as epm_aggregate
from backend.epm import payload as epm_payload
from backend.epm.home import fetch_epm_home_projects, merge_epm_linkage
from backend.epm.rollup import EpmRollupDependencies, build_per_project_rollup
from backend.epm.scope import build_epm_scope_clause, normalize_epm_sprint_field, should_apply_epm_sprint
from planning import Issue, ScheduledIssue, ScenarioConfig, compute_slack, schedule_issues
from backend.auth.cache_policy import (
    build_jira_home_process_cache_key,
    jira_home_partitioned_process_cache_enabled,
    jira_home_process_cache_enabled,
)
from backend.auth.context import RequestAuthContext, build_auth_cache_key, stable_local_workspace_id
from backend.auth.admin_bootstrap import bootstrap_first_tool_admin
from backend.auth.csrf import validate_csrf_token
from backend.auth.db_context import is_db_auth_context, resolve_db_request_auth_context
from backend.auth.db_tokens import db_oauth_session_data, store_oauth_callback_tokens
from backend.auth.key_provider import key_provider_from_env
from backend.auth.local_oauth_store import LocalOAuthStoreConfig, LocalOAuthTokenStore
from backend.auth.project_access import project_access_denied_response
from backend.auth.service_integrations import register_service_integration_cache_invalidator
from backend.config.repository import (
    ConfigStorageError,
    config_storage_db_enabled,
    db_repository as build_db_config_repository,
    json_repository as build_json_config_repository,
    validate_config_storage_startup,
)
from backend.db.engine import DatabaseConfigurationError, database_storage_enabled, session_scope
from backend.auth.jira_auth import (
    AUTH_MODE_ATLASSIAN_OAUTH,
    AUTH_MODE_BASIC,
    AuthConfig,
    AuthError,
    build_authorize_url,
    build_pkce_challenge,
    choose_accessible_resource,
    exchange_authorization_code,
    fetch_accessible_resources,
    fetch_current_user,
    jira_get,
    jira_post,
    jira_request,
    missing_oauth_scopes,
    new_oauth_state,
    new_pkce_verifier,
    token_session_payload,
    validate_auth_config,
)

# Load environment variables from .env file before constructing the Flask app.
load_dotenv()
from backend.app import create_app
from backend import config_store as _config_store
from backend import jira_client as _jira_client
from backend.services import capacity as _capacity_service
from backend.services import sprints as _sprints_service
from backend.services import stats_cache as _stats_cache_service
from backend.services import update_check as _update_check_service
from backend.epm import projects as epm_projects
from backend.security.policy import (
    is_oauth_ready_api_path as policy_is_oauth_ready_api_path,
    oauth_ready_api_paths,
    shared_config_write_paths,
)

# Reuse a single HTTP session to avoid reconnect overhead on repeated calls
HTTP_SESSION = Session()
logger = logging.getLogger(__name__)

# CONFIGURATION - Load from environment variables
JIRA_URL = os.getenv('JIRA_URL')
JIRA_EMAIL = os.getenv('JIRA_EMAIL')
JIRA_TOKEN = os.getenv('JIRA_TOKEN')
JIRA_AUTH_MODE = os.getenv('JIRA_AUTH_MODE', AUTH_MODE_BASIC).strip() or AUTH_MODE_BASIC
ATLASSIAN_CLIENT_ID = os.getenv('ATLASSIAN_CLIENT_ID', '').strip()
ATLASSIAN_CLIENT_SECRET = os.getenv('ATLASSIAN_CLIENT_SECRET', '').strip()
ATLASSIAN_REDIRECT_URI = os.getenv('ATLASSIAN_REDIRECT_URI', '').strip()
ATLASSIAN_SCOPES = os.getenv(
    'ATLASSIAN_SCOPES',
    'read:me read:jira-work read:jira-user read:board-scope:jira-software read:sprint:jira-software read:project:jira offline_access',
).strip()
FLASK_SECRET_KEY = os.getenv('FLASK_SECRET_KEY', '').strip()
app = create_app()
app.secret_key = FLASK_SECRET_KEY or os.urandom(32)
APP_ENVIRONMENT_KEY = os.getenv('APP_ENVIRONMENT_KEY', 'local').strip() or 'local'
OAUTH_LOCAL_TOKEN_STORE_ALLOWED = os.getenv('OAUTH_LOCAL_TOKEN_STORE_ALLOWED', '').strip().lower() in {'1', 'true', 'yes'}
OAUTH_TOKEN_STORE_TTL_SECONDS = int(os.getenv('OAUTH_TOKEN_STORE_TTL_SECONDS', '2592000'))
OAUTH_TOKEN_STORE_MIN_TTL_SECONDS = 900
OAUTH_TOKEN_STORE_PATH = os.getenv('OAUTH_TOKEN_STORE_PATH', '.oauth-token-store.json').strip()
JQL_QUERY = os.getenv('JQL_QUERY', '').strip()
JIRA_BOARD_ID = os.getenv('JIRA_BOARD_ID')  # Optional: board ID for faster sprint fetching
JIRA_PRODUCT_PROJECT = os.getenv('JIRA_PRODUCT_PROJECT', 'PRODUCT ROADMAPS')
JIRA_TECH_PROJECT = os.getenv('JIRA_TECH_PROJECT', 'TECHNICAL ROADMAP')
SERVER_PORT = int(os.getenv('SERVER_PORT', '5050'))
EPIC_EMPTY_EXCLUDED_STATUSES = [s.strip() for s in os.getenv('EPIC_EMPTY_EXCLUDED_STATUSES', 'Killed,Done,Incomplete').split(',') if s.strip()]
EPIC_EMPTY_TEAM_IDS = [s.strip() for s in os.getenv('EPIC_EMPTY_TEAM_IDS', '').split(',') if s.strip()]
MISSING_INFO_COMPONENT = os.getenv('MISSING_INFO_COMPONENT', '').strip()
MISSING_INFO_TEAM_IDS = [s.strip() for s in os.getenv('MISSING_INFO_TEAM_IDS', '').split(',') if s.strip()]
STATS_JQL_BASE = os.getenv('STATS_JQL_BASE', '').strip()
STATS_JQL_ORDER_BY = os.getenv('STATS_JQL_ORDER_BY', 'ORDER BY cf[30101] ASC, status ASC').strip()
STATS_PRODUCT_PROJECTS = [s.strip() for s in os.getenv('STATS_PRODUCT_PROJECTS', JIRA_PRODUCT_PROJECT).split(',') if s.strip()]
STATS_TECH_PROJECTS = [s.strip() for s in os.getenv('STATS_TECH_PROJECTS', JIRA_TECH_PROJECT).split(',') if s.strip()]
STATS_TEAM_IDS = [s.strip() for s in os.getenv('STATS_TEAM_IDS', '').split(',') if s.strip()]
STATS_PRIORITY_WEIGHTS = os.getenv('STATS_PRIORITY_WEIGHTS', '').strip()
CAPACITY_PROJECT = os.getenv('CAPACITY_PROJECT', '').strip()
CAPACITY_FIELD_ID = os.getenv('CAPACITY_FIELD_ID', '').strip()
CAPACITY_FIELD_NAME = os.getenv('CAPACITY_FIELD_NAME', '').strip()
GROUPS_CONFIG_PATH = os.getenv('GROUPS_CONFIG_PATH', '').strip()
DASHBOARD_CONFIG_PATH = os.getenv('DASHBOARD_CONFIG_PATH', '').strip()
TEAM_CATALOG_PATH = os.getenv('TEAM_CATALOG_PATH', '').strip()
SCENARIO_OVERRIDES_PATH = os.getenv('SCENARIO_OVERRIDES_PATH', '').strip()
TEAM_GROUPS_JSON = os.getenv('TEAM_GROUPS_JSON', '').strip()
JQL_QUERY_TEMPLATE = os.getenv('JQL_QUERY_TEMPLATE', '').strip()
DEBUG_MODE = os.getenv('DEBUG_MODE', 'false').lower() == 'true'
SETTINGS_ADMIN_ONLY = os.getenv('SETTINGS_ADMIN_ONLY', 'true').lower() not in ('0', 'false', 'no')
UPDATE_CHECK_ENABLED = os.getenv('UPDATE_CHECK', 'true').lower() not in ('0', 'false', 'no')
UPDATE_CHECK_REMOTE = os.getenv('UPDATE_CHECK_REMOTE', 'origin').strip() or 'origin'
UPDATE_CHECK_BRANCH = os.getenv('UPDATE_CHECK_BRANCH', 'main').strip() or 'main'
UPDATE_CHECK_TTL_SECONDS = int(os.getenv('UPDATE_CHECK_TTL_SECONDS', '300'))
UPDATE_CHECK_RELEASE_INFO = os.getenv('UPDATE_CHECK_RELEASE_INFO', 'release-info.json').strip() or 'release-info.json'
LOG_LEVEL = os.getenv('LOG_LEVEL', 'INFO').strip().upper() or 'INFO'
JIRA_RETRY_MAX_ATTEMPTS = int(os.getenv('JIRA_RETRY_MAX_ATTEMPTS', '4'))
JIRA_RETRY_MAX_ELAPSED_SECONDS = float(os.getenv('JIRA_RETRY_MAX_ELAPSED_SECONDS', '10'))
JIRA_RETRY_BASE_DELAY_SECONDS = float(os.getenv('JIRA_RETRY_BASE_DELAY_SECONDS', '0.5'))
JIRA_RETRY_MAX_DELAY_SECONDS = float(os.getenv('JIRA_RETRY_MAX_DELAY_SECONDS', '3'))
JIRA_CIRCUIT_FAILURE_THRESHOLD = int(os.getenv('JIRA_CIRCUIT_FAILURE_THRESHOLD', '5'))
JIRA_CIRCUIT_OPEN_SECONDS = float(os.getenv('JIRA_CIRCUIT_OPEN_SECONDS', '30'))
STATS_BURNOUT_TIMEZONE = 'Europe/Berlin'
EPIC_COHORT_CACHE_TTL_SECONDS = int(os.getenv('EPIC_COHORT_CACHE_TTL_SECONDS', '300'))
EPIC_COHORT_ENRICH_MAX_ISSUES = int(os.getenv('EPIC_COHORT_ENRICH_MAX_ISSUES', '200'))
EPIC_COHORT_ENRICH_WORKERS = int(os.getenv('EPIC_COHORT_ENRICH_WORKERS', '4'))
EPIC_COHORT_ENRICH_TIMEOUT_SECONDS = float(os.getenv('EPIC_COHORT_ENRICH_TIMEOUT_SECONDS', '10'))
EXCLUDED_CAPACITY_STATS_MAX_SPRINTS = int(os.getenv('EXCLUDED_CAPACITY_STATS_MAX_SPRINTS', '24'))
EXCLUDED_CAPACITY_STATS_MAX_ISSUES = int(os.getenv('EXCLUDED_CAPACITY_STATS_MAX_ISSUES', '2000'))
EXCLUDED_CAPACITY_STATS_MAX_EPICS = int(os.getenv('EXCLUDED_CAPACITY_STATS_MAX_EPICS', '200'))
EXCLUDED_CAPACITY_STATS_SOURCE_CACHE_TTL_SECONDS = int(os.getenv('EXCLUDED_CAPACITY_STATS_SOURCE_CACHE_TTL_SECONDS', '300'))
EXCLUDED_CAPACITY_EPIC_SUMMARY_CACHE_TTL_SECONDS = int(os.getenv('EXCLUDED_CAPACITY_EPIC_SUMMARY_CACHE_TTL_SECONDS', '3600'))
EXCLUDED_CAPACITY_EPIC_SUMMARY_BATCH_SIZE = int(os.getenv('EXCLUDED_CAPACITY_EPIC_SUMMARY_BATCH_SIZE', '100'))

SCENARIO_CACHE = {'generatedAt': None, 'data': None}
TASKS_CACHE = {}
TASKS_CACHE_TTL_SECONDS = 60 * 5
TASKS_CACHE_SCHEMA_VERSION = 'v2-empty-epic-actionable'
MISSING_INFO_CACHE = {}
MISSING_INFO_CACHE_TTL_SECONDS = 60 * 5
DEPENDENCIES_CACHE = {}
DEPENDENCIES_CACHE_TTL_SECONDS = 60 * 5
UPDATE_CHECK_CACHE = {'ts': 0, 'data': None}
EPIC_COHORT_CACHE = {}
EXCLUDED_CAPACITY_STATS_SOURCE_CACHE = {}
EXCLUDED_CAPACITY_EPIC_SUMMARY_CACHE = {}
EPM_PROJECTS_CACHE = {}
EPM_ISSUES_CACHE = {}
EPM_ROLLUP_CACHE = {}
OAUTH_TOKEN_STORE = {}
OAUTH_TOKEN_STORE_LOCK = threading.RLock()
OAUTH_REFRESH_LOCKS = {}
EPM_PROJECTS_CACHE_TTL_SECONDS = 300
EPM_ISSUES_CACHE_TTL_SECONDS = 300
EPM_ROLLUP_CACHE_TTL_SECONDS = 300
EPM_ROLLUP_QUERY_MAX_RESULTS = 2000
_epm_cache_lock = threading.Lock()

# Single lock for all global caches — kept simple since these are not hot paths.
_cache_lock = threading.RLock()

# Cache settings
SPRINTS_CACHE_FILE = 'sprints_cache.json'
STATS_CACHE_FILE = 'stats_cache.json'
CACHE_EXPIRY_HOURS = 24
GROUPS_CONFIG_VERSION = 1
GROUPS_MAX_TEAMS = 12
PRIORITY_WEIGHT_DEFAULTS = [
    {'priority': 'Blocker', 'weight': 0.4},
    {'priority': 'Critical', 'weight': 0.3},
    {'priority': 'Major', 'weight': 0.2},
    {'priority': 'Minor', 'weight': 0.06},
    {'priority': 'Low', 'weight': 0.03},
    {'priority': 'Trivial', 'weight': 0.01},
]
PRIORITY_WEIGHT_NAME_ALIASES = {
    'highest': 'blocker',
    'high': 'major',
    'medium': 'minor',
    'lowest': 'trivial',
}


SENSITIVE_CALLBACK_QUERY_RE = re.compile(
    r'(/api/auth/atlassian/callback)\?[^ \t\r\n"]+'
)


def redact_sensitive_log_text(value):
    if not isinstance(value, str):
        return value
    return SENSITIVE_CALLBACK_QUERY_RE.sub(r'\1?[redacted]', value)


class SensitiveLogRedactionFilter(logging.Filter):
    def filter(self, record):
        record.msg = redact_sensitive_log_text(record.msg)
        if isinstance(record.args, tuple):
            record.args = tuple(redact_sensitive_log_text(arg) for arg in record.args)
        elif isinstance(record.args, dict):
            record.args = {
                key: redact_sensitive_log_text(value)
                for key, value in record.args.items()
            }
        return True


def _install_sensitive_log_filter(target_logger):
    if any(isinstance(existing, SensitiveLogRedactionFilter) for existing in target_logger.filters):
        return
    target_logger.addFilter(SensitiveLogRedactionFilter())


def configure_logging():
    """Initialize process logging once with a readable default format."""
    class CsvLineFormatter(logging.Formatter):
        """Emit one CSV record per log line: timestamp,level,logger,message."""
        def format(self, record):
            message = record.getMessage()
            if record.exc_info:
                message = f'{message}\n{self.formatException(record.exc_info)}'
            # Keep one CSV record per visible log line.
            lines = str(message).replace('\r', '\\r').split('\n')
            ts = self.formatTime(record, self.datefmt)
            rows = []
            for line in lines:
                buf = io.StringIO()
                writer = csv.writer(buf)
                writer.writerow([ts, record.levelname, record.name, line])
                rows.append(buf.getvalue().rstrip('\r\n'))
            return '\n'.join(rows)

    formatter = CsvLineFormatter(datefmt='%Y-%m-%dT%H:%M:%S')
    root_logger = logging.getLogger()
    _install_sensitive_log_filter(root_logger)
    _install_sensitive_log_filter(logging.getLogger('werkzeug'))
    _install_sensitive_log_filter(logger)
    if not root_logger.handlers:
        logging.basicConfig(
            level=getattr(logging, LOG_LEVEL, logging.INFO),
            handlers=[logging.StreamHandler()]
        )
    else:
        root_logger.setLevel(getattr(logging, LOG_LEVEL, logging.INFO))
    for handler in root_logger.handlers:
        handler.setFormatter(formatter)
    logger.setLevel(getattr(logging, LOG_LEVEL, logging.INFO))


def _format_log_parts(parts):
    if not parts:
        return ''
    return ' '.join(str(part) for part in parts)


def log_debug(*parts):
    message = _format_log_parts(parts)
    if not message:
        return
    logger.debug(message)


def log_info(*parts):
    message = _format_log_parts(parts)
    if not message:
        return
    logger.info(message)


def log_warning(*parts):
    message = _format_log_parts(parts)
    if not message:
        return
    logger.warning(message)


def log_error(*parts):
    message = _format_log_parts(parts)
    if not message:
        return
    logger.error(message)


configure_logging()


UNSAFE_METHODS = {'POST', 'PUT', 'PATCH', 'DELETE'}
OAUTH_DASHBOARD_ENTRY_PATHS = {'/', '/jira-dashboard.html'}
OAUTH_READY_API_PATHS = oauth_ready_api_paths()
OAUTH_SHARED_CONFIG_WRITE_PATHS = shared_config_write_paths()


def is_oauth_ready_api_path(path):
    return policy_is_oauth_ready_api_path(path)


def bootstrap_tool_admin_account_ids():
    raw = os.getenv('TOOL_ADMIN_ATLASSIAN_ACCOUNT_IDS', '')
    return {account_id.strip() for account_id in raw.split(',') if account_id.strip()}


def is_pre_db_tool_admin_account(atlassian_account_id):
    account_id = str(atlassian_account_id or '').strip()
    return bool(account_id)


def store_db_oauth_callback_session_metadata(token_data, resource, user_profile):
    if not database_storage_enabled():
        return {}
    with session_scope() as db_session:
        stored = store_oauth_callback_tokens(
            db_session,
            token_data=token_data,
            resource=resource,
            user_profile=user_profile,
            environment_key=APP_ENVIRONMENT_KEY.strip().lower() or 'local',
            configured_jira_url=JIRA_URL or '',
            key_provider=key_provider_from_env(),
            requested_scopes=ATLASSIAN_SCOPES,
        )
        bootstrap_first_tool_admin(
            db_session,
            workspace_id=stored.workspace_id,
            user_id=stored.user_id,
            atlassian_account_id=(user_profile or {}).get('account_id'),
        )
        clear_auth_sensitive_caches('oauth_reconnect')
        return stored.session_metadata


def current_auth_config():
    return AuthConfig(
        auth_mode=JIRA_AUTH_MODE,
        jira_url=JIRA_URL or '',
        jira_email=JIRA_EMAIL or '',
        jira_token=JIRA_TOKEN or '',
        client_id=ATLASSIAN_CLIENT_ID,
        client_secret=ATLASSIAN_CLIENT_SECRET,
        redirect_uri=ATLASSIAN_REDIRECT_URI,
        scopes=ATLASSIAN_SCOPES,
        flask_secret_key=FLASK_SECRET_KEY,
    )


def _local_oauth_store_config():
    return LocalOAuthStoreConfig(
        auth_mode=JIRA_AUTH_MODE,
        oauth_mode=AUTH_MODE_ATLASSIAN_OAUTH,
        environment_key=APP_ENVIRONMENT_KEY,
        persistence_allowed=OAUTH_LOCAL_TOKEN_STORE_ALLOWED,
        token_store_path=OAUTH_TOKEN_STORE_PATH,
        ttl_seconds=OAUTH_TOKEN_STORE_TTL_SECONDS,
    )


_LOCAL_OAUTH_STORE = LocalOAuthTokenStore(
    token_store=OAUTH_TOKEN_STORE,
    refresh_locks=OAUTH_REFRESH_LOCKS,
    store_lock=OAUTH_TOKEN_STORE_LOCK,
    config=_local_oauth_store_config,
    new_session_id=new_oauth_state,
    now=lambda: time.time(),
    logger=logger,
)


def _drop_oauth_session(session_id):
    _LOCAL_OAUTH_STORE.drop_session(session_id)


def _oauth_token_store_persistence_enabled():
    return _LOCAL_OAUTH_STORE.persistence_enabled()


def _db_oauth_browser_session_payload(data):
    if JIRA_AUTH_MODE != AUTH_MODE_ATLASSIAN_OAUTH or not database_storage_enabled():
        return {}
    connection_id = str((data or {}).get('db_auth_connection_id') or '').strip()
    if not connection_id:
        return {}
    return {'db_auth_connection_id': connection_id}


def remember_db_oauth_browser_session(data):
    payload = _db_oauth_browser_session_payload(data)
    if payload:
        session['db_oauth_session'] = payload
    else:
        session.pop('db_oauth_session', None)


def db_oauth_browser_session_data():
    if JIRA_AUTH_MODE != AUTH_MODE_ATLASSIAN_OAUTH or not database_storage_enabled():
        return {}
    stored = session.get('db_oauth_session')
    if isinstance(stored, dict):
        payload = _db_oauth_browser_session_payload(stored)
        if payload:
            return payload
    local_session = oauth_session_data()
    payload = _db_oauth_browser_session_payload(local_session)
    if payload:
        session['db_oauth_session'] = payload
    return payload


def strict_db_oauth_browser_session_data():
    if JIRA_AUTH_MODE != AUTH_MODE_ATLASSIAN_OAUTH or not database_storage_enabled():
        return {}
    stored = session.get('db_oauth_session')
    if isinstance(stored, dict):
        return _db_oauth_browser_session_payload(stored)
    return {}


def save_oauth_session(data):
    if not data:
        session.pop('db_oauth_session', None)
        session_id = session.pop('atlassian_oauth_session_id', None)
        if session_id:
            refresh_lock = _LOCAL_OAUTH_STORE.existing_refresh_lock(session_id)
            if refresh_lock:
                with refresh_lock:
                    _drop_oauth_session(session_id)
            else:
                _drop_oauth_session(session_id)
        return
    remember_db_oauth_browser_session(data)
    _LOCAL_OAUTH_STORE.save_session(session, data)


def oauth_refresh_lock():
    return _LOCAL_OAUTH_STORE.refresh_lock(session)


def oauth_session_data():
    return _LOCAL_OAUTH_STORE.session_data(session)


def jira_session_data():
    if JIRA_AUTH_MODE == AUTH_MODE_ATLASSIAN_OAUTH:
        return oauth_session_data()
    return {}


def oauth_session_id_from_auth_context(context):
    connection_id = getattr(context, 'auth_connection_id', '') or ''
    prefix = 'local-oauth-connection:'
    if connection_id.startswith(prefix):
        return connection_id[len(prefix):]
    return ''


def oauth_session_data_for_auth_context(context):
    if JIRA_AUTH_MODE != AUTH_MODE_ATLASSIAN_OAUTH:
        return {}
    session_id = oauth_session_id_from_auth_context(context)
    if not session_id and is_db_auth_context(context):
        return db_oauth_session_data_for_auth_context(context)
    if not session_id:
        return {}
    return _LOCAL_OAUTH_STORE.session_data_for_id(session_id)


def db_oauth_session_data_for_auth_context(context):
    with session_scope() as db_session:
        return db_oauth_session_data(
            db_session,
            context,
            config=current_auth_config(),
            key_provider=key_provider_from_env(),
            http_post=HTTP_SESSION.post,
        )


def save_oauth_session_for_auth_context(context, data):
    session_id = oauth_session_id_from_auth_context(context)
    if not session_id:
        return
    _LOCAL_OAUTH_STORE.save_session_for_id(session_id, data)


def oauth_refresh_lock_for_auth_context(context):
    session_id = oauth_session_id_from_auth_context(context)
    return _LOCAL_OAUTH_STORE.refresh_lock_for_id(session_id)


def current_jira_session_data(context=None):
    if context is not None and JIRA_AUTH_MODE == AUTH_MODE_ATLASSIAN_OAUTH:
        return oauth_session_data_for_auth_context(context)
    return jira_session_data()


def current_request_auth_context():
    if JIRA_AUTH_MODE == AUTH_MODE_ATLASSIAN_OAUTH and database_storage_enabled():
        db_session_data = db_oauth_browser_session_data()
        if db_session_data:
            return resolve_db_request_auth_context(
                db_session_data,
                required_scopes=ATLASSIAN_SCOPES,
            )
    session_data = jira_session_data()
    site_url = (session_data.get('site_url') or JIRA_URL or '').strip().rstrip('/')
    cloud_id = session_data.get('cloudid', '')
    workspace_id = stable_local_workspace_id(APP_ENVIRONMENT_KEY, site_url, cloud_id)
    if JIRA_AUTH_MODE == AUTH_MODE_ATLASSIAN_OAUTH:
        session_id = session.get('atlassian_oauth_session_id') or ''
        account_id = session_data.get('account_id', '')
        return RequestAuthContext(
            auth_mode=AUTH_MODE_ATLASSIAN_OAUTH,
            user_id=f'local-oauth-user:{account_id}',
            stable_subject=account_id,
            atlassian_account_id=account_id,
            workspace_id=workspace_id,
            auth_connection_id=f'local-oauth-connection:{session_id}',
            cloud_id=cloud_id,
            site_url=site_url,
            token_version=str(session_data.get('stored_at', '1')),
            account_status=session_data.get('account_status', ''),
            is_admin=is_pre_db_tool_admin_account(account_id),
        )
    return RequestAuthContext(
        auth_mode=AUTH_MODE_BASIC,
        user_id='local-basic-user',
        stable_subject='local-basic',
        atlassian_account_id='',
        workspace_id=workspace_id,
        auth_connection_id='local-basic-connection',
        cloud_id='',
        site_url=site_url,
        token_version='1',
        account_status='active',
        is_admin=True,
    )


def oauth_auth_required_payload():
    save_oauth_session({})
    return {
        'error': 'auth_required',
        'message': 'Your Jira sign-in expired. Sign in again to continue.',
        'loginUrl': '/login?reason=session_expired',
    }, 401


def csrf_session_data_for_request():
    if JIRA_AUTH_MODE == AUTH_MODE_ATLASSIAN_OAUTH and database_storage_enabled():
        context = scenario_draft_request_auth_context()
        return {
            'db_auth_connection_id': context.auth_connection_id,
            'db_token_version': context.token_version,
            'account_id': context.atlassian_account_id,
        }
    return oauth_session_data()


def scenario_draft_request_auth_context():
    if JIRA_AUTH_MODE == AUTH_MODE_ATLASSIAN_OAUTH and database_storage_enabled():
        db_session_data = strict_db_oauth_browser_session_data()
        if not db_session_data:
            raise AuthError('auth_required', 'Atlassian authentication is required.')
        return resolve_db_request_auth_context(
            db_session_data,
            required_scopes=ATLASSIAN_SCOPES,
        )
    return current_request_auth_context()


def current_jira_auth_context(context=None):
    if context is not None:
        return context
    if JIRA_AUTH_MODE == AUTH_MODE_ATLASSIAN_OAUTH and not has_request_context():
        raise AuthError('auth_required', 'Atlassian authentication is required.')
    return current_request_auth_context()


def current_oauth_session_callbacks(context=None):
    if JIRA_AUTH_MODE != AUTH_MODE_ATLASSIAN_OAUTH:
        return {}
    if context is not None:
        if is_db_auth_context(context):
            return {
                'save_session': lambda data: None,
                'reload_session': lambda: current_jira_session_data(context),
                'refresh_lock': nullcontext(),
            }
        return {
            'save_session': lambda data: save_oauth_session_for_auth_context(context, data),
            'reload_session': lambda: oauth_session_data_for_auth_context(context),
            'refresh_lock': oauth_refresh_lock_for_auth_context(context),
        }
    if not has_request_context():
        raise AuthError('auth_required', 'Atlassian authentication is required.')
    return {
        'save_session': save_oauth_session,
        'reload_session': oauth_session_data,
        'refresh_lock': oauth_refresh_lock(),
    }


def current_jira_get(path, *, params=None, timeout=30, context=None):
    explicit_context = context is not None
    auth_context = current_jira_auth_context(context)
    session_context = auth_context if explicit_context or is_db_auth_context(auth_context) else None
    session_data = current_jira_session_data(session_context)
    session_callbacks = current_oauth_session_callbacks(session_context)

    def request_get(url, **kwargs):
        return resilient_jira_get(
            url,
            session=HTTP_SESSION,
            breaker=JIRA_SEARCH_CIRCUIT_BREAKER,
            **kwargs,
        )

    return jira_get(
        current_auth_config(),
        auth_context,
        session_data,
        path,
        http_get=request_get,
        params=params,
        timeout=timeout,
        **session_callbacks,
    )


def current_jira_search(payload, *, context=None, timeout=30):
    return current_jira_get(
        '/rest/api/3/search/jql',
        params=_jira_client.build_jira_search_params(payload),
        timeout=timeout,
        context=context,
    )


def current_jira_request(method, path, *, json_body=None, params=None, timeout=30, context=None):
    explicit_context = context is not None
    auth_context = current_jira_auth_context(context)
    session_context = auth_context if explicit_context or is_db_auth_context(auth_context) else None
    session_data = current_jira_session_data(session_context)
    session_callbacks = current_oauth_session_callbacks(session_context)

    def request_fn(method_name, url, **kwargs):
        return HTTP_SESSION.request(method_name, url, **kwargs)

    kwargs = {'timeout': timeout}
    if json_body is not None:
        kwargs['json'] = json_body
    if params is not None:
        kwargs['params'] = params
    return jira_request(
        current_auth_config(),
        auth_context,
        session_data,
        method,
        path,
        request_fn,
        **kwargs,
        **session_callbacks,
    )


def _cache_policy_context(context=None):
    if context is not None:
        return context
    if has_request_context():
        return current_request_auth_context()
    return None


def auth_error_response(error, status=401):
    payload = {'error': error.code, 'message': str(error)}
    recovery_url = auth_recovery_url(error.code)
    if recovery_url:
        payload['recoveryUrl'] = recovery_url
    if error.code == 'auth_required':
        payload['loginUrl'] = '/login?reason=session_expired'
    return jsonify(payload), status


def admin_required_payload():
    return {
        'error': 'admin_required',
        'message': 'Admin access is required for this configuration change.',
        'recoveryUrl': '/auth/admin-required',
    }, 403


def auth_recovery_url(error_code):
    return {
        'account_disabled': '/auth/account-disabled',
        'auth_connection_revoked': '/auth/reconnect',
        'auth_connection_stale': '/auth/reconnect',
        'missing_oauth_scope': '/login?reason=missing_scope',
        'missing_project_access': '/auth/missing-project-access',
    }.get(error_code)


def validate_local_token_store_allowed():
    if JIRA_AUTH_MODE != AUTH_MODE_ATLASSIAN_OAUTH:
        return
    environment = APP_ENVIRONMENT_KEY.strip().lower()
    if environment not in {'local', 'dev'} or not OAUTH_LOCAL_TOKEN_STORE_ALLOWED:
        raise AuthError(
            'local_token_store_not_allowed',
            'Local OAuth token storage requires APP_ENVIRONMENT_KEY=local or dev and OAUTH_LOCAL_TOKEN_STORE_ALLOWED=true',
        )
    if OAUTH_TOKEN_STORE_TTL_SECONDS < OAUTH_TOKEN_STORE_MIN_TTL_SECONDS:
        raise AuthError(
            'oauth_token_store_ttl_too_low',
            f'OAUTH_TOKEN_STORE_TTL_SECONDS must be at least {OAUTH_TOKEN_STORE_MIN_TTL_SECONDS} seconds for Atlassian OAuth local testing',
        )


def validate_startup_auth_config():
    validate_auth_config(current_auth_config())
    validate_local_token_store_allowed()
    try:
        validate_config_storage_startup()
    except ConfigStorageError as error:
        raise AuthError('config_storage_invalid', str(error))


def _env_flag(name):
    return os.getenv(name, '').strip().lower() in {'1', 'true', 'yes'}


def default_bind_host():
    """Return the default Flask bind host for local execution."""
    return os.getenv('APP_BIND_HOST', '127.0.0.1').strip() or '127.0.0.1'


def validate_network_bind(host):
    """Validate and return the requested bind host, or raise AuthError with a stable code."""
    bind_host = (host or '').strip() or '127.0.0.1'
    if bind_host in {'127.0.0.1', 'localhost', '::1'}:
        return bind_host

    if not _env_flag('ALLOW_NETWORK_BIND'):
        raise AuthError(
            'network_bind_not_allowed',
            'Network bind requires ALLOW_NETWORK_BIND=true.',
        )

    if JIRA_AUTH_MODE == AUTH_MODE_BASIC:
        if not _env_flag('ALLOW_BASIC_AUTH_ON_NETWORK') or (os.getenv('APP_ENVIRONMENT_KEY', APP_ENVIRONMENT_KEY).strip().lower() != 'local'):
            raise AuthError(
                'basic_network_bind_not_allowed',
                'Basic auth network bind requires ALLOW_BASIC_AUTH_ON_NETWORK=true and APP_ENVIRONMENT_KEY=local.',
            )
        return bind_host

    if JIRA_AUTH_MODE == AUTH_MODE_ATLASSIAN_OAUTH:
        if not _env_flag('SESSION_COOKIE_SECURE'):
            raise AuthError(
                'secure_cookie_required',
                'OAuth network bind requires SESSION_COOKIE_SECURE=true.',
            )
        allowed_origins = [origin.strip() for origin in os.getenv('APP_ALLOWED_ORIGINS', '').split(',') if origin.strip()]
        if not allowed_origins or '*' in allowed_origins:
            raise AuthError(
                'allowed_origins_required',
                'OAuth network bind requires explicit APP_ALLOWED_ORIGINS without *.',
            )
        if not os.getenv('FLASK_SECRET_KEY', '').strip():
            raise AuthError(
                'flask_secret_required',
                'OAuth network bind requires FLASK_SECRET_KEY.',
            )
        if _env_flag('OAUTH_LOCAL_TOKEN_STORE_ALLOWED'):
            raise AuthError(
                'local_token_store_network_bind_not_allowed',
                'Local OAuth token storage cannot be used with network bind.',
            )
    return bind_host


def dev_diagnostics_allowed():
    return (
        os.getenv('APP_ENVIRONMENT_KEY', APP_ENVIRONMENT_KEY).strip().lower() in {'local', 'dev'}
        and _env_flag('ALLOW_DEV_DIAGNOSTIC_ENDPOINTS')
        and (request.remote_addr or '').strip().lower() in {'127.0.0.1', '::1', 'localhost'}
    )


@app.before_request
def redirect_unauthenticated_oauth_dashboard_entry():
    if JIRA_AUTH_MODE != AUTH_MODE_ATLASSIAN_OAUTH:
        return None
    if request.path not in OAUTH_DASHBOARD_ENTRY_PATHS:
        return None
    if database_storage_enabled() and db_oauth_browser_session_data():
        try:
            current_request_auth_context()
            return None
        except AuthError:
            pass
    data = oauth_session_data()
    if data.get('access_token') and data.get('cloudid'):
        return None
    return redirect('/login?reason=session_expired' if session.get('atlassian_oauth_session_id') else '/login')


def utc_now_iso(timespec=None):
    now = datetime.now(timezone.utc)
    value = now.isoformat(timespec=timespec) if timespec else now.isoformat()
    return value.replace('+00:00', 'Z')


RETRYABLE_JIRA_STATUS_CODES = _jira_client.RETRYABLE_JIRA_STATUS_CODES
SyntheticJiraResponse = _jira_client.SyntheticJiraResponse
JiraCircuitBreaker = _jira_client.JiraCircuitBreaker


JIRA_SEARCH_CIRCUIT_BREAKER = JiraCircuitBreaker(
    failure_threshold=JIRA_CIRCUIT_FAILURE_THRESHOLD,
    open_seconds=JIRA_CIRCUIT_OPEN_SECONDS
)


def _build_jira_unavailable_response(message, attempts=0, elapsed_seconds=0.0, upstream_status=None, circuit=None):
    return _jira_client._build_jira_unavailable_response(
        message,
        attempts=attempts,
        elapsed_seconds=elapsed_seconds,
        upstream_status=upstream_status,
        circuit=circuit,
        response_cls=SyntheticJiraResponse
    )


def resilient_jira_get(url, *, params=None, headers=None, timeout=30, session=None, breaker=None,
                       now_fn=None, sleep_fn=None, rand_fn=None,
                       max_attempts=None, max_elapsed_seconds=None,
                       base_delay_seconds=None, max_delay_seconds=None):
    """GET with bounded retries + circuit breaker for Jira upstream calls."""
    return _jira_client.resilient_jira_get(
        url,
        params=params,
        headers=headers,
        timeout=timeout,
        session=session or HTTP_SESSION,
        breaker=breaker or JIRA_SEARCH_CIRCUIT_BREAKER,
        now_fn=now_fn,
        sleep_fn=sleep_fn,
        rand_fn=rand_fn,
        max_attempts=JIRA_RETRY_MAX_ATTEMPTS if max_attempts is None else max_attempts,
        max_elapsed_seconds=JIRA_RETRY_MAX_ELAPSED_SECONDS if max_elapsed_seconds is None else max_elapsed_seconds,
        base_delay_seconds=JIRA_RETRY_BASE_DELAY_SECONDS if base_delay_seconds is None else base_delay_seconds,
        max_delay_seconds=JIRA_RETRY_MAX_DELAY_SECONDS if max_delay_seconds is None else max_delay_seconds,
        log_debug_fn=log_debug,
        log_info_fn=log_info,
        log_warning_fn=log_warning,
        log_error_fn=log_error,
        unavailable_response_fn=_build_jira_unavailable_response,
        retryable_status_codes=RETRYABLE_JIRA_STATUS_CODES
    )

def parse_args():
    """Parse CLI arguments to optionally override environment variables."""
    parser = argparse.ArgumentParser(description='Jira proxy server')
    parser.add_argument('--server_port', type=int, help='Port to run the server on (defaults to 5050 or SERVER_PORT env)')
    parser.add_argument('--jira_email', help='Jira account email (overrides JIRA_EMAIL env)')
    parser.add_argument('--jira_token', help='Jira API token (overrides JIRA_TOKEN env)')
    parser.add_argument('--jira_url', help='Base Jira URL, e.g. https://your-domain.atlassian.net (overrides JIRA_URL env)')
    parser.add_argument('--jira_query', help='JQL query to use for fetching issues (overrides JQL_QUERY env)')
    return parser.parse_args()


def add_clause_to_jql(jql: str, clause: str) -> str:
    """Append a clause to JQL before ORDER BY if present."""
    if not clause:
        return jql

    if 'ORDER BY' in jql:
        parts = jql.split('ORDER BY')
        base = parts[0].strip()
        order = parts[1].strip()
        if base:
            return f"{base} AND {clause} ORDER BY {order}"
        return f"{clause} ORDER BY {order}"
    if not jql.strip():
        return clause
    return f"{jql} AND {clause}"


def parse_iso_date(value):
    if not value:
        return None
    try:
        return datetime.strptime(value, '%Y-%m-%d').date()
    except Exception:
        return None


def resolve_sprint_label(sprint_value, cache_enabled=None):
    if sprint_value is None:
        return None
    sprint_str = str(sprint_value).strip()
    if not sprint_str:
        return None
    if sprint_str.isdigit():
        if cache_enabled is None:
            cache_enabled = jira_home_process_cache_enabled(_cache_policy_context())
        if cache_enabled:
            cache = load_sprints_cache() or {}
            for sprint in cache.get('sprints', []) or []:
                if str(sprint.get('id')) == sprint_str:
                    return sprint.get('name') or sprint_str
    return sprint_str


def quarter_dates_from_label(label):
    if not label:
        return None, None
    match = re.match(r'^(\d{4})Q([1-4])$', str(label).strip(), re.IGNORECASE)
    if not match:
        return None, None
    year = int(match.group(1))
    quarter = int(match.group(2))
    if quarter == 1:
        return date(year, 1, 1), date(year, 3, 31)
    if quarter == 2:
        return date(year, 4, 1), date(year, 6, 30)
    if quarter == 3:
        return date(year, 7, 1), date(year, 9, 30)
    return date(year, 10, 1), date(year, 12, 31)


def month_dates_from_iso(label):
    if not label:
        return None, None
    match = re.match(r'^(\d{4})-(\d{2})$', str(label).strip())
    if not match:
        return None, None
    year = int(match.group(1))
    month = int(match.group(2))
    if month < 1 or month > 12:
        return None, None
    if month == 12:
        next_month = date(year + 1, 1, 1)
    else:
        next_month = date(year, month + 1, 1)
    end = next_month - timedelta(days=1)
    return date(year, month, 1), end


def generate_period_labels(start_quarter, end_quarter, group_by):
    start_date, _ = quarter_dates_from_label(start_quarter)
    _, end_date = quarter_dates_from_label(end_quarter)
    if not start_date or not end_date:
        return []

    labels = []
    current = start_date
    if group_by == 'month':
        while current <= end_date:
            labels.append(f'{current.year}-{current.month:02d}')
            if current.month == 12:
                current = date(current.year + 1, 1, 1)
            else:
                current = date(current.year, current.month + 1, 1)
        return labels

    while current <= end_date:
        quarter = ((current.month - 1) // 3) + 1
        labels.append(f'{current.year}Q{quarter}')
        month = current.month + 3
        year = current.year
        if month > 12:
            month -= 12
            year += 1
        current = date(year, month, 1)
    return labels


def assign_to_period(day_value, group_by):
    if not isinstance(day_value, date):
        return None
    if group_by == 'month':
        return f'{day_value.year}-{day_value.month:02d}'
    quarter = ((day_value.month - 1) // 3) + 1
    return f'{day_value.year}Q{quarter}'


def compute_elapsed_period_index(created_day, terminal_day, group_by):
    if not isinstance(created_day, date) or not isinstance(terminal_day, date):
        return None
    if group_by == 'month':
        return max(0, (terminal_day.year - created_day.year) * 12 + (terminal_day.month - created_day.month))
    created_quarter = ((created_day.month - 1) // 3) + 1
    terminal_quarter = ((terminal_day.month - 1) // 3) + 1
    return max(0, (terminal_day.year - created_day.year) * 4 + (terminal_quarter - created_quarter))


def normalize_epic_status(status_name):
    normalized = str(status_name or '').strip().lower()
    if normalized == 'done':
        return 'Done'
    if normalized == 'killed':
        return 'Killed'
    if normalized == 'incomplete':
        return 'Incomplete'
    if normalized == 'postponed':
        return 'Postponed'
    return 'open'


def is_terminal_epic_status(status_name):
    return normalize_epic_status(status_name) in ('Done', 'Killed', 'Incomplete', 'Postponed')


def resolve_terminal_date_from_history(histories, terminal_status):
    target = normalize_epic_status(terminal_status).lower()
    if target == 'open':
        return None
    for history in sorted(histories or [], key=lambda item: item.get('created') or ''):
        event_dt = parse_jira_datetime(history.get('created'))
        if not event_dt:
            continue
        for item in history.get('items') or []:
            field_name = str(item.get('field') or '').strip().lower()
            if field_name != 'status':
                continue
            to_status = normalize_epic_status(item.get('toString')).lower()
            if to_status == target:
                return event_dt.date()
    return None


def _build_epic_cohort_cache_key(start_quarter, team_ids, projects, components=None):
    normalized_teams = ','.join(normalize_team_ids(team_ids or []))
    normalized_projects = ','.join(sorted(str(project or '').strip() for project in (projects or []) if str(project or '').strip()))
    normalized_components = ','.join(
        sorted(
            str(component or '').strip()
            for component in (components or [])
            if str(component or '').strip()
        )
    )
    return (
        f'{str(start_quarter or "").strip().upper()}'
        f'::{normalized_teams or "all"}'
        f'::{normalized_projects or "none"}'
        f'::{normalized_components or "no-components"}'
    )


TEAM_FIELD_CACHE = None
PARENT_NAME_FIELD_CACHE = None
EPIC_LINK_FIELD_CACHE = None
CAPACITY_FIELD_CACHE = None


def resolve_team_field_id(headers, context=None):
    """Resolve the Jira custom field ID for Team[Team]."""
    global TEAM_FIELD_CACHE
    cache_enabled = jira_home_process_cache_enabled(_cache_policy_context(context))
    with _cache_lock:
        if cache_enabled and TEAM_FIELD_CACHE:
            return TEAM_FIELD_CACHE
        # Check dashboard config first
        configured = get_team_field_id()
        if configured:
            if cache_enabled:
                TEAM_FIELD_CACHE = configured
                return TEAM_FIELD_CACHE
            return configured

        try:
            response = current_jira_get('/rest/api/3/field', timeout=20, context=context)
            if response.status_code != 200:
                return None

            fields = response.json() or []
            for field in fields:
                name = str(field.get('name', '')).strip().lower()
                if name == 'team[team]':
                    field_id = field.get('id')
                    if cache_enabled:
                        TEAM_FIELD_CACHE = field_id
                        return TEAM_FIELD_CACHE
                    return field_id
        except Exception:
            return None

        return None


def resolve_epic_link_field_id(headers, names_map=None, context=None):
    """Resolve the Jira custom field ID for Epic Link."""
    global EPIC_LINK_FIELD_CACHE
    cache_enabled = jira_home_process_cache_enabled(_cache_policy_context(context))
    with _cache_lock:
        if cache_enabled and EPIC_LINK_FIELD_CACHE:
            return EPIC_LINK_FIELD_CACHE

        if names_map:
            for field_id, field_name in (names_map or {}).items():
                if str(field_name).strip().lower() == 'epic link':
                    if cache_enabled:
                        EPIC_LINK_FIELD_CACHE = field_id
                        return EPIC_LINK_FIELD_CACHE
                    return field_id

        try:
            response = current_jira_get('/rest/api/3/field', timeout=20, context=context)
            if response.status_code != 200:
                return None

            fields = response.json() or []
            for field in fields:
                name = str(field.get('name', '')).strip().lower()
                if name == 'epic link':
                    field_id = field.get('id')
                    if cache_enabled:
                        EPIC_LINK_FIELD_CACHE = field_id
                        return EPIC_LINK_FIELD_CACHE
                    return field_id
        except Exception:
            return None

        return None


def resolve_capacity_field_id(headers, context=None):
    """Resolve the Jira custom field ID for Team capacity."""
    global CAPACITY_FIELD_CACHE
    cache_enabled = jira_home_process_cache_enabled(_cache_policy_context(context))
    with _cache_lock:
        if cache_enabled and CAPACITY_FIELD_CACHE:
            return CAPACITY_FIELD_CACHE
        cap = get_capacity_config()
        if cap['fieldId']:
            if cache_enabled:
                CAPACITY_FIELD_CACHE = cap['fieldId']
                return CAPACITY_FIELD_CACHE
            return cap['fieldId']

        field_name = cap['fieldName']
        if not field_name:
            return None

        try:
            response = current_jira_get('/rest/api/3/field', timeout=20, context=context)
            if response.status_code != 200:
                return None

            fields = response.json() or []
            target = field_name.strip().lower()
            for field in fields:
                name = str(field.get('name', '')).strip().lower()
                if name == target:
                    field_id = field.get('id')
                    if cache_enabled:
                        CAPACITY_FIELD_CACHE = field_id
                        return CAPACITY_FIELD_CACHE
                    return field_id
        except Exception:
            return None

        return None


def extract_team_name(value):
    """Extract a readable team name from Jira Team field values."""
    if value is None:
        return None
    if isinstance(value, list):
        names = [extract_team_name(item) for item in value]
        names = [name for name in names if name]
        return ', '.join(names) if names else None
    if isinstance(value, dict):
        for key in ('name', 'title', 'value', 'displayName', 'teamName'):
            if value.get(key):
                return value.get(key)
        return value.get('id')
    return str(value)


def extract_team_ids(value):
    """Extract Team[Team] ids from Jira Team field values."""
    if value is None:
        return []
    if isinstance(value, list):
        ids = []
        for item in value:
            ids.extend(extract_team_ids(item))
        return [team_id for team_id in ids if team_id]
    if isinstance(value, dict):
        team_id = value.get('id') or value.get('teamId')
        return [team_id] if team_id else []
    return []


def normalize_team_value(value):
    """Normalize Team field values to human-readable names."""
    if isinstance(value, list):
        return [normalize_team_value(item) for item in value if item]
    if isinstance(value, dict):
        return value.get('name') or value.get('value') or value.get('displayName') or value.get('teamName') or value.get('title') or value.get('id')
    return value


def normalize_capacity_team_name(team_name):
    """Strip prefixes to match capacity team labels."""
    return _capacity_service.normalize_capacity_team_name(team_name)


def build_team_value(raw_team):
    """Build a consistent team payload with id/name when possible."""
    if isinstance(raw_team, dict):
        return {
            'id': raw_team.get('id') or raw_team.get('teamId'),
            'name': raw_team.get('name') or raw_team.get('title') or raw_team.get('value') or raw_team.get('displayName')
        }
    return raw_team


def jira_search_request(payload, *, context=None):
    """Call Jira search endpoint through the active request auth boundary."""
    if context is None:
        return current_jira_search(payload)
    return current_jira_search(payload, context=context)


def fetch_teams_from_jira_api():
    """Fetch teams directly from Jira Teams REST API (team registry, not issues).

    This catches teams that may not have any issues in PRODUCT/TECH projects.
    Uses /rest/teams/1.0/teams/find which is the same API the Jira team picker uses.
    """
    teams = {}
    try:
        start_at = 0
        max_results = 200
        while True:
            response = current_jira_get(
                '/rest/teams/1.0/teams/find',
                params={'query': '', 'maxResults': max_results, 'startAt': start_at},
                timeout=30
            )
            if response.status_code != 200:
                log_warning(f'[teams-api] Teams API returned {response.status_code}, falling back to issue scan only')
                break
            data = response.json()
            team_list = data if isinstance(data, list) else data.get('teams', [])
            if not team_list:
                break
            for t in team_list:
                tid = t.get('id') or t.get('teamId')
                tname = t.get('title') or t.get('name') or t.get('displayName')
                if tid and tname:
                    teams[str(tid)] = {'id': str(tid), 'name': tname}
            # If we got fewer than max, we've fetched everything
            if len(team_list) < max_results:
                break
            start_at += max_results
        if teams:
            log_info(f'[teams-api] Fetched {len(teams)} teams from Jira Teams API')
    except Exception as e:
        log_warning(f'[teams-api] Teams API unavailable ({e}), using issue scan only')
    return teams


def build_capacity_jql(sprint_name, team_names=None):
    return _capacity_service.build_capacity_jql(
        sprint_name,
        team_names,
        capacity_project=get_effective_capacity_project(),
    )


def fetch_capacity_for_sprint(sprint_name, headers, debug=False, team_names=None):
    return _capacity_service.fetch_capacity_for_sprint(
        sprint_name,
        headers,
        debug=debug,
        team_names=team_names,
        capacity_project=get_effective_capacity_project(),
        resolve_capacity_field_id=resolve_capacity_field_id,
        search_request=jira_search_request,
        build_capacity_jql_fn=build_capacity_jql,
        normalize_capacity_team_name_fn=normalize_capacity_team_name,
    )


def fetch_watchers_count(issue_key):
    """Fetch watchers count for an issue (fallback if watches field is missing)."""
    return _capacity_service.fetch_watchers_count(
        issue_key,
        current_jira_get=current_jira_get,
        log_warning_fn=log_warning,
        logger=logger,
    )


def fetch_capacity_team_sizes(sprint_name, headers, team_names=None):
    """Fetch team sizes from Jira capacity issues (watchers count)."""
    return _capacity_service.fetch_capacity_team_sizes(
        sprint_name,
        headers,
        team_names=team_names,
        capacity_project=get_effective_capacity_project(),
        search_request=jira_search_request,
        fetch_watchers_count=fetch_watchers_count,
        build_capacity_jql_fn=build_capacity_jql,
        normalize_capacity_team_name_fn=normalize_capacity_team_name,
        log_warning_fn=log_warning,
        log_debug_fn=log_debug,
    )


# Cache helper functions
def load_sprints_cache():
    """Load sprints from cache file"""
    return _sprints_service.load_sprints_cache(
        SPRINTS_CACHE_FILE,
        log_warning_fn=log_warning,
    )


def save_sprints_cache(sprints):
    """Save sprints to cache file"""
    return _sprints_service.save_sprints_cache(
        sprints,
        cache_file=SPRINTS_CACHE_FILE,
        board_id=get_effective_board_id(),
        now_fn=datetime.now,
        log_info_fn=log_info,
        log_warning_fn=log_warning,
    )


def is_cache_valid():
    """Check if cache exists, is not expired, and matches the current board config"""
    cache_data = load_sprints_cache()
    return _sprints_service.is_sprints_cache_valid(
        cache_data,
        current_board_id=get_effective_board_id(),
        cache_expiry_hours=CACHE_EXPIRY_HOURS,
        now_fn=datetime.now,
        log_info_fn=log_info,
        log_debug_fn=log_debug,
        log_warning_fn=log_warning,
    )


def load_stats_cache():
    """Load stats cache from disk."""
    return _stats_cache_service.load_stats_cache(STATS_CACHE_FILE, log_warning_fn=log_warning)


def save_stats_cache(cache_data):
    """Persist stats cache to disk."""
    return _stats_cache_service.save_stats_cache(
        cache_data,
        cache_file=STATS_CACHE_FILE,
        log_warning_fn=log_warning,
    )


def build_stats_cache_key(sprint_name, base_jql, team_ids, group_id=None):
    return _stats_cache_service.build_stats_cache_key(
        sprint_name,
        base_jql,
        team_ids,
        order_by=STATS_JQL_ORDER_BY,
        group_id=group_id,
    )


def strip_sprint_clause(jql):
    """Remove Sprint clause if the base query already includes it."""
    if not jql:
        return jql
    jql = re.sub(r'\s+AND\s+Sprint\s+in\s+\([^)]+\)', '', jql, flags=re.IGNORECASE)
    return re.sub(r'\s+AND\s+Sprint\s*=\s*[^ ]+', '', jql, flags=re.IGNORECASE)


def extract_team_ids_from_jql(jql):
    """Extract Team[Team] ids from a JQL clause if present."""
    if not jql:
        return []
    match_in = re.search(r'"Team\[Team\]"\s+in\s*\(([^)]+)\)', jql, flags=re.IGNORECASE)
    if match_in:
        raw = match_in.group(1)
        parts = [p.strip() for p in raw.split(',')]
        ids = []
        for part in parts:
            part = part.strip().strip('"').strip("'")
            if part:
                ids.append(part)
        return ids
    match_eq = re.search(r'"Team\[Team\]"\s*=\s*("?)([^")\s]+)\1', jql, flags=re.IGNORECASE)
    if match_eq:
        return [match_eq.group(2)]
    return []


def remove_team_filter_from_jql(jql):
    """Remove Team[Team] filter from JQL query to fetch all teams."""
    if not jql:
        return jql
    # Remove "Team[Team]" in (...) pattern
    jql = re.sub(r'\s+AND\s+"Team\[Team\]"\s+in\s*\([^)]+\)', '', jql, flags=re.IGNORECASE)
    jql = re.sub(r'"Team\[Team\]"\s+in\s*\([^)]+\)\s+AND\s+', '', jql, flags=re.IGNORECASE)
    # Remove "Team[Team]" = "..." pattern
    jql = re.sub(r'\s+AND\s+"Team\[Team\]"\s*=\s*[^\s]+', '', jql, flags=re.IGNORECASE)
    jql = re.sub(r'"Team\[Team\]"\s*=\s*[^\s]+\s+AND\s+', '', jql, flags=re.IGNORECASE)
    # Remove standalone Team[Team] filter (if it's the only filter before ORDER BY)
    jql = re.sub(r'"Team\[Team\]"\s+in\s*\([^)]+\)\s*(?=ORDER BY|$)', '', jql, flags=re.IGNORECASE)
    jql = re.sub(r'"Team\[Team\]"\s*=\s*[^\s]+\s*(?=ORDER BY|$)', '', jql, flags=re.IGNORECASE)
    return jql.strip()


def remove_project_filter_from_jql(jql):
    """Remove project IN (...) and project = "..." filters from JQL query."""
    if not jql:
        return jql
    # Remove project IN (...) pattern
    jql = re.sub(r'\s+AND\s+project\s+IN\s*\([^)]+\)', '', jql, flags=re.IGNORECASE)
    jql = re.sub(r'project\s+IN\s*\([^)]+\)\s+AND\s+', '', jql, flags=re.IGNORECASE)
    # Remove project = "..." pattern
    jql = re.sub(r'\s+AND\s+project\s*=\s*"[^"]+"', '', jql, flags=re.IGNORECASE)
    jql = re.sub(r'project\s*=\s*"[^"]+"\s+AND\s+', '', jql, flags=re.IGNORECASE)
    # Remove standalone project filter (only filter before ORDER BY)
    jql = re.sub(r'project\s+IN\s*\([^)]+\)\s*(?=ORDER BY|$)', '', jql, flags=re.IGNORECASE)
    jql = re.sub(r'project\s*=\s*"[^"]+"\s*(?=ORDER BY|$)', '', jql, flags=re.IGNORECASE)
    return jql.strip()


def get_stats_team_ids():
    """Resolve stats team IDs from env or JQL configuration."""
    if STATS_TEAM_IDS:
        return STATS_TEAM_IDS
    base_jql = STATS_JQL_BASE or build_base_jql()
    return extract_team_ids_from_jql(base_jql)


def normalize_team_ids(team_ids):
    seen = set()
    normalized = []
    for team_id in team_ids or []:
        value = str(team_id or '').strip()
        if not value or value in seen:
            continue
        seen.add(value)
        normalized.append(value)
    return normalized


def normalize_epic_keys(epic_keys):
    seen = set()
    normalized = []
    for epic_key in epic_keys or []:
        value = str(epic_key or '').strip().upper()
        if not value or value in seen:
            continue
        seen.add(value)
        normalized.append(value)
    return normalized


def resolve_groups_config_path():
    return _config_store.resolve_groups_config_path(GROUPS_CONFIG_PATH)


def load_groups_config_file(path):
    return _config_store.load_groups_config_file(path, log_warning_fn=log_warning)


def resolve_dashboard_config_path():
    return _config_store.resolve_dashboard_config_path(DASHBOARD_CONFIG_PATH)


def _json_config_repository():
    return build_json_config_repository(
        dashboard_path=resolve_dashboard_config_path(),
        groups_path=resolve_groups_config_path(),
        load_groups_config_file_fn=load_groups_config_file,
        log_warning_fn=log_warning,
    )


def _load_dashboard_config_json():
    return _json_config_repository().load_dashboard_config()


def _normalize_dashboard_config_source(source):
    value = str(source or 'auto').strip().lower()
    if value in {'auto', 'jsonfile', 'db'}:
        return value
    raise ConfigStorageError('dashboard config source must be auto, jsonfile, or db')


def _current_dashboard_config_context_or_error():
    if has_request_context():
        return current_request_auth_context()
    raise ConfigStorageError(
        'CONFIG_STORAGE_BACKEND=db requires request context for dashboard config; '
        'pass source="jsonfile" for explicit legacy JSON access'
    )


def load_dashboard_config(*, source='auto'):
    """Load the unified dashboard config."""
    source = _normalize_dashboard_config_source(source)
    if source == 'jsonfile':
        return _load_dashboard_config_json()
    if source == 'db' or config_storage_db_enabled():
        context = _current_dashboard_config_context_or_error()
        return build_db_config_repository().load_dashboard_config(
            context,
            fallback_loader=_load_dashboard_config_json,
        )
    return _load_dashboard_config_json()


def _save_dashboard_config_json(config):
    return _config_store.save_dashboard_config(config, resolve_dashboard_config_path())


def save_dashboard_config(config, *, source='auto'):
    """Write the unified dashboard config."""
    source = _normalize_dashboard_config_source(source)
    if source == 'jsonfile':
        return _save_dashboard_config_json(config)
    if source == 'db' or config_storage_db_enabled():
        context = _current_dashboard_config_context_or_error()
        return build_db_config_repository().save_dashboard_config(context, config)
    return _save_dashboard_config_json(config)


def resolve_team_catalog_path():
    return _config_store.resolve_team_catalog_path(TEAM_CATALOG_PATH)


def load_team_catalog():
    return _config_store.load_team_catalog(
        resolve_team_catalog_path(),
        normalize_team_catalog_fn=normalize_team_catalog,
        normalize_team_catalog_meta_fn=normalize_team_catalog_meta,
        log_warning_fn=log_warning
    )


def save_team_catalog_file(catalog_data):
    return _config_store.save_team_catalog_file(
        catalog_data,
        resolve_team_catalog_path(),
        normalize_team_catalog_fn=normalize_team_catalog,
        normalize_team_catalog_meta_fn=normalize_team_catalog_meta
    )


def migrate_team_catalog_from_config():
    """One-time migration: extract teamCatalog from dashboard-config.json into team-catalog.json."""
    catalog_path = resolve_team_catalog_path()
    if os.path.exists(catalog_path):
        return  # Already migrated or manually created
    dashboard_config = load_dashboard_config(source='jsonfile')
    if not dashboard_config:
        return
    team_groups = dashboard_config.get('teamGroups')
    if not isinstance(team_groups, dict):
        return
    raw_catalog = team_groups.get('teamCatalog') or {}
    raw_meta = team_groups.get('teamCatalogMeta') or {}
    catalog = normalize_team_catalog(raw_catalog)
    if not catalog:
        return  # Nothing to migrate
    save_team_catalog_file({'catalog': catalog, 'meta': raw_meta})
    log_info('Migrated teamCatalog from dashboard-config.json to team-catalog.json')


def resolve_scenario_overrides_path():
    return SCENARIO_OVERRIDES_PATH or './scenario-overrides.json'


def load_scenario_overrides():
    """Load scenario overrides from disk."""
    path = resolve_scenario_overrides_path()
    if os.path.exists(path):
        try:
            with open(path, 'r') as handle:
                data = json.load(handle)
                if isinstance(data, dict) and 'version' in data:
                    return data
        except Exception as e:
            log_warning(f'Failed to read scenario overrides: {e}')
    return {'version': 1, 'scenarios': {}}


def save_scenario_overrides(data):
    """Write scenario overrides to disk."""
    path = resolve_scenario_overrides_path()
    directory = os.path.dirname(path)
    if directory:
        os.makedirs(directory, exist_ok=True)
    with _cache_lock:
        with open(path, 'w') as handle:
            json.dump(data, handle, indent=2)


def get_selected_projects():
    """Return the list of selected project keys from dashboard config."""
    config = load_dashboard_config()
    if not config:
        return []
    selected = config.get('projects', {}).get('selected', [])
    # Normalise: support both old string format and new {key, type} format
    keys = []
    for item in selected:
        if isinstance(item, str):
            keys.append(item)
        elif isinstance(item, dict) and item.get('key'):
            keys.append(item['key'])
    return keys


def build_base_jql():
    """Return base JQL: env JQL_QUERY if set, else derive from dashboard-config projects."""
    if JQL_QUERY:
        return JQL_QUERY
    projects = get_selected_projects()
    if not projects:
        return ''
    quoted = ', '.join(f'"{p}"' for p in projects)
    return f'project in ({quoted}) ORDER BY created DESC'


def get_selected_projects_typed():
    """Return the list of selected projects with their product/tech type."""
    config = load_dashboard_config()
    if not config:
        return []
    selected = config.get('projects', {}).get('selected', [])
    result = []
    for item in selected:
        if isinstance(item, str):
            result.append({'key': item, 'type': 'product'})
        elif isinstance(item, dict) and item.get('key'):
            result.append({'key': item['key'], 'type': item.get('type', 'product')})
    return result


def get_capacity_config():
    """Return capacity config from dashboard config, falling back to env vars."""
    config = load_dashboard_config()
    if config and 'capacity' in config:
        cap = config['capacity']
        return {
            'project': cap.get('project', ''),
            'fieldId': cap.get('fieldId', ''),
            'fieldName': cap.get('fieldName', ''),
        }
    return {
        'project': CAPACITY_PROJECT,
        'fieldId': CAPACITY_FIELD_ID,
        'fieldName': CAPACITY_FIELD_NAME,
    }


def get_board_config():
    """Return dashboard Jira board config, falling back to env var."""
    config = load_dashboard_config()
    if config and 'board' in config:
        board = config.get('board') or {}
        return {
            'boardId': str(board.get('boardId', '') or '').strip(),
            'boardName': str(board.get('boardName', '') or '').strip(),
            'source': 'config'
        }
    return {
        'boardId': str(JIRA_BOARD_ID or '').strip(),
        'boardName': '',
        'source': 'env' if JIRA_BOARD_ID else 'default'
    }


def get_effective_board_id():
    return get_board_config().get('boardId', '').strip()


def build_epm_home_projects_cache_key(epm_scope):
    return epm_projects.build_epm_home_projects_cache_key(epm_scope)


def clear_epm_project_cache():
    with _epm_cache_lock:
        EPM_PROJECTS_CACHE.clear()


def clear_epm_rollup_caches():
    with _epm_cache_lock:
        EPM_ISSUES_CACHE.clear()
        EPM_ROLLUP_CACHE.clear()


def clear_epm_caches():
    with _epm_cache_lock:
        EPM_PROJECTS_CACHE.clear()
        EPM_ISSUES_CACHE.clear()
        EPM_ROLLUP_CACHE.clear()


def invalidate_stats_cache():
    return _stats_cache_service.invalidate_stats_cache(STATS_CACHE_FILE, log_warning_fn=log_warning)


def clear_auth_sensitive_caches(reason='auth_context_change'):
    global TEAM_FIELD_CACHE, PARENT_NAME_FIELD_CACHE, EPIC_LINK_FIELD_CACHE, CAPACITY_FIELD_CACHE
    with _cache_lock:
        TASKS_CACHE.clear()
        MISSING_INFO_CACHE.clear()
        DEPENDENCIES_CACHE.clear()
        EPIC_COHORT_CACHE.clear()
        EXCLUDED_CAPACITY_STATS_SOURCE_CACHE.clear()
        EXCLUDED_CAPACITY_EPIC_SUMMARY_CACHE.clear()
        SCENARIO_CACHE['generatedAt'] = None
        SCENARIO_CACHE['data'] = None
        if 'PROJECTS_CACHE' in globals():
            PROJECTS_CACHE['data'] = None
            PROJECTS_CACHE['timestamp'] = 0
        if 'COMPONENTS_CACHE' in globals():
            COMPONENTS_CACHE['data'] = None
            COMPONENTS_CACHE['timestamp'] = 0
        if 'EPICS_SEARCH_CACHE' in globals():
            EPICS_SEARCH_CACHE.clear()
        if 'LABELS_CACHE' in globals():
            LABELS_CACHE['data'] = None
            LABELS_CACHE['timestamp'] = 0
        if 'ISSUE_TYPES_CACHE' in globals():
            ISSUE_TYPES_CACHE['data'] = None
            ISSUE_TYPES_CACHE['timestamp'] = 0
        TEAM_FIELD_CACHE = None
        PARENT_NAME_FIELD_CACHE = None
        EPIC_LINK_FIELD_CACHE = None
        CAPACITY_FIELD_CACHE = None
    clear_epm_caches()
    invalidate_sprints_cache()
    invalidate_stats_cache()
    log_info(f'Cleared auth-sensitive caches reason={reason}')


register_service_integration_cache_invalidator(clear_auth_sensitive_caches)


def build_epm_projects_dependencies(context=None):
    auth_context = context if context is not None else (current_request_auth_context() if has_request_context() else None)
    fetch_context = auth_context if auth_context is not None and not jira_home_process_cache_enabled(auth_context) else None
    return epm_projects.EpmProjectsDependencies(
        fetch_epm_home_projects=(
            lambda epm_scope: fetch_epm_home_projects(epm_scope, context=fetch_context)
            if fetch_context is not None
            else fetch_epm_home_projects(epm_scope)
        ),
        merge_epm_linkage=merge_epm_linkage,
        normalize_epm_config=normalize_epm_config,
        utc_now_iso=utc_now_iso,
        cache=EPM_PROJECTS_CACHE,
        cache_lock=_epm_cache_lock,
        cache_ttl_seconds=EPM_PROJECTS_CACHE_TTL_SECONDS,
        home_project_limit=epm_home.HOME_MAX_PROJECTS_PER_GOAL,
        get_epm_config=get_epm_config,
        abort_not_found=abort,
        context=auth_context,
    )


def build_epm_home_projects_state(epm_scope, force_refresh=False):
    return epm_projects.build_epm_home_projects_state(
        epm_scope,
        build_epm_projects_dependencies(),
        force_refresh=force_refresh,
    )


def get_cached_epm_home_projects(epm_scope, force_refresh=False):
    return epm_projects.get_cached_epm_home_projects(
        epm_scope,
        build_epm_projects_dependencies(),
        force_refresh=force_refresh,
    )


def build_jira_headers():
    if JIRA_AUTH_MODE != AUTH_MODE_BASIC:
        raise AuthError(
            'route_not_oauth_ready',
            'This Jira route has not been migrated to Atlassian OAuth yet',
        )
    credentials = base64.b64encode(f"{JIRA_EMAIL or ''}:{JIRA_TOKEN or ''}".encode()).decode()
    return {
        'Authorization': f'Basic {credentials}',
        'Accept': 'application/json',
        'Content-Type': 'application/json',
    }


def build_epm_fields_list():
    fields_list = ['summary', 'status', 'assignee', 'priority', 'issuetype', 'parent', 'labels', 'created', 'updated']
    story_points_field = get_story_points_field_id()
    if story_points_field and story_points_field not in fields_list:
        fields_list.append(story_points_field)
    return fields_list


def build_epm_rollup_fields_list(epic_link_field_id=None, team_field_id=None):
    fields_list = build_epm_fields_list()
    sprint_field_id = get_sprint_field_id()
    if sprint_field_id and sprint_field_id not in fields_list:
        fields_list.append(sprint_field_id)
    if epic_link_field_id and epic_link_field_id not in fields_list:
        fields_list.append(epic_link_field_id)
    if team_field_id and team_field_id not in fields_list:
        fields_list.append(team_field_id)
    return fields_list


def shape_epm_issue_payload(issues, team_field_id=None, include_card_fields=False):
    story_points_field = get_story_points_field_id() if include_card_fields else None
    team_field_id = team_field_id or (get_team_field_id() if include_card_fields else None)
    slim_issues = []
    epic_details = {}
    for issue in issues or []:
        fields = issue.get('fields') or {}
        parent = fields.get('parent') or {}
        parent_key = parent.get('key') or ''
        priority = fields.get('priority') or {}
        raw_team = fields.get(team_field_id) if team_field_id else None
        team_name = extract_team_name(raw_team) or ''
        team_ids = extract_team_ids(raw_team)
        if parent_key and parent_key not in epic_details:
            parent_fields = parent.get('fields') or {}
            epic_details[parent_key] = {
                'key': parent_key,
                'summary': parent_fields.get('summary') or '',
                'issueType': (parent_fields.get('issuetype') or {}).get('name') or '',
            }
        slim_issue = {
            'key': issue.get('key'),
            'summary': fields.get('summary') or '',
            'status': (fields.get('status') or {}).get('name') or '',
            'assignee': (fields.get('assignee') or {}).get('displayName') or '',
            'issueType': (fields.get('issuetype') or {}).get('name') or '',
            'parentKey': parent_key,
            'labels': list(fields.get('labels') or []),
        }
        if include_card_fields:
            slim_issue.update({
                'id': issue.get('id'),
                'priority': priority.get('name') if priority else '',
                'storyPoints': fields.get(story_points_field),
                'updated': fields.get('updated'),
                'teamName': team_name,
                'teamId': team_ids[0] if team_ids else '',
            })
        slim_issues.append(slim_issue)
    return slim_issues, epic_details


def shape_epm_rollup_issue_payload(issues, epic_link_field_id=None, team_field_id=None):
    slim_issues, epic_details = shape_epm_issue_payload(issues, team_field_id=team_field_id, include_card_fields=True)
    sprint_field_id = get_sprint_field_id()
    for raw_issue, slim_issue in zip(issues or [], slim_issues):
        fields = raw_issue.get('fields') or {}
        if not slim_issue.get('parentKey') and epic_link_field_id:
            legacy_parent_key = normalize_epm_text(fields.get(epic_link_field_id))
            if legacy_parent_key:
                slim_issue['parentKey'] = legacy_parent_key
                epic_details.setdefault(legacy_parent_key, {
                    'key': legacy_parent_key,
                    'summary': '',
                    'issueType': '',
                })
        slim_issue['sprint'] = normalize_epm_sprint_field(fields.get(sprint_field_id))
    return slim_issues, epic_details


dedupe_issues_by_key = epm_payload.dedupe_issues_by_key
validate_epm_tab_sprint = epm_payload.validate_epm_tab_sprint
normalize_epm_issue_type_sets = epm_payload.normalize_epm_issue_type_sets
build_empty_epm_rollup_payload = epm_payload.build_empty_epm_rollup_payload
build_epm_rollup_hierarchy = epm_payload.build_epm_rollup_hierarchy


def fetch_epm_rollup_query(jql, query_name, headers, fields_list, truncated_queries, context=None):
    raw_issues = fetch_issues_by_jql(
        jql,
        fields_list,
        max_results=EPM_ROLLUP_QUERY_MAX_RESULTS + 1,
        context=context,
    )
    if len(raw_issues) > EPM_ROLLUP_QUERY_MAX_RESULTS:
        truncated_queries.append(query_name)
        return raw_issues[:EPM_ROLLUP_QUERY_MAX_RESULTS]
    return raw_issues


def build_epm_rollup_dependencies(sub_goal_keys=None):
    auth_context = current_request_auth_context() if has_request_context() else None
    return EpmRollupDependencies(
        find_epm_project_or_404=lambda project_id: find_epm_project_or_404(
            project_id,
            sub_goal_keys=sub_goal_keys,
            context=auth_context,
        ),
        normalize_epm_text=normalize_epm_text,
        validate_epm_tab_sprint=validate_epm_tab_sprint,
        build_empty_epm_rollup_payload=build_empty_epm_rollup_payload,
        build_base_jql=build_base_jql,
        add_clause_to_jql=add_clause_to_jql,
        build_jira_headers=(
            (lambda: {}) if auth_context is not None and auth_context.auth_mode != AUTH_MODE_BASIC
            else build_jira_headers
        ),
        resolve_epic_link_field_id=(
            lambda headers: resolve_epic_link_field_id(headers, context=auth_context)
            if auth_context is not None
            else resolve_epic_link_field_id(headers)
        ),
        resolve_team_field_id=(
            lambda headers: resolve_team_field_id(headers, context=auth_context)
            if auth_context is not None
            else resolve_team_field_id(headers)
        ),
        build_epm_rollup_fields_list=build_epm_rollup_fields_list,
        get_epm_config=get_epm_config,
        normalize_epm_issue_type_sets=normalize_epm_issue_type_sets,
        fetch_epm_rollup_query=(
            lambda jql, query_name, headers, fields_list, truncated_queries:
            fetch_epm_rollup_query(jql, query_name, headers, fields_list, truncated_queries, context=auth_context)
        ),
        shape_epm_rollup_issue_payload=shape_epm_rollup_issue_payload,
        dedupe_issues_by_key=dedupe_issues_by_key,
        build_epm_rollup_hierarchy=build_epm_rollup_hierarchy,
        cache=EPM_ROLLUP_CACHE,
        cache_lock=_epm_cache_lock,
        cache_ttl_seconds=EPM_ROLLUP_CACHE_TTL_SECONDS,
        context=auth_context,
    )


def normalize_epm_label_prefix_mask(label_prefix):
    return epm_projects.normalize_epm_label_prefix_mask(label_prefix)


def filter_epm_home_tag_matches(home_project, label_prefix):
    return epm_projects.filter_epm_home_tag_matches(home_project, label_prefix)


def resolve_epm_project_label(home_project, config_row, label_prefix):
    return epm_projects.resolve_epm_project_label(home_project, config_row, label_prefix)


def build_epm_project_payload(home_project, config_row, label_prefix=None):
    return epm_projects.build_epm_project_payload(
        home_project,
        config_row,
        label_prefix,
        merge_epm_linkage=merge_epm_linkage,
    )


def build_custom_project_payload(row):
    return epm_projects.build_custom_project_payload(row)


def find_epm_config_row(projects, project_id):
    return epm_projects.find_epm_config_row(projects, project_id)


def build_epm_projects_payload(epm_config, force_refresh=False, tab=None, sub_goal_keys=None, context=None):
    return epm_projects.build_epm_projects_payload(
        epm_config,
        build_epm_projects_dependencies(context=context),
        force_refresh=force_refresh,
        tab=tab,
        sub_goal_keys=sub_goal_keys,
    )


def filter_epm_projects_for_tab(projects, tab, now=None):
    return epm_projects.filter_epm_projects_for_tab(projects, tab, now=now)


def collect_epm_rollup_issue_keys(rollup):
    return epm_aggregate.collect_epm_rollup_issue_keys(rollup, normalize_epm_text)


def build_all_epm_projects_rollup(tab, sprint, sub_goal_keys=None):
    return epm_aggregate.build_all_epm_projects_rollup(
        tab,
        sprint,
        epm_aggregate.EpmAggregateDependencies(
            normalize_epm_text=normalize_epm_text,
            validate_epm_tab_sprint=validate_epm_tab_sprint,
            get_epm_config=get_epm_config,
            build_epm_projects_payload=build_epm_projects_payload,
            filter_epm_projects_for_tab=filter_epm_projects_for_tab,
            build_epm_rollup_dependencies=build_epm_rollup_dependencies,
            get_epm_project_payload_identity=get_epm_project_payload_identity,
            build_empty_epm_rollup_payload=build_empty_epm_rollup_payload,
            build_per_project_rollup=build_per_project_rollup,
            logger=logger,
        ),
        sub_goal_keys=sub_goal_keys,
    )


def find_epm_project_or_404(project_id, sub_goal_keys=None, context=None):
    requested_sub_goal_keys = epm_projects.normalize_epm_sub_goal_keys(sub_goal_keys)
    if requested_sub_goal_keys:
        epm_config = get_epm_config()
        projects_payload = build_epm_projects_payload(
            epm_config,
            sub_goal_keys=requested_sub_goal_keys,
            context=context,
        )
        for project in projects_payload.get('projects') or []:
            candidates = [
                normalize_epm_text(project.get('id')),
                normalize_epm_text(project.get('homeProjectId')),
            ]
            if normalize_epm_text(project_id) in candidates:
                return project
        abort(404)
    return epm_projects.find_epm_project_or_404(project_id, build_epm_projects_dependencies(context=context))


def get_epm_project_payload_identity(project):
    return epm_projects.get_epm_project_payload_identity(project)


def normalize_epm_text(value):
    return str(value or '').strip()


def normalize_epm_upper_text(value):
    return normalize_epm_text(value).upper()


def normalize_epm_sub_goal_keys(values):
    return epm_projects.normalize_epm_sub_goal_keys(values)


def parse_epm_sub_goal_keys_param(value):
    if isinstance(value, list):
        raw_values = value
    else:
        raw_values = str(value or '').split(',')
    return normalize_epm_sub_goal_keys(raw_values)


DEFAULT_EPM_LABEL_PREFIX = epm_config.DEFAULT_EPM_LABEL_PREFIX
DEFAULT_EPM_ISSUE_TYPES = epm_config.DEFAULT_EPM_ISSUE_TYPES
normalize_epm_scope = epm_config.normalize_epm_scope
normalize_epm_issue_types = epm_config.normalize_epm_issue_types
is_epm_v2_config = epm_config.is_epm_v2_config
normalize_epm_project_row = epm_config.normalize_epm_project_row
normalize_epm_project_output_key = epm_config.normalize_epm_project_output_key
normalize_epm_config = epm_config.normalize_epm_config


def get_epm_config():
    config = load_dashboard_config() or {}
    return normalize_epm_config(config.get('epm') or {})


def fetch_home_site_cloud_id(context=None):
    auth_context = _cache_policy_context(context)
    if auth_context is not None:
        return epm_home.fetch_home_site_cloud_id(context=auth_context)
    return epm_home.fetch_home_site_cloud_id()


def _build_epm_home_graphql_client(context=None):
    auth_context = _cache_policy_context(context)
    credential = epm_home._read_metadata_credential(auth_context)
    return (
        epm_home.build_home_graphql_client(credential)
        if credential is not None
        else epm_home.build_home_graphql_client()
    )


def fetch_epm_goal_catalog(context=None):
    auth_context = _cache_policy_context(context)
    client = _build_epm_home_graphql_client(context=auth_context)
    cloud_id = fetch_home_site_cloud_id(context=auth_context)
    container_id = epm_home._container_id_from_cloud(cloud_id)
    return client.execute_paginated(
        epm_home.QUERY_GOALS_SEARCH,
        {'containerId': container_id, 'first': epm_home.HOME_PAGE_SIZE},
        'goals_search',
    )


def fetch_epm_sub_goals(root_goal_key, context=None):
    auth_context = _cache_policy_context(context)
    client = _build_epm_home_graphql_client(context=auth_context)
    cloud_id = fetch_home_site_cloud_id(context=auth_context)
    container_id = epm_home._container_id_from_cloud(cloud_id)
    return epm_home.fetch_sub_goals_for_root_key(client, root_goal_key, container_id, context=auth_context)


# --- Custom field config getters ---
SPRINT_FIELD_DEFAULT = 'customfield_10101'
STORY_POINTS_FIELD_DEFAULT = 'customfield_10004'
PARENT_NAME_FIELD_DEFAULT = 'customfield_10011'
TEAM_FIELD_DEFAULT = 'customfield_30101'


def get_sprint_field_config():
    config = load_dashboard_config()
    if config and 'sprintField' in config:
        sf = config['sprintField']
        return {'fieldId': sf.get('fieldId', ''), 'fieldName': sf.get('fieldName', '')}
    return {'fieldId': SPRINT_FIELD_DEFAULT, 'fieldName': ''}


def get_sprint_field_id():
    return get_sprint_field_config()['fieldId'] or SPRINT_FIELD_DEFAULT


def get_story_points_field_config():
    config = load_dashboard_config()
    if config and 'storyPointsField' in config:
        sp = config['storyPointsField']
        return {'fieldId': sp.get('fieldId', ''), 'fieldName': sp.get('fieldName', '')}
    return {'fieldId': STORY_POINTS_FIELD_DEFAULT, 'fieldName': ''}


def get_story_points_field_id():
    return get_story_points_field_config()['fieldId'] or STORY_POINTS_FIELD_DEFAULT


def get_parent_name_field_config():
    config = load_dashboard_config()
    if config and 'parentNameField' in config:
        el = config['parentNameField']
        return {'fieldId': el.get('fieldId', ''), 'fieldName': el.get('fieldName', '')}
    return {'fieldId': '', 'fieldName': ''}


def get_parent_name_field_id():
    cfg = get_parent_name_field_config()
    return cfg['fieldId'] or ''


def get_team_field_config():
    config = load_dashboard_config()
    if config and 'teamField' in config:
        tf = config['teamField']
        return {'fieldId': tf.get('fieldId', ''), 'fieldName': tf.get('fieldName', '')}
    return {'fieldId': TEAM_FIELD_DEFAULT, 'fieldName': ''}


def get_team_field_id():
    return get_team_field_config()['fieldId'] or TEAM_FIELD_DEFAULT


def get_configured_issue_types():
    """Return configured issue types from dashboard config. Default: ['Story']."""
    config = load_dashboard_config()
    if not config:
        return ['Story']
    types = config.get('issueTypes', None)
    if types is None:
        return ['Story']
    return [str(t).strip() for t in types if str(t).strip()]


def normalize_priority_weight_name(name):
    key = str(name or '').strip().lower()
    return PRIORITY_WEIGHT_NAME_ALIASES.get(key, key)


def build_priority_weight_defaults():
    return [dict(item) for item in PRIORITY_WEIGHT_DEFAULTS]


def normalize_priority_weight_rows(rows):
    """Validate and normalize weight rows into canonical list format."""
    if not isinstance(rows, list):
        raise ValueError('weights must be an array')
    normalized = []
    seen = set()
    for item in rows:
        if not isinstance(item, dict):
            raise ValueError('each weight entry must be an object')
        priority = str(item.get('priority', '') or '').strip()
        if not priority:
            raise ValueError('priority is required')
        norm_name = normalize_priority_weight_name(priority)
        if norm_name in seen:
            raise ValueError(f'duplicate priority: {priority}')
        raw_weight = item.get('weight', None)
        try:
            weight = float(raw_weight)
        except (TypeError, ValueError):
            raise ValueError(f'invalid weight for {priority}')
        if weight < 0:
            raise ValueError(f'weight must be non-negative for {priority}')
        seen.add(norm_name)
        normalized.append({'priority': priority, 'weight': weight})
    return normalized


def parse_stats_priority_weights_env(raw):
    if not raw:
        return None
    rows = []
    for chunk in str(raw).split(','):
        token = chunk.strip()
        if not token:
            continue
        if ':' not in token:
            raise ValueError(f'invalid STATS_PRIORITY_WEIGHTS token: {token}')
        name, weight = token.split(':', 1)
        rows.append({'priority': name.strip(), 'weight': weight.strip()})
    return normalize_priority_weight_rows(rows)


def get_priority_weights_config():
    """Return effective stats priority weights with source metadata."""
    config = load_dashboard_config()
    if config and 'statsPriorityWeights' in config:
        try:
            rows = normalize_priority_weight_rows(config.get('statsPriorityWeights') or [])
            return {'weights': rows, 'source': 'config'}
        except ValueError as e:
            log_warning(f'Invalid statsPriorityWeights in dashboard config; falling back: {e}')

    if STATS_PRIORITY_WEIGHTS:
        try:
            rows = parse_stats_priority_weights_env(STATS_PRIORITY_WEIGHTS)
            if rows:
                return {'weights': rows, 'source': 'env'}
        except ValueError as e:
            log_warning(f'Invalid STATS_PRIORITY_WEIGHTS env; using defaults: {e}')

    return {'weights': build_priority_weight_defaults(), 'source': 'default'}


def get_effective_capacity_project():
    """Return the effective capacity project name."""
    return get_capacity_config()['project']


def parse_groups_config_env():
    if not TEAM_GROUPS_JSON:
        return None
    try:
        return json.loads(TEAM_GROUPS_JSON)
    except Exception as e:
        log_warning(f'Failed to parse TEAM_GROUPS_JSON: {e}')
        return None


def invalidate_sprints_cache():
    return _sprints_service.invalidate_sprints_cache(
        SPRINTS_CACHE_FILE,
        log_warning_fn=log_warning,
    )


def run_git_command(args):
    return _update_check_service.run_git_command(
        args,
        repo_dir=os.path.dirname(os.path.abspath(__file__)),
    )


def load_release_info():
    return _update_check_service.load_release_info(
        UPDATE_CHECK_RELEASE_INFO,
        base_dir=os.path.dirname(os.path.abspath(__file__)),
        log_warning_fn=log_warning,
    )


def build_update_check_payload():
    return _update_check_service.build_update_check_payload(
        remote=UPDATE_CHECK_REMOTE,
        branch=UPDATE_CHECK_BRANCH,
        run_git_command_fn=run_git_command,
        load_release_info_fn=load_release_info,
        now_iso_fn=utc_now_iso,
    )


def normalize_team_catalog(raw):
    catalog = {}
    if isinstance(raw, list):
        for item in raw:
            if not isinstance(item, dict):
                continue
            team_id = str(item.get('id') or '').strip()
            name = str(item.get('name') or '').strip()
            if not team_id or not name:
                continue
            catalog[team_id] = {'id': team_id, 'name': name}
    elif isinstance(raw, dict):
        for key, value in raw.items():
            if isinstance(value, dict):
                team_id = str(value.get('id') or key or '').strip()
                name = str(value.get('name') or '').strip()
            else:
                team_id = str(key or '').strip()
                name = str(value or '').strip()
            if not team_id or not name:
                continue
            catalog[team_id] = {'id': team_id, 'name': name}
    return catalog


def normalize_team_catalog_meta(raw):
    if not isinstance(raw, dict):
        return {}
    meta = {}
    for key in ('updatedAt', 'sprintId', 'sprintName', 'source', 'resolvedAt'):
        value = raw.get(key)
        if value is None:
            continue
        meta[key] = str(value)
    return meta


def normalize_group_team_labels(raw, team_ids):
    if not isinstance(raw, dict):
        return {}
    allowed_ids = set(normalize_team_ids(team_ids or []))
    labels = {}
    for raw_team_id, raw_label in raw.items():
        team_id = str(raw_team_id or '').strip()
        label = str(raw_label or '').strip()
        if not team_id or not label or team_id not in allowed_ids:
            continue
        labels[team_id] = label
    return labels


def validate_groups_config(payload, allow_empty=False):
    errors = []
    warnings = []
    if not isinstance(payload, dict):
        errors.append('Config must be an object.')
        return None, errors, warnings

    groups_raw = payload.get('groups')
    if not isinstance(groups_raw, list):
        errors.append('groups must be a list.')
        return None, errors, warnings

    normalized_groups = []
    seen_ids = set()
    seen_names = set()
    for idx, group in enumerate(groups_raw):
        if not isinstance(group, dict):
            errors.append(f'Group at index {idx} must be an object.')
            continue
        group_id = str(group.get('id') or '').strip()
        name = str(group.get('name') or '').strip()
        if not group_id:
            errors.append(f'Group at index {idx} is missing id.')
            continue
        if not name:
            errors.append(f'Group "{group_id}" is missing name.')
            continue
        if group_id.lower() in seen_ids:
            errors.append(f'Duplicate group id "{group_id}".')
            continue
        if name.lower() in seen_names:
            errors.append(f'Duplicate group name "{name}".')
            continue
        seen_ids.add(group_id.lower())
        seen_names.add(name.lower())

        team_ids = normalize_team_ids(group.get('teamIds') or [])
        if not team_ids and not allow_empty:
            errors.append(f'Group "{name}" must include at least one team.')
        if len(team_ids) > GROUPS_MAX_TEAMS:
            errors.append(f'Group "{name}" exceeds {GROUPS_MAX_TEAMS} teams.')
        raw_components = group.get('missingInfoComponents')
        if isinstance(raw_components, list):
            missing_info_components = [str(c).strip() for c in raw_components if str(c).strip()]
        elif isinstance(raw_components, str) and raw_components.strip():
            missing_info_components = [raw_components.strip()]
        else:
            # Backwards compat: accept old singular field
            old_single = str(group.get('missingInfoComponent') or '').strip()
            missing_info_components = [old_single] if old_single else []
        raw_excluded_epics = group.get('excludedCapacityEpics')
        if isinstance(raw_excluded_epics, list):
            excluded_capacity_epics = normalize_epic_keys(raw_excluded_epics)
        elif isinstance(raw_excluded_epics, str) and raw_excluded_epics.strip():
            excluded_capacity_epics = normalize_epic_keys([raw_excluded_epics.strip()])
        else:
            excluded_capacity_epics = []
        team_labels = normalize_group_team_labels(group.get('teamLabels') or {}, team_ids)
        normalized_groups.append({
            'id': group_id,
            'name': name,
            'teamIds': team_ids,
            'missingInfoComponents': missing_info_components,
            'excludedCapacityEpics': excluded_capacity_epics,
            'teamLabels': team_labels
        })

    default_group_id = str(payload.get('defaultGroupId') or '').strip()
    if default_group_id:
        if default_group_id not in {g['id'] for g in normalized_groups}:
            errors.append('defaultGroupId must reference an existing group.')

    normalized = {
        'version': payload.get('version') or GROUPS_CONFIG_VERSION,
        'groups': normalized_groups,
        'defaultGroupId': default_group_id,
    }
    return normalized, errors, warnings


def build_default_groups_config():
    warnings = []
    team_ids = normalize_team_ids(extract_team_ids_from_jql(build_base_jql()))
    if len(team_ids) > GROUPS_MAX_TEAMS:
        warnings.append(f'Found more than {GROUPS_MAX_TEAMS} teams in JQL_QUERY; truncated to first {GROUPS_MAX_TEAMS}.')
        team_ids = team_ids[:GROUPS_MAX_TEAMS]
    if not team_ids:
        warnings.append('No teams found in JQL_QUERY. Default group is empty; add teams manually.')

    config = {
        'version': GROUPS_CONFIG_VERSION,
        'groups': [{
            'id': 'default',
            'name': 'Default',
            'teamIds': team_ids,
            'missingInfoComponents': [MISSING_INFO_COMPONENT] if MISSING_INFO_COMPONENT else [],
            'excludedCapacityEpics': []
        }],
        'defaultGroupId': 'default',
    }
    return config, warnings


def apply_team_ids_to_template(team_ids):
    if not JQL_QUERY_TEMPLATE:
        return None
    ids = normalize_team_ids(team_ids)
    quoted = ', '.join(f'"{team_id}"' for team_id in ids)
    return JQL_QUERY_TEMPLATE.replace('{TEAM_IDS}', quoted)


def build_tasks_cache_key(sprint, group_id, project_filter, team_ids, include_team_name, use_template, purpose='dashboard', epic_keys=None):
    epic_signature = ','.join(sorted({str(k).strip() for k in (epic_keys or []) if str(k).strip()}))
    raw = f"{TASKS_CACHE_SCHEMA_VERSION}::{purpose}::{sprint}::{group_id}::{project_filter}::{','.join(team_ids)}::{include_team_name}::{use_template}::{epic_signature}"
    digest = hashlib.sha1(raw.encode('utf-8')).hexdigest()[:12]
    return f"tasks:{digest}"


def classify_project(project_name, project_key=None):
    """Classify projects into product/tech buckets based on dashboard config, falling back to env vars."""
    if not project_name and not project_key:
        return 'other'
    # First check dashboard config for typed classification
    typed = get_selected_projects_typed()
    if typed and project_key:
        for item in typed:
            if item['key'].upper() == project_key.upper():
                return item['type']
    # Fallback to env-var-based classification by project name
    if project_name:
        normalized = str(project_name).strip().lower()
        if any(normalized == p.lower() for p in STATS_PRODUCT_PROJECTS):
            return 'product'
        if any(normalized == p.lower() for p in STATS_TECH_PROJECTS):
            return 'tech'
    return 'other'


def fetch_board_sprint_ids(board_id, headers):
    """Fetch the set of sprint IDs that originated on a specific board.
    Uses originBoardId to exclude cross-board sprints that Jira includes
    in the board API response."""
    return _sprints_service.fetch_board_sprint_ids(
        board_id,
        jira_get=current_jira_get,
        auth_error_class=AuthError,
        log_warning_fn=log_warning,
    )


def deduplicate_sprints_by_name(sprints, board_sprint_ids=None):
    """When multiple sprints share a name, keep the one on the configured board.
    If neither or both are on the board, prefer active > closed > future."""
    return _sprints_service.deduplicate_sprints_by_name(sprints, board_sprint_ids)


def fetch_sprints_from_jira():
    """Fetch sprints from Jira (not from cache)"""
    return _sprints_service.fetch_sprints_from_jira(
        board_id=get_effective_board_id(),
        stats_jql_base=STATS_JQL_BASE,
        product_project=JIRA_PRODUCT_PROJECT,
        tech_project=JIRA_TECH_PROJECT,
        jira_get=current_jira_get,
        jira_search_request=jira_search_request,
        get_sprint_field_id=get_sprint_field_id,
        strip_sprint_clause=strip_sprint_clause,
        add_clause_to_jql=add_clause_to_jql,
        auth_error_class=AuthError,
        fetch_board_sprint_ids_fn=fetch_board_sprint_ids,
        log_info_fn=log_info,
        log_warning_fn=log_warning,
    )


def fetch_epic_details_bulk(epic_keys, headers, epic_name_field):
    """Fetch epic details in small batches to avoid per-epic network calls."""
    epic_details = {}
    if not epic_keys:
        return epic_details

    epic_field = epic_name_field or PARENT_NAME_FIELD_DEFAULT
    keys_list = list(epic_keys)
    batch_size = 40  # keep JQL length reasonable for GET

    for start in range(0, len(keys_list), batch_size):
        batch_keys = keys_list[start:start + batch_size]
        jql = f'issueKey in ({",".join(batch_keys)})'
        payload = {
            'jql': jql,
            'maxResults': len(batch_keys),
            'fields': ['summary', 'status', 'reporter', 'assignee', 'parent', epic_field]
        }

        try:
            resp = jira_search_request(payload)
            if resp.status_code != 200:
                log_warning(f'Epic batch {start}-{start + len(batch_keys)} failed: status={resp.status_code}')
                continue

            data = resp.json()
            for issue in data.get('issues', []):
                fields = issue.get('fields', {}) or {}
                key = issue.get('key')
                epic_details[key] = {
                    'key': key,
                    'summary': fields.get('summary'),
                    'status': (fields.get('status') or {}).get('name') or '',
                    'reporter': (fields.get('reporter') or {}).get('displayName'),
                    'assignee': {'displayName': (fields.get('assignee') or {}).get('displayName')} if fields.get('assignee') else None,
                }
                # Extract initiative from parent if present
                parent = fields.get('parent')
                if parent and parent.get('key'):
                    parent_fields = parent.get('fields') or {}
                    parent_type = parent_fields.get('issuetype') or {}
                    type_name = (parent_type.get('name') or '').lower()
                    hierarchy_level = parent_type.get('hierarchyLevel')
                    if type_name == 'initiative' or hierarchy_level == 0:
                        epic_details[key]['initiative'] = {
                            'key': parent['key'],
                            'summary': parent_fields.get('summary'),
                        }
        except Exception as exc:
            log_warning(f'Epic batch fetch error: {exc}')

    return epic_details


def derive_epic_jql(base_jql: str, team_ids=None) -> str:
    """Attempt to derive an epic query from a story query by swapping Story→Epic."""
    if not base_jql:
        base_jql = ''

    jql = base_jql
    replacements = [
        ('type = Story', 'type = Epic'),
        ('type=Story', 'type=Epic'),
        ('type = "Story"', 'type = "Epic"'),
        ("type = 'Story'", "type = 'Epic'"),
        ('issuetype = Story', 'issuetype = Epic'),
        ('issuetype=Story', 'issuetype=Epic'),
        ('issuetype = "Story"', 'issuetype = "Epic"'),
        ("issuetype = 'Story'", "issuetype = 'Epic'"),
    ]
    replaced = False
    for old, new in replacements:
        if old in jql:
            jql = jql.replace(old, new)
            replaced = True

    if not replaced:
        jql = add_clause_to_jql(jql, 'type = Epic') if jql else 'type = Epic'

    if EPIC_EMPTY_EXCLUDED_STATUSES:
        quoted = ', '.join(f'"{s}"' for s in EPIC_EMPTY_EXCLUDED_STATUSES)
        jql = add_clause_to_jql(jql, f'status not in ({quoted})')

    team_ids = [t.strip() for t in (team_ids or []) if t and str(t).strip()]
    if team_ids:
        if len(team_ids) == 1:
            jql = add_clause_to_jql(jql, f'"Team[Team]" = "{team_ids[0]}"')
        else:
            quoted_teams = ', '.join(f'"{t}"' for t in team_ids)
            jql = add_clause_to_jql(jql, f'"Team[Team]" in ({quoted_teams})')
    return jql


def issue_has_sprint(value):
    if value is None:
        return False
    if isinstance(value, list):
        return any(issue_has_sprint(item) for item in value)
    if isinstance(value, dict):
        return any(value.get(key) for key in ('id', 'name', 'state'))
    if isinstance(value, str):
        return bool(value.strip())
    return True


def fetch_epics_for_empty_alert(jql, headers, team_field_id, epic_name_field, sprint_field_id=None):
    """Fetch epics matching the current sprint/team filters so UI can flag epics with 0 stories."""
    epic_jql = derive_epic_jql(jql, EPIC_EMPTY_TEAM_IDS)
    epic_field = epic_name_field or PARENT_NAME_FIELD_DEFAULT

    fields_list = ['summary', 'status', 'assignee', 'labels', epic_field]
    if team_field_id and team_field_id not in fields_list:
        fields_list.append(team_field_id)
    if sprint_field_id and sprint_field_id not in fields_list:
        fields_list.append(sprint_field_id)

    payload = {
        'jql': epic_jql,
        'maxResults': 250,
        'fields': fields_list
    }
    resp = jira_search_request(payload)
    if resp.status_code != 200:
        log_warning(f'Epic empty-state fetch failed: status={resp.status_code}')
        return []

    data = resp.json() or {}
    issues = data.get('issues', []) or []
    epics = []
    for issue in issues:
        fields = issue.get('fields', {}) or {}

        raw_team = None
        if team_field_id and fields.get(team_field_id) is not None:
            raw_team = fields.get(team_field_id)

        team_value = build_team_value(raw_team) if raw_team is not None else None
        team_name = extract_team_name(raw_team) if raw_team is not None else None

        assignee = fields.get('assignee') or {}
        status = fields.get('status') or {}
        epics.append({
            'key': issue.get('key'),
            'summary': fields.get('summary'),
            'status': {'name': status.get('name')} if status else None,
            'assignee': {'displayName': assignee.get('displayName')} if assignee else None,
            'labels': fields.get('labels') or [],
            'team': team_value,
            'teamName': team_name,
            'teamId': team_value.get('id') if isinstance(team_value, dict) else None,
            'fields': {
                'customfield_10101': fields.get(sprint_field_id) if sprint_field_id else None
            },
        })
    return epics


def fetch_backlog_epics_for_alert(jql, headers, team_field_id, sprint_field_id, epic_link_field):
    """Fetch backlog epics and count open child stories that are still sprinted."""
    epic_jql = derive_epic_jql(jql, EPIC_EMPTY_TEAM_IDS)
    epic_jql = add_clause_to_jql(epic_jql, 'assignee is not EMPTY')
    epic_jql = add_clause_to_jql(epic_jql, 'component is not EMPTY')
    if sprint_field_id:
        epic_jql = add_clause_to_jql(epic_jql, f'"{sprint_field_id}" is EMPTY')

    epic_fields = ['summary', 'status', 'assignee', 'components']
    if team_field_id and team_field_id not in epic_fields:
        epic_fields.append(team_field_id)
    if sprint_field_id and sprint_field_id not in epic_fields:
        epic_fields.append(sprint_field_id)

    payload = {
        'jql': epic_jql,
        'maxResults': 250,
        'fields': epic_fields
    }
    resp = jira_search_request(payload)
    if resp.status_code != 200:
        log_warning(f'Backlog epic fetch failed: status={resp.status_code}')
        return []

    issues = (resp.json() or {}).get('issues', []) or []
    epics = []
    epic_keys = []
    for issue in issues:
        fields = issue.get('fields', {}) or {}
        raw_team = fields.get(team_field_id) if team_field_id else None
        team_value = build_team_value(raw_team) if raw_team is not None else None
        team_name = extract_team_name(raw_team) if raw_team is not None else None
        assignee = fields.get('assignee') or {}
        status = fields.get('status') or {}
        epics.append({
            'key': issue.get('key'),
            'summary': fields.get('summary'),
            'status': {'name': status.get('name')} if status else None,
            'assignee': {'displayName': assignee.get('displayName')} if assignee else None,
            'components': [c.get('name') for c in (fields.get('components') or []) if c.get('name')],
            'team': team_value,
            'teamName': team_name,
            'teamId': team_value.get('id') if isinstance(team_value, dict) else None,
            'fields': {
                'customfield_10101': fields.get(sprint_field_id) if sprint_field_id else None
            },
            'cleanupStoryCount': 0,
        })
        if issue.get('key'):
            epic_keys.append(issue.get('key'))

    if not epics or not sprint_field_id:
        return epics

    quoted_keys = ', '.join(epic_keys)
    children_jql = f'("Epic Link" in ({quoted_keys}) OR parent in ({quoted_keys})) AND issuetype != Epic'
    child_payload = {
        'jql': children_jql,
        'maxResults': 250,
        'fields': [epic_link_field, 'parent', 'status', sprint_field_id]
    }
    child_resp = jira_search_request(child_payload)
    if child_resp.status_code != 200:
        log_warning(f'Backlog child fetch failed: status={child_resp.status_code}')
        return epics

    cleanup_counts = {}
    for issue in (child_resp.json() or {}).get('issues', []) or []:
        fields = issue.get('fields', {}) or {}
        epic_key = fields.get(epic_link_field) if epic_link_field else None
        if not epic_key:
            epic_key = (fields.get('parent') or {}).get('key')
        if not epic_key:
            continue
        status_name = str((fields.get('status') or {}).get('name') or '').strip().lower()
        if status_name in {'done', 'killed', 'incomplete'}:
            continue
        if not issue_has_sprint(fields.get(sprint_field_id)):
            continue
        cleanup_counts[epic_key] = cleanup_counts.get(epic_key, 0) + 1

    for epic in epics:
        epic['cleanupStoryCount'] = cleanup_counts.get(epic.get('key'), 0)
    return epics


def fetch_story_counts_for_epics(epic_keys, headers, epic_link_field):
    """Return total Story counts for each epic key.

    Prefer counting via the Epic Link field (company-managed Jira). If that yields zero for an epic,
    fall back to counting via `parent` (team-managed projects / some Jira configs).
    """
    epic_keys = [k for k in (epic_keys or []) if k]
    if not epic_keys:
        return {}

    def count_by_query(batch_keys, jql, fields):
        next_page_token = None
        max_results = 250
        local_counts = {k: 0 for k in batch_keys}

        while True:
            payload = {
                'jql': jql,
                'maxResults': max_results,
                'fields': fields
            }
            if next_page_token:
                payload['nextPageToken'] = next_page_token
            resp = jira_search_request(payload)
            if resp.status_code != 200:
                return local_counts

            data = resp.json() or {}
            issues = data.get('issues', []) or []
            if not issues:
                break

            for issue in issues:
                fields_obj = issue.get('fields', {}) or {}
                epic_key = fields_obj.get(epic_link_field) if epic_link_field else None
                if not epic_key:
                    epic_key = (fields_obj.get('parent') or {}).get('key')
                if epic_key in local_counts:
                    local_counts[epic_key] += 1

            next_page_token = data.get('nextPageToken')
            if data.get('isLast', not next_page_token) or not next_page_token:
                break

        return local_counts

    counts = {k: 0 for k in epic_keys}
    batch_size = 40

    for start in range(0, len(epic_keys), batch_size):
        batch = epic_keys[start:start + batch_size]
        quoted_keys = ', '.join(f'"{k}"' for k in batch)
        if epic_link_field:
            combined_jql = f'(("Epic Link" in ({quoted_keys})) OR (parent in ({quoted_keys}))) AND issuetype != Epic'
            fields = [epic_link_field, 'parent']
        else:
            combined_jql = f'parent in ({quoted_keys}) AND issuetype != Epic'
            fields = ['parent']

        batch_counts = count_by_query(batch, combined_jql, fields)
        for k, v in batch_counts.items():
            counts[k] += v

    return counts


def fetch_story_distribution_for_epics(epic_keys, headers, epic_link_field, selected_sprint):
    """Return selected/future not-completed story counts for each epic key."""
    epic_keys = [k for k in (epic_keys or []) if k]
    if not epic_keys:
        return {}

    distribution = {
        key: {
            'selectedStories': 0,
            'selectedActionableStories': 0,
            'futureOpenStories': 0,
            'openStoriesOutsideSelected': 0
        } for key in epic_keys
    }
    batch_size = 40

    def is_actionable_selected_status(raw_status):
        name = str(raw_status or '').strip().lower()
        return name not in ('blocked', 'done', 'killed', 'incomplete')

    def count_for_batch(batch_keys, where_clause, bucket_name):
        if not where_clause:
            return
        quoted_keys = ', '.join(f'"{k}"' for k in batch_keys)

        def run_query(jql, fields):
            next_page_token = None
            max_results = 250
            while True:
                payload = {
                    'jql': jql,
                    'maxResults': max_results,
                    'fields': fields
                }
                if next_page_token:
                    payload['nextPageToken'] = next_page_token
                resp = jira_search_request(payload)
                if resp.status_code != 200:
                    return
                data = resp.json() or {}
                issues = data.get('issues', []) or []
                if not issues:
                    break
                for issue in issues:
                    fields_obj = issue.get('fields', {}) or {}
                    epic_key = fields_obj.get(epic_link_field) if epic_link_field else None
                    if not epic_key:
                        epic_key = (fields_obj.get('parent') or {}).get('key')
                    if epic_key in distribution:
                        distribution[epic_key][bucket_name] += 1
                next_page_token = data.get('nextPageToken')
                if data.get('isLast', not next_page_token) or not next_page_token:
                    break

        if epic_link_field:
            combined_jql = (
                f'{where_clause} AND (("Epic Link" in ({quoted_keys})) OR (parent in ({quoted_keys})))'
            )
            run_query(combined_jql, [epic_link_field, 'parent'])
        else:
            jql_parent = f'{where_clause} AND parent in ({quoted_keys})'
            run_query(jql_parent, ['parent'])

    def count_selected_for_batch(batch_keys):
        if not selected_sprint:
            return
        quoted_keys = ', '.join(f'"{k}"' for k in batch_keys)
        where_clause = f'Sprint = {selected_sprint} AND issuetype != Epic'

        def run_query(jql, fields):
            next_page_token = None
            max_results = 250
            while True:
                payload = {
                    'jql': jql,
                    'maxResults': max_results,
                    'fields': fields
                }
                if next_page_token:
                    payload['nextPageToken'] = next_page_token
                resp = jira_search_request(payload)
                if resp.status_code != 200:
                    return
                data = resp.json() or {}
                issues = data.get('issues', []) or []
                if not issues:
                    break
                for issue in issues:
                    fields_obj = issue.get('fields', {}) or {}
                    epic_key = fields_obj.get(epic_link_field) if epic_link_field else None
                    if not epic_key:
                        epic_key = (fields_obj.get('parent') or {}).get('key')
                    if epic_key not in distribution:
                        continue
                    distribution[epic_key]['selectedStories'] += 1
                    status_name = ((fields_obj.get('status') or {}).get('name') if isinstance(fields_obj.get('status'), dict) else None)
                    if is_actionable_selected_status(status_name):
                        distribution[epic_key]['selectedActionableStories'] += 1
                next_page_token = data.get('nextPageToken')
                if data.get('isLast', not next_page_token) or not next_page_token:
                    break

        if epic_link_field:
            combined_jql = (
                f'{where_clause} AND (("Epic Link" in ({quoted_keys})) OR (parent in ({quoted_keys})))'
            )
            run_query(combined_jql, [epic_link_field, 'parent', 'status'])
        else:
            jql_parent = f'{where_clause} AND parent in ({quoted_keys})'
            run_query(jql_parent, ['parent', 'status'])

    selected_clause = ''
    if selected_sprint:
        selected_clause = f'Sprint = {selected_sprint} AND issuetype != Epic'
    future_clause = 'Sprint in futureSprints() AND issuetype != Epic AND status not in ("Done","Killed","Incomplete")'
    outside_selected_clause = (
        f'issuetype != Epic AND status not in ("Done","Killed","Incomplete") AND (Sprint != {selected_sprint} OR Sprint is EMPTY)'
        if selected_sprint else
        'issuetype != Epic AND status not in ("Done","Killed","Incomplete")'
    )

    for start in range(0, len(epic_keys), batch_size):
        batch = epic_keys[start:start + batch_size]
        if selected_clause:
            count_selected_for_batch(batch)
        count_for_batch(batch, future_clause, 'futureOpenStories')
        count_for_batch(batch, outside_selected_clause, 'openStoriesOutsideSelected')

    return distribution


def fetch_tasks(include_team_name=False):
    """Fetch tasks from Jira API."""
    try:
        request_started = time.perf_counter()
        timings_ms = {}
        include_debug_timings = request.args.get('debugTimings', '').strip().lower() in ('1', 'true', 'yes')
        def record_timing(name, started_at):
            timings_ms[name] = round((time.perf_counter() - started_at) * 1000, 1)

        # Get sprint parameter from query string
        parse_started = time.perf_counter()
        sprint = request.args.get('sprint', '')
        team = request.args.get('team', '').strip()
        group_id = request.args.get('groupId', '').strip() or 'default'
        team_ids_param = request.args.get('teamIds', '').strip()
        epic_keys_param = request.args.get('epicKeys', '').strip()
        project_filter = request.args.get('project', '').strip().lower()
        request_purpose = request.args.get('purpose', 'dashboard').strip().lower() or 'dashboard'
        force_refresh = request.args.get('refresh', '').lower() in ('1', 'true')
        team_ids = normalize_team_ids([t.strip() for t in team_ids_param.split(',') if t.strip()])
        epic_keys_filter = sorted({t.strip() for t in epic_keys_param.split(',') if t.strip()})
        use_template = bool(team_ids and JQL_QUERY_TEMPLATE)
        lightweight_ready_to_close = request_purpose == 'ready-to-close'
        raw_cache_key = build_tasks_cache_key(
            sprint,
            group_id,
            project_filter,
            team_ids if use_template else [],
            include_team_name,
            use_template,
            request_purpose,
            epic_keys_filter
        )
        record_timing('parse_params', parse_started)
        auth_context = current_request_auth_context()
        if project_filter in ('product', 'tech'):
            denied_response, denied_status = project_access_denied_response(auth_context, project_filter)
            if denied_response is not None:
                return denied_response, denied_status
        cache_enabled = jira_home_partitioned_process_cache_enabled(auth_context)
        cache_key = build_jira_home_process_cache_key(auth_context, raw_cache_key)
        cached_entry = None
        if cache_enabled:
            with _cache_lock:
                cached_entry = TASKS_CACHE.get(cache_key)
        if cache_enabled and not force_refresh and cached_entry and (time.time() - cached_entry.get('timestamp', 0)) < TASKS_CACHE_TTL_SECONDS:
            cached_response = jsonify(cached_entry.get('data') or {})
            cached_response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
            cached_response.headers['Pragma'] = 'no-cache'
            cached_response.headers['Expires'] = '0'
            cached_response.headers['Server-Timing'] = 'cache;dur=1'
            return cached_response

        auth_started = time.perf_counter()
        headers = None
        record_timing('auth_headers', auth_started)

        # Build JQL query with sprint filter if provided
        jql_started = time.perf_counter()
        if use_template:
            jql = apply_team_ids_to_template(team_ids)
            if not jql:
                jql = build_base_jql()
        elif team_ids:
            # If team_ids provided (from group), filter by those teams
            # Remove existing team filter from base JQL and add our own
            jql = remove_team_filter_from_jql(build_base_jql())
            if len(team_ids) == 1:
                jql = add_clause_to_jql(jql, f'"Team[Team]" = "{team_ids[0]}"')
            else:
                quoted_teams = ', '.join(f'"{tid}"' for tid in team_ids)
                jql = add_clause_to_jql(jql, f'"Team[Team]" in ({quoted_teams})')
        else:
            jql = build_base_jql()

        if sprint:
            jql = add_clause_to_jql(jql, f"Sprint = {sprint}")

        if team and team.lower() != 'all' and not use_template and not team_ids:
            jql = add_clause_to_jql(jql, f'"Team[Team]" = {team}')

        # Apply project filter — use config-typed projects when available, else env vars
        if project_filter in ('product', 'tech'):
            typed = get_selected_projects_typed()
            if typed:
                matching_keys = [item['key'] for item in typed if item['type'] == project_filter]
                if matching_keys:
                    jql = remove_project_filter_from_jql(jql)
                    if len(matching_keys) == 1:
                        jql = add_clause_to_jql(jql, f'project = "{matching_keys[0]}"')
                    else:
                        quoted = ', '.join(f'"{k}"' for k in matching_keys)
                        jql = add_clause_to_jql(jql, f'project in ({quoted})')
            else:
                project_name = JIRA_PRODUCT_PROJECT if project_filter == 'product' else JIRA_TECH_PROJECT
                jql = add_clause_to_jql(jql, f'project = "{project_name}"')

        # Apply issue type filter from config
        issue_types = get_configured_issue_types()
        if issue_types:
            if len(issue_types) == 1:
                jql = add_clause_to_jql(jql, f'type = "{issue_types[0]}"')
            else:
                quoted_types = ', '.join(f'"{t}"' for t in issue_types)
                jql = add_clause_to_jql(jql, f'type in ({quoted_types})')

        tasks_jql = jql
        if epic_keys_filter:
            quoted_epics = ', '.join(f'"{key}"' for key in epic_keys_filter)
            tasks_jql = add_clause_to_jql(tasks_jql, f'("Epic Link" in ({quoted_epics}) OR parent in ({quoted_epics}))')
        record_timing('build_jql', jql_started)

        team_field_id = resolve_team_field_id(headers, context=auth_context)
        epic_link_field_id = resolve_epic_link_field_id(headers, context=auth_context)

        sprint_field_id = get_sprint_field_id()

        # Prepare request parameters for search endpoint
        if lightweight_ready_to_close:
            fields_list = [
                'status',
                'parent'
            ]
        else:
            fields_list = [
                'summary',
                'status',
                'priority',
                'issuetype',
                'assignee',
                'updated',
                get_story_points_field_id(),  # Story Points
                'parent',
                'project'
            ]
        if sprint_field_id and sprint_field_id not in fields_list:
            fields_list.append(sprint_field_id)
        if epic_link_field_id and epic_link_field_id not in fields_list:
            fields_list.append(epic_link_field_id)
        if team_field_id:
            fields_list.append(team_field_id)

        max_results = 250
        page_size = 100
        next_page_token = None
        collected_issues = []
        names_map = {}
        total_issues = None

        log_info(
            f'Jira task fetch start purpose={request_purpose} sprint={sprint or "all"} '
            f'project={project_filter or "all"} mode={"lightweight" if lightweight_ready_to_close else "full"}'
        )

        jira_fetch_started = time.perf_counter()
        while len(collected_issues) < max_results:
            remaining = max_results - len(collected_issues)
            page_limit = min(page_size, remaining)
            payload = {
                'jql': tasks_jql,
                'maxResults': page_limit,
                'fields': fields_list
            }
            if next_page_token:
                payload['nextPageToken'] = next_page_token

            response = jira_search_request(payload)
            log_debug(f'Jira search page response status={response.status_code}')

            if response.status_code != 200:
                error_text = response.text
                log_error(f'Jira search failed: status={response.status_code}')

                try:
                    error_json = response.json()
                    log_debug(f'Jira error payload keys={sorted((error_json or {}).keys()) if isinstance(error_json, dict) else "non-dict"}')
                except Exception:
                    pass

                error_response = jsonify({
                    'error': f'Jira API error: {response.status_code}',
                    'details': error_text,
                    'jql_used': tasks_jql
                })
                error_response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
                error_response.headers['Pragma'] = 'no-cache'
                error_response.headers['Expires'] = '0'
                return error_response, response.status_code

            data = response.json()
            if not names_map:
                names_map = data.get('names', {}) or {}
            total_issues = data.get('total', total_issues)

            issues = data.get('issues', [])
            if not issues:
                break

            collected_issues.extend(issues)
            next_page_token = data.get('nextPageToken')
            if data.get('isLast', not next_page_token) or not next_page_token:
                break
        record_timing('jira_search', jira_fetch_started)

        data = {
            'issues': collected_issues,
            'names': names_map,
            'total': total_issues,
            'startAt': 0,
            'maxResults': max_results
        }

        if total_issues is None:
            total_issues = len(collected_issues)

        if not team_field_id:
            team_field_id = next((k for k, v in names_map.items() if str(v).lower() == 'team[team]'), None)
        epic_link_field = epic_link_field_id or resolve_epic_link_field_id(headers, names_map, context=auth_context)
        epic_name_field = next((k for k, v in names_map.items() if str(v).lower() == 'epic name'), None)
        epic_keys = set()
        normalize_started = time.perf_counter()
        for issue in collected_issues:
            fields = issue.get('fields', {})

            raw_team = None
            if team_field_id and fields.get(team_field_id) is not None:
                raw_team = fields.get(team_field_id)

            if raw_team is not None:
                team_name = extract_team_name(raw_team)
                fields['team'] = build_team_value(raw_team)
                fields['teamName'] = team_name
                fields['teamId'] = fields['team'].get('id') if isinstance(fields['team'], dict) else None

            parent_fields = fields.get('parent', {}).get('fields', {})
            if parent_fields.get('summary'):
                fields['parentSummary'] = parent_fields.get('summary')

            epic_key = None
            if epic_link_field and fields.get(epic_link_field):
                epic_key = fields.get(epic_link_field)
            elif fields.get('parent') and fields['parent'].get('key') and \
                    fields['parent'].get('fields', {}).get('issuetype', {}).get('name', '').lower() == 'epic':
                epic_key = fields['parent'].get('key')

            if epic_key:
                fields['epicKey'] = epic_key
                epic_keys.add(epic_key)
        record_timing('normalize_tasks', normalize_started)

        enrich_epics_started = time.perf_counter()
        if lightweight_ready_to_close:
            epic_details = {}
            epics_in_scope = fetch_epics_for_empty_alert(
                jql,
                headers,
                team_field_id,
                epic_name_field,
                sprint_field_id=sprint_field_id
            )
        else:
            if JIRA_AUTH_MODE == AUTH_MODE_ATLASSIAN_OAUTH:
                epic_details = fetch_epic_details_bulk(epic_keys, headers, epic_name_field)
                epics_in_scope = fetch_epics_for_empty_alert(
                    jql,
                    headers,
                    team_field_id,
                    epic_name_field,
                    sprint_field_id
                )
            else:
                with ThreadPoolExecutor(max_workers=2) as pool:
                    future_epic_details = pool.submit(fetch_epic_details_bulk, epic_keys, headers, epic_name_field)
                    future_epics_in_scope = pool.submit(
                        fetch_epics_for_empty_alert,
                        jql,
                        headers,
                        team_field_id,
                        epic_name_field,
                        sprint_field_id
                    )
                    epic_details = future_epic_details.result()
                    epics_in_scope = future_epics_in_scope.result()
        record_timing('epic_enrichment', enrich_epics_started)

        if epic_keys_filter:
            epic_filter_set = set(epic_keys_filter)
            epics_in_scope = [epic for epic in epics_in_scope if epic.get('key') in epic_filter_set]
        if not lightweight_ready_to_close:
            enrich_counts_started = time.perf_counter()
            epic_scope_keys = [e.get('key') for e in epics_in_scope]
            if JIRA_AUTH_MODE == AUTH_MODE_ATLASSIAN_OAUTH:
                epic_story_counts = (
                    fetch_story_counts_for_epics(epic_scope_keys, headers, epic_link_field)
                    if epic_link_field else None
                )
                epic_story_distribution = fetch_story_distribution_for_epics(epic_scope_keys, headers, epic_link_field, sprint)
            else:
                with ThreadPoolExecutor(max_workers=2) as pool:
                    future_epic_story_counts = (
                        pool.submit(fetch_story_counts_for_epics, epic_scope_keys, headers, epic_link_field)
                        if epic_link_field else None
                    )
                    future_epic_story_distribution = pool.submit(
                        fetch_story_distribution_for_epics, epic_scope_keys, headers, epic_link_field, sprint
                    )
                    epic_story_counts = future_epic_story_counts.result() if future_epic_story_counts else None
                    epic_story_distribution = future_epic_story_distribution.result()
            record_timing('epic_counts_distribution', enrich_counts_started)
            for epic in epics_in_scope:
                key = epic.get('key')
                epic['totalStories'] = epic_story_counts.get(key) if (epic_story_counts and key) else None
                if key and epic_story_distribution.get(key):
                    epic['selectedStories'] = epic_story_distribution[key].get('selectedStories', 0)
                    epic['selectedActionableStories'] = epic_story_distribution[key].get('selectedActionableStories', 0)
                    epic['futureOpenStories'] = epic_story_distribution[key].get('futureOpenStories', 0)
                else:
                    epic['selectedStories'] = 0
                    epic['selectedActionableStories'] = 0
                    epic['futureOpenStories'] = 0
        slim_build_started = time.perf_counter()
        slim_issues = []
        for issue in collected_issues:
            fields = issue.get('fields', {})
            status = fields.get('status') or {}
            priority = fields.get('priority') or {}
            issuetype = fields.get('issuetype') or {}
            assignee = fields.get('assignee') or {}
            project_field = fields.get('project') or {}
            if lightweight_ready_to_close:
                slim_issues.append({
                    'key': issue.get('key'),
                    'fields': {
                        'status': {'name': status.get('name')} if status else None,
                        'team': fields.get('team'),
                        'teamName': fields.get('teamName'),
                        'teamId': fields.get('teamId'),
                        'epicKey': fields.get('epicKey'),
                        'customfield_10101': fields.get(sprint_field_id) if sprint_field_id else None
                    }
                })
            else:
                slim_issues.append({
                    'id': issue.get('id'),
                    'key': issue.get('key'),
                    'fields': {
                        'summary': fields.get('summary'),
                        'status': {'name': status.get('name')} if status else None,
                        'priority': {'name': priority.get('name')} if priority else None,
                        'issuetype': {'name': issuetype.get('name')} if issuetype else None,
                        'assignee': {'displayName': assignee.get('displayName')} if assignee else None,
                        'updated': fields.get('updated'),
                        'customfield_10004': fields.get(get_story_points_field_id()),
                        'team': fields.get('team'),
                        'teamName': fields.get('teamName'),
                        'teamId': fields.get('teamId'),
                        'epicKey': fields.get('epicKey'),
                        'customfield_10101': fields.get(sprint_field_id) if sprint_field_id else None,
                        'parentSummary': fields.get('parentSummary'),
                        'projectKey': project_field.get('key', ''),
                        'projectName': project_field.get('name', '')
                    }
                })
        record_timing('build_response', slim_build_started)

        data['issues'] = slim_issues
        data['epics'] = epic_details
        data['epicsInScope'] = epics_in_scope
        data['teamFieldId'] = team_field_id
        if include_debug_timings:
            timings_ms['issueCount'] = len(slim_issues)
            timings_ms['epicKeyCount'] = len(epic_keys)
            timings_ms['epicsInScopeCount'] = len(epics_in_scope)
            data['debugTimingsMs'] = timings_ms

        log_info(f'Tasks fetch success issues={len(data.get("issues", []))}')
        timings_ms['total'] = round((time.perf_counter() - request_started) * 1000, 1)
        log_info(
            f'⏱️ tasks-with-team-name timing purpose={request_purpose} sprint={sprint or "all"} '
            f'project={project_filter or "all"} issues={len(slim_issues)} epics={len(epics_in_scope)} '
            f'timings_ms={timings_ms}'
        )
        if cache_enabled:
            cache_store_started = time.perf_counter()
            with _cache_lock:
                TASKS_CACHE[cache_key] = {
                    'timestamp': time.time(),
                    'data': data
                }
            record_timing('cache_store', cache_store_started)

        success_response = jsonify(data)
        success_response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
        success_response.headers['Pragma'] = 'no-cache'
        success_response.headers['Expires'] = '0'
        server_timing_parts = []
        for key in ('jira_search', 'normalize_tasks', 'epic_enrichment', 'epic_counts_distribution', 'build_response'):
            value = timings_ms.get(key)
            if value is not None:
                token = key.replace('_', '-')
                server_timing_parts.append(f'{token};dur={value}')
        if server_timing_parts:
            success_response.headers['Server-Timing'] = ', '.join(server_timing_parts)
        return success_response

    except AuthError as error:
        if error.code == "auth_required":
            payload, status = oauth_auth_required_payload()
            return jsonify(payload), status
        raise
    except Exception as e:
        logger.exception('Failed to fetch tasks from Jira')
        error_response = jsonify({
            'error': 'Failed to fetch tasks from Jira',
            'message': str(e)
        })
        error_response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
        return error_response, 500


def fetch_issues_by_keys(keys, fields_list, context=None):
    """Fetch issues by keys in batches."""
    return _jira_client.fetch_issues_by_keys(
        keys,
        fields_list,
        search_request=jira_search_request,
        context=context,
        log_warning_fn=log_warning,
    )


def fetch_issues_by_jql(jql, fields_list, max_results=500, context=None):
    """Fetch issues by JQL with pagination."""
    return _jira_client.fetch_issues_by_jql(
        jql,
        fields_list,
        max_results=max_results,
        search_request=jira_search_request,
        context=context,
        log_warning_fn=log_warning,
    )


def build_scenario_jql(filters):
    sprint = filters.get('sprint')
    teams = [t for t in (filters.get('teams') or []) if t]
    if teams and JQL_QUERY_TEMPLATE:
        jql = apply_team_ids_to_template(teams)
    else:
        jql = build_base_jql()
        if teams:
            # Strip any Team[Team] filters from the base JQL before applying group teams.
            jql = remove_team_filter_from_jql(jql)
            quoted = ', '.join(f'"{t}"' for t in teams)
            jql = add_clause_to_jql(jql, f'"Team[Team]" in ({quoted})')

    if sprint:
        if str(sprint).isdigit():
            jql = add_clause_to_jql(jql, f"Sprint = {sprint}")
        else:
            jql = add_clause_to_jql(jql, f'Sprint = "{sprint}"')

    projects = [p for p in (filters.get('projects') or []) if p]
    if projects:
        quoted = ', '.join(f'"{p}"' for p in projects)
        jql = add_clause_to_jql(jql, f'project in ({quoted})')

    epics = [e for e in (filters.get('epics') or []) if e]
    if epics:
        quoted = ', '.join(f'"{e}"' for e in epics)
        jql = add_clause_to_jql(jql, f'("Epic Link" in ({quoted}) OR parent in ({quoted}))')

    # Apply issue type filter from config (exclude epics etc.)
    issue_types = get_configured_issue_types()
    if issue_types:
        if len(issue_types) == 1:
            jql = add_clause_to_jql(jql, f'type = "{issue_types[0]}"')
        else:
            quoted_types = ', '.join(f'"{t}"' for t in issue_types)
            jql = add_clause_to_jql(jql, f'type in ({quoted_types})')

    return jql


def build_issue_snapshot(issue, team_field_id=None, epic_link_field_id=None):
    """Build a compact issue snapshot for dependency rendering."""
    fields = issue.get('fields', {}) or {}
    raw_team = None
    if team_field_id and fields.get(team_field_id) is not None:
        raw_team = fields.get(team_field_id)

    team_name = None
    team_id = None
    if raw_team is not None:
        team_name = extract_team_name(raw_team)
        team_value = build_team_value(raw_team)
        if isinstance(team_value, dict):
            team_id = team_value.get('id')

    epic_key = None
    if epic_link_field_id and fields.get(epic_link_field_id):
        epic_key = fields.get(epic_link_field_id)
    elif fields.get('parent') and fields['parent'].get('key') and \
            fields['parent'].get('fields', {}).get('issuetype', {}).get('name', '').lower() == 'epic':
        epic_key = fields['parent'].get('key')

    assignee_name = None
    if fields.get('assignee'):
        assignee_name = fields['assignee'].get('displayName') or fields['assignee'].get('name')

    return {
        'key': issue.get('key'),
        'summary': fields.get('summary'),
        'issuetype': fields.get('issuetype', {}).get('name') if fields.get('issuetype') else None,
        'status': fields.get('status', {}).get('name') if fields.get('status') else None,
        'priority': fields.get('priority', {}).get('name') if fields.get('priority') else None,
        'storyPoints': fields.get(get_story_points_field_id()),
        'teamName': team_name,
        'teamId': team_id,
        'epicKey': epic_key,
        'assignee': assignee_name
    }


def collect_dependencies(keys, context=None):
    """Fetch dependency links for a set of issues."""
    keys = sorted({str(k).strip() for k in keys if str(k).strip()})
    if not keys:
        return {}

    def normalize_link_text(value):
        return str(value or '').strip().lower()

    def has_block_marker(*values):
        return any('block' in normalize_link_text(value) for value in values if value)

    def has_depend_marker(*values):
        return any('depend' in normalize_link_text(value) for value in values if value)

    def resolve_link_direction(category, relation, direction, base_key, linked_key):
        relation_text = normalize_link_text(relation)
        if category == 'block':
            base_blocks = 'block' in relation_text and 'blocked by' not in relation_text
            base_blocked = 'blocked by' in relation_text
            if base_blocks:
                return base_key, linked_key
            if base_blocked:
                return linked_key, base_key
            return (base_key, linked_key) if direction == 'outward' else (linked_key, base_key)
        if category == 'dependency':
            base_depends = 'depend' in relation_text and 'depended on' not in relation_text
            base_depended = 'depended on' in relation_text
            if base_depends:
                return linked_key, base_key
            if base_depended:
                return base_key, linked_key
            return (linked_key, base_key) if direction == 'outward' else (base_key, linked_key)
        return None, None

    auth_context = context if context is not None else (current_request_auth_context() if has_request_context() else None)
    team_field_id = resolve_team_field_id(None, context=auth_context)
    epic_link_field_id = resolve_epic_link_field_id(None, context=auth_context)

    fields_list = [
        'summary',
        'status',
        'priority',
        'issuetype',
        get_story_points_field_id(),
        'parent',
        'issuelinks'
    ]
    if epic_link_field_id and epic_link_field_id not in fields_list:
        fields_list.append(epic_link_field_id)
    if team_field_id and team_field_id not in fields_list:
        fields_list.append(team_field_id)

    issues = fetch_issues_by_keys(keys, fields_list)
    issue_map = {}
    linked_keys = set()
    for issue in issues:
        snapshot = build_issue_snapshot(issue, team_field_id, epic_link_field_id)
        if snapshot.get('key'):
            issue_map[snapshot['key']] = snapshot
        for link in issue.get('fields', {}).get('issuelinks', []) or []:
            linked = link.get('outwardIssue') or link.get('inwardIssue')
            if linked and linked.get('key'):
                linked_keys.add(linked['key'])

    missing_linked = sorted(linked_keys - set(issue_map.keys()))
    if missing_linked:
        linked_issues = fetch_issues_by_keys(missing_linked, fields_list)
        for issue in linked_issues:
            snapshot = build_issue_snapshot(issue, team_field_id, epic_link_field_id)
            if snapshot.get('key'):
                issue_map[snapshot['key']] = snapshot

    dependencies = {}
    for issue in issues:
        base_key = issue.get('key')
        if not base_key:
            continue
        links = issue.get('fields', {}).get('issuelinks', []) or []
        entries = []
        for link in links:
            type_info = link.get('type', {}) or {}
            type_name = type_info.get('name')
            type_inward = type_info.get('inward')
            type_outward = type_info.get('outward')
            linked_issue = link.get('outwardIssue')
            direction = 'outward'
            relation = type_outward
            if linked_issue is None:
                linked_issue = link.get('inwardIssue')
                direction = 'inward'
                relation = type_inward
            if not linked_issue or not linked_issue.get('key'):
                continue
            category = None
            if has_block_marker(type_name, type_inward, type_outward):
                category = 'block'
            elif has_depend_marker(type_name, type_inward, type_outward):
                category = 'dependency'
            if category is None:
                continue
            linked_key = linked_issue.get('key')
            prereq_key, dependent_key = resolve_link_direction(
                category,
                relation,
                direction,
                base_key,
                linked_key
            )
            linked_snapshot = issue_map.get(linked_key)
            if not linked_snapshot:
                linked_snapshot = {
                    'key': linked_key,
                    'summary': linked_issue.get('fields', {}).get('summary'),
                    'issuetype': linked_issue.get('fields', {}).get('issuetype', {}).get('name'),
                    'status': linked_issue.get('fields', {}).get('status', {}).get('name'),
                    'storyPoints': linked_issue.get('fields', {}).get(get_story_points_field_id()),
                    'teamName': None,
                    'teamId': None,
                    'epicKey': None
                }
            entries.append({
                **linked_snapshot,
                'direction': direction,
                'relation': relation or type_name,
                'category': category,
                'typeName': type_name,
                'typeInward': type_inward,
                'typeOutward': type_outward,
                'prereqKey': prereq_key,
                'dependentKey': dependent_key
            })
        if entries:
            dependencies[base_key] = entries

    return dependencies


def scenario_planner():
    """Scenario planner endpoint."""
    auth_context = current_request_auth_context()
    cache_enabled = jira_home_process_cache_enabled(auth_context)
    if request.method == 'GET':
        if cache_enabled:
            with _cache_lock:
                if not SCENARIO_CACHE.get('data'):
                    return jsonify({'error': 'No scenario cached'}), 404
                return jsonify(SCENARIO_CACHE)
        return jsonify({'error': 'No scenario cached'}), 404

    try:
        payload = request.get_json(silent=True) or {}
        config_payload = payload.get('config') or {}
        filters = payload.get('filters') or {}

        sprint_label = resolve_sprint_label(filters.get('sprint'), cache_enabled=cache_enabled)
        quarter_start, quarter_end = quarter_dates_from_label(sprint_label)

        # Build sprint boundaries (selected + previous/next neighbors)
        sprint_boundaries = None
        if cache_enabled and sprint_label:
            cache_data = load_sprints_cache() or {}
            cached_sprints = cache_data.get('sprints') or []
            # Sort chronologically by name (e.g. 2025Q4, 2026Q1, 2026Q2)
            sorted_sprints = sorted(cached_sprints, key=lambda s: s.get('name', ''))
            selected_idx = None
            for i, s in enumerate(sorted_sprints):
                if s.get('name') == sprint_label:
                    selected_idx = i
                    break
            if selected_idx is not None:
                def _sprint_boundary(s):
                    return {'id': s.get('id'), 'name': s.get('name'),
                            'startDate': s.get('startDate'), 'endDate': s.get('endDate')}
                sprint_boundaries = {
                    'selected': _sprint_boundary(sorted_sprints[selected_idx]),
                    'previous': _sprint_boundary(sorted_sprints[selected_idx - 1]) if selected_idx > 0 else None,
                    'next': _sprint_boundary(sorted_sprints[selected_idx + 1]) if selected_idx < len(sorted_sprints) - 1 else None,
                }
        start_date = parse_iso_date(config_payload.get('start_date')) or quarter_start or date.today()
        quarter_end_date = parse_iso_date(config_payload.get('quarter_end_date')) or quarter_end or (start_date + timedelta(days=90))
        anchor_date = parse_iso_date(config_payload.get('anchor_date')) if config_payload.get('anchor_date') else None

        scenario_config = ScenarioConfig(
            start_date=start_date,
            quarter_end_date=quarter_end_date,
            anchor_date=anchor_date,
            sp_to_weeks=2.0,
            team_sizes={},
            vacation_weeks={},
            sickleave_buffer=0.0,
            wip_limit=1,
            lane_mode=config_payload.get('lane_mode', 'team'),
        )

        headers = None
        team_field_id = resolve_team_field_id(None, context=auth_context)
        epic_link_field_id = resolve_epic_link_field_id(None, context=auth_context)

        fields_list = [
            'summary',
            'status',
            'priority',
            'issuetype',
            'assignee',
            'updated',
            get_story_points_field_id(),
            'parent',
            'startDate',
            'duedate',
            'timetracking',
        ]
        if epic_link_field_id and epic_link_field_id not in fields_list:
            fields_list.append(epic_link_field_id)
        if team_field_id and team_field_id not in fields_list:
            fields_list.append(team_field_id)

        search_query = (filters.get('search') or '').strip().lower()
        team_filter_ids = {t for t in (filters.get('teams') or []) if t}
        scenario_jql = build_scenario_jql(filters)
        issues_raw = fetch_issues_by_jql(scenario_jql, fields_list)

        issues = []
        issue_keys = []
        issue_by_key = {}
        team_names = set()
        epic_keys = set()
        for issue in issues_raw:
            fields = issue.get('fields', {}) or {}
            raw_team = None
            if team_field_id and fields.get(team_field_id) is not None:
                raw_team = fields.get(team_field_id)
            team_name = extract_team_name(raw_team) if raw_team is not None else None
            team_id = None
            if raw_team is not None:
                team_value = build_team_value(raw_team)
                if isinstance(team_value, dict):
                    team_id = team_value.get('id')

            epic_key = None
            if epic_link_field_id and fields.get(epic_link_field_id):
                epic_key = fields.get(epic_link_field_id)
            elif fields.get('parent') and fields['parent'].get('key') and \
                    fields['parent'].get('fields', {}).get('issuetype', {}).get('name', '').lower() == 'epic':
                epic_key = fields['parent'].get('key')

            assignee = fields.get('assignee') or {}
            issue_type = (fields.get('issuetype') or {}).get('name') or ''
            story_points = fields.get(get_story_points_field_id())
            priority = (fields.get('priority') or {}).get('name')
            status = (fields.get('status') or {}).get('name')
            jira_start_date = fields.get('startDate')   # ISO string or None
            jira_due_date = fields.get('duedate')       # ISO string or None
            time_tracking = fields.get('timetracking') or {}
            time_spent_seconds = time_tracking.get('timeSpentSeconds')
            issue_obj = Issue(
                key=issue.get('key'),
                summary=fields.get('summary') or '',
                issue_type=issue_type,
                team=team_name,
                assignee=assignee.get('displayName'),
                story_points=story_points,
                priority=priority,
                status=status,
                epic_key=epic_key,
                team_id=team_id,
            )
            if issue_obj.key:
                issues.append(issue_obj)
                issue_keys.append(issue_obj.key)
                if team_name:
                    team_names.add(team_name)
                if epic_key:
                    epic_keys.add(epic_key)
                issue_by_key[issue_obj.key] = {
                    'key': issue_obj.key,
                    'summary': issue_obj.summary,
                    'type': issue_type,
                    'team': team_name,
                    'team_id': team_id,
                    'assignee': issue_obj.assignee,
                    'sp': story_points,
                    'priority': priority,
                    'status': status,
                    'epicKey': epic_key,
                    'jiraStartDate': jira_start_date,
                    'jiraDueDate': jira_due_date,
                    'timeSpentSeconds': time_spent_seconds,
                }

        dependencies = collect_dependencies(issue_keys)
        dependency_edges = {}
        edge_list = []
        edge_set = set()
        dependency_snapshots = {}
        for deps in dependencies.values():
            for dep in deps:
                if dep.get('key'):
                    dependency_snapshots[dep['key']] = dep

        for issue_key, deps in dependencies.items():
            for dep in deps:
                prereq_key = dep.get('prereqKey')
                dependent_key = dep.get('dependentKey')
                if not prereq_key or not dependent_key:
                    continue
                category = dep.get('category')
                edge_type = 'dependency' if category == 'dependency' else 'block' if category == 'block' else None
                if edge_type is None:
                    continue
                if prereq_key == dependent_key:
                    continue
                edge_id = (prereq_key, dependent_key, edge_type)
                if edge_id in edge_set:
                    continue
                edge_set.add(edge_id)
                edge_list.append({'from': prereq_key, 'to': dependent_key, 'type': edge_type})
                if edge_type in ('dependency', 'block'):
                    dependency_edges.setdefault(dependent_key, []).append(prereq_key)

        def matches_search(entry):
            if not search_query:
                return True
            key = (entry.get('key') or '').lower()
            summary = (entry.get('summary') or '').lower()
            return search_query in key or search_query in summary

        focus_keys = [
            key for key, entry in issue_by_key.items()
            if matches_search(entry) and (not team_filter_ids or entry.get('team_id') in team_filter_ids)
        ]

        adjacency = {}
        for edge in edge_list:
            adjacency.setdefault(edge['from'], set()).add(edge['to'])
            adjacency.setdefault(edge['to'], set()).add(edge['from'])

        focus_set = set(focus_keys)
        context_keys = set()
        for key in focus_set:
            context_keys.update(adjacency.get(key, set()))
        context_keys -= focus_set

        included_keys = focus_set | context_keys
        if not focus_set and not search_query:
            included_keys = set(issue_by_key.keys())
            focus_set = set(included_keys)

        for key in context_keys:
            if key in issue_by_key:
                continue
            snapshot = dependency_snapshots.get(key)
            if not snapshot:
                continue
            issue_by_key[key] = {
                'key': snapshot.get('key'),
                'summary': snapshot.get('summary') or '',
                'type': snapshot.get('issuetype'),
                'team': snapshot.get('teamName'),
                'team_id': snapshot.get('teamId'),
                'assignee': None,
                'sp': snapshot.get('storyPoints'),
                'priority': snapshot.get('priority'),
                'status': snapshot.get('status'),
                'epicKey': snapshot.get('epicKey'),
            }
            if snapshot.get('teamName'):
                team_names.add(snapshot.get('teamName'))
            if snapshot.get('epicKey'):
                epic_keys.add(snapshot.get('epicKey'))

        capacity_details = {}
        if sprint_label:
            capacity_keys = {}
            for name in team_names:
                normalized = normalize_capacity_team_name(name)
                if normalized:
                    capacity_keys[name] = normalized
            capacity_sizes, capacity_details = fetch_capacity_team_sizes(
                sprint_label,
                None,
                team_names=sorted(set(capacity_keys.values()))
            )
            scenario_config.team_sizes = {
                name: capacity_sizes.get(norm)
                for name, norm in capacity_keys.items()
                if capacity_sizes.get(norm) is not None
            }

        epic_summary_by_key = {}
        if epic_keys:
            epic_issues = fetch_issues_by_keys(sorted(epic_keys), ['summary'])
            for epic in epic_issues:
                fields = epic.get('fields') or {}
                epic_summary_by_key[epic.get('key')] = fields.get('summary')

        jira_base_url = auth_context.site_url or (JIRA_URL or '').rstrip('/')

        capacity_by_team = {}
        if sprint_label:
            for team_name in sorted(team_names):
                normalized = normalize_capacity_team_name(team_name)
                detail = capacity_details.get(normalized) if normalized else None
                size = detail.get('watchers') if detail else None
                capacity_by_team[team_name] = {
                    'size': size,
                    'capacityIssueKey': detail.get('issue_key') if detail else None,
                    'watchersCount': detail.get('watchers') if detail else None,
                    'devLead': detail.get('reporter') if detail else None
                }

        # Separate excluded-capacity issues (Ad Hoc, Interrupt, DevLead, etc.)
        # from the scheduling pipeline.  They get fixed sprint-window dates
        # instead of consuming assignee time slots.
        excluded_capacity_epics_raw = config_payload.get('excluded_capacity_epics') or []
        excluded_epic_set = {str(k).strip().upper() for k in excluded_capacity_epics_raw if k}
        log_info(f'[Scenario] excluded_capacity_epics from config: {excluded_capacity_epics_raw}')
        log_info(f'[Scenario] excluded_epic_set (normalized): {excluded_epic_set}')

        issue_objs = []
        excluded_issue_entries = []
        for key in included_keys:
            entry = issue_by_key.get(key)
            if not entry:
                continue
            epic = (entry.get('epicKey') or '').strip().upper()
            if excluded_epic_set and epic in excluded_epic_set:
                excluded_issue_entries.append(entry)
                continue
            issue_objs.append(Issue(
                key=entry.get('key'),
                summary=entry.get('summary') or '',
                issue_type=entry.get('type') or '',
                team=entry.get('team'),
                assignee=entry.get('assignee'),
                story_points=entry.get('sp'),
                priority=entry.get('priority'),
                status=entry.get('status'),
                epic_key=entry.get('epicKey'),
                team_id=entry.get('team_id'),
            ))

        log_info(f'[Scenario] excluded {len(excluded_issue_entries)} issues, scheduling {len(issue_objs)} regular issues')
        if excluded_issue_entries:
            log_info(f'[Scenario] excluded issue keys: {[e.get("key") for e in excluded_issue_entries]}')

        scheduled_list, scheduled_map = schedule_issues(issue_objs, dependency_edges, scenario_config)

        # Give excluded-capacity issues SP-proportional durations within the sprint.
        # Each starts at sprint start with width reflecting its SP — they run in
        # parallel (background capacity), not sequentially.
        total_weeks = max(1.0, (quarter_end_date - start_date).days / 7.0)
        for entry in excluded_issue_entries:
            sp = entry.get('sp')
            sp_val = float(sp) if sp is not None else 0.0
            duration_weeks = max(0.5, sp_val * scenario_config.sp_to_weeks) if sp_val > 0 else total_weeks
            exc_start = start_date
            exc_end = start_date + timedelta(weeks=duration_weeks)
            if exc_end > quarter_end_date:
                exc_end = quarter_end_date
            item = ScheduledIssue(
                key=entry.get('key'),
                summary=entry.get('summary') or '',
                lane=entry.get('team') or 'Unassigned',
                start_date=exc_start,
                end_date=exc_end,
                blocked_by=[],
                scheduled_reason='excluded_capacity',
                duration_weeks=duration_weeks,
                assignee=entry.get('assignee'),
            )
            scheduled_list.append(item)
            scheduled_map[entry.get('key')] = item
        scheduled_by_key = {item.key: item for item in scheduled_list}
        slack, critical = compute_slack(scheduled_map, dependency_edges, scenario_config.quarter_end_date)
        if app.debug:
            blocked_edges = [edge for edge in edge_list if edge.get('type') == 'block']
            for edge in blocked_edges[:20]:
                prereq_key = edge.get('from')
                dependent_key = edge.get('to')
                assert prereq_key != dependent_key
                prereq_item = scheduled_by_key.get(prereq_key)
                dependent_item = scheduled_by_key.get(dependent_key)
                prereq_start = prereq_item.start_date.isoformat() if prereq_item and prereq_item.start_date else None
                prereq_end = prereq_item.end_date.isoformat() if prereq_item and prereq_item.end_date else None
                dependent_start = dependent_item.start_date.isoformat() if dependent_item and dependent_item.start_date else None
                dependent_end = dependent_item.end_date.isoformat() if dependent_item and dependent_item.end_date else None
                log_debug(
                    "scenario blocked_by edge timing",
                    {
                        "prereqScheduled": bool(prereq_item),
                        "dependentScheduled": bool(dependent_item),
                        "prereqStart": prereq_start,
                        "prereqEnd": prereq_end,
                        "dependentStart": dependent_start,
                        "dependentEnd": dependent_end,
                    },
                )

        total_weeks = max(1.0, (scenario_config.quarter_end_date - scenario_config.start_date).days / 7.0)
        lane_usage = {}
        for item in scheduled_list:
            if item.duration_weeks is None:
                continue
            lane_usage[item.lane] = lane_usage.get(item.lane, 0.0) + item.duration_weeks

        bottleneck_lanes = sorted(lane_usage.keys(), key=lambda lane: lane_usage[lane], reverse=True)[:3]
        late_items = []
        unschedulable = []

        for item in scheduled_list:
            if item.key in slack:
                item.slack_weeks = slack[item.key]
                item.is_critical = item.key in critical
            if item.end_date and item.end_date > scenario_config.quarter_end_date:
                item.is_late = True
                late_items.append(item.key)
            if item.scheduled_reason != 'scheduled' and item.scheduled_reason != 'already_done':
                unschedulable.append(item.key)

        response_issues = []
        for key in sorted(included_keys):
            entry = issue_by_key.get(key)
            item = scheduled_by_key.get(key)
            if not entry:
                continue
            epic_key = entry.get('epicKey')
            response_issues.append({
                'key': key,
                'summary': entry.get('summary'),
                'type': entry.get('type'),
                'team': entry.get('team'),
                'team_id': entry.get('team_id'),
                'assignee': entry.get('assignee'),
                'sp': entry.get('sp'),
                'priority': entry.get('priority'),
                'status': entry.get('status'),
                'epicKey': epic_key,
                'epicSummary': epic_summary_by_key.get(epic_key),
                'start': item.start_date.isoformat() if item and item.start_date else None,
                'end': item.end_date.isoformat() if item and item.end_date else None,
                'blockedBy': item.blocked_by if item else [],
                'scheduledReason': item.scheduled_reason if item else 'context_only',
                'durationWeeks': item.duration_weeks if item else None,
                'slackWeeks': item.slack_weeks if item else None,
                'progressPct': item.progress_pct if item and item.progress_pct is not None else None,
                'isCritical': item.is_critical if item else False,
                'isLate': item.is_late if item else False,
                'isContext': key in context_keys,
                'url': f'{jira_base_url}/browse/{key}' if jira_base_url else None,
                'jiraStartDate': entry.get('jiraStartDate'),
                'jiraDueDate': entry.get('jiraDueDate'),
                'timeSpentSeconds': entry.get('timeSpentSeconds'),
            })

        result = {
            'generatedAt': datetime.now().isoformat(),
            'jira_base_url': jira_base_url,
            'config': {
                'start_date': scenario_config.start_date.isoformat(),
                'quarter_end_date': scenario_config.quarter_end_date.isoformat(),
                'sp_to_weeks': scenario_config.sp_to_weeks,
                'wip_limit': scenario_config.wip_limit,
                'sickleave_buffer': scenario_config.sickleave_buffer,
                'lane_mode': scenario_config.lane_mode,
                'sprint': sprint_label
            },
            'summary': {
                'critical_path': critical,
                'bottleneck_lanes': bottleneck_lanes,
                'late_items': late_items,
                'unschedulable': unschedulable,
                'deadline_met': len(late_items) == 0 and len(unschedulable) == 0,
            },
            'issues': response_issues,
            'dependencies': [edge for edge in edge_list if edge['from'] in included_keys and edge['to'] in included_keys],
            'capacity_by_team': capacity_by_team,
            'focus_set': {
                'focused_issue_keys': sorted(focus_set),
                'context_issue_keys': sorted(context_keys),
            },
            'sprintBoundaries': sprint_boundaries,
        }

        if cache_enabled:
            with _cache_lock:
                SCENARIO_CACHE['generatedAt'] = result['generatedAt']
                SCENARIO_CACHE['data'] = result

        return jsonify(result)
    except AuthError:
        payload, status = oauth_auth_required_payload()
        return jsonify(payload), status
    except Exception as e:
        logger.exception('Scenario error')
        return jsonify({'error': 'Failed to compute scenario', 'message': str(e)}), 500


def get_scenario_overrides():
    """Return overrides for a given scope_key, or empty overrides."""
    scope_key = request.args.get('scope_key', '').strip()
    if not scope_key:
        return jsonify({'overrides': {}})
    if database_storage_enabled():
        try:
            from backend.scenario_drafts import get_active_draft

            response = get_active_draft(
                scenario_draft_request_auth_context(),
                scope_key,
                legacy_loader=load_scenario_overrides,
            )
            active = response.get('activeDraft') or {}
            return jsonify({
                'overrides': active.get('overrides', {}),
                'activeDraft': response.get('activeDraft'),
                'versions': response.get('versions', []),
                'storage': 'db',
            })
        except AuthError as error:
            return auth_error_response(error, 401)
        except DatabaseConfigurationError:
            return jsonify({
                'error': 'config_storage_unavailable',
                'message': 'Scenario drafts require database-backed configuration storage.',
            }), 503
    data = load_scenario_overrides()
    entry = data.get('scenarios', {}).get(scope_key, {})
    return jsonify({'overrides': entry.get('overrides', {})})


def post_scenario_overrides():
    """Upsert overrides for a scope_key."""
    body = request.get_json(force=True, silent=True) or {}
    scope_key = (body.get('scope_key') or '').strip()
    if database_storage_enabled():
        if JIRA_AUTH_MODE == AUTH_MODE_ATLASSIAN_OAUTH:
            try:
                data = csrf_session_data_for_request()
            except AuthError as error:
                return auth_error_response(error, 401)
            except DatabaseConfigurationError:
                return jsonify({
                    'error': 'config_storage_unavailable',
                    'message': 'Scenario drafts require database-backed configuration storage.',
                }), 503
            if not validate_csrf_token(session, data, request.headers.get('X-CSRF-Token')):
                return jsonify({
                    'error': 'csrf_required',
                    'message': 'A valid CSRF token is required for this request.',
                }), 403
        if not scope_key:
            return jsonify({'error': 'scope_key is required'}), 400
        try:
            from backend.scenario_drafts import get_active_draft

            response = get_active_draft(
                scenario_draft_request_auth_context(),
                scope_key,
            )
        except AuthError as error:
            return auth_error_response(error, 401)
        except DatabaseConfigurationError:
            return jsonify({
                'error': 'config_storage_unavailable',
                'message': 'Scenario drafts require database-backed configuration storage.',
            }), 503
        if body.get('baseDraftRevision') is None:
            return jsonify({
                'error': 'scenario_draft_revision_required',
                'message': 'Scenario draft writes require baseDraftRevision. Use POST /api/scenario/drafts.',
                'activeDraft': response.get('activeDraft'),
                'versions': response.get('versions', []),
                'storage': 'db',
            }), 409
        return jsonify({
            'error': 'scenario_draft_api_required',
            'message': 'Use POST /api/scenario/drafts to save database-backed scenario drafts.',
            'storage': 'db',
        }), 409
    if not scope_key:
        return jsonify({'error': 'scope_key is required'}), 400
    overrides = body.get('overrides', {})
    name = body.get('name', '')
    data = load_scenario_overrides()
    data.setdefault('scenarios', {})[scope_key] = {
        'scope_key': scope_key,
        'name': name,
        'updated_at': utc_now_iso(timespec='seconds'),
        'overrides': overrides,
    }
    save_scenario_overrides(data)
    return jsonify({'ok': True})


def fetch_stats_for_sprint(sprint_name, headers, team_field_id, team_ids=None):
    """Fetch stories for a sprint and aggregate delivery stats by team/project."""
    base_jql = STATS_JQL_BASE or f'project in ("{JIRA_PRODUCT_PROJECT}","{JIRA_TECH_PROJECT}")'
    base_jql = strip_sprint_clause(base_jql)
    stats_team_ids = team_ids or get_stats_team_ids()
    if stats_team_ids and not re.search(r'"Team\[Team\]"\s+in\s*\(', base_jql, flags=re.IGNORECASE) and \
            not re.search(r'"Team\[Team\]"\s*=\s*', base_jql, flags=re.IGNORECASE):
        if len(stats_team_ids) == 1:
            base_jql = add_clause_to_jql(base_jql, f'"Team[Team]" = "{stats_team_ids[0]}"')
        else:
            quoted = ', '.join(f'"{team_id}"' for team_id in stats_team_ids)
            base_jql = add_clause_to_jql(base_jql, f'"Team[Team]" in ({quoted})')
    jql = add_clause_to_jql(base_jql, f'Sprint in ("{sprint_name}")')
    if STATS_JQL_ORDER_BY and not re.search(r'order\s+by', jql, flags=re.IGNORECASE):
        jql = f"{jql} {STATS_JQL_ORDER_BY}"

    fields_list = [
        'status',
        'project',
        'priority',
        get_story_points_field_id()  # Story Points
    ]
    if team_field_id and team_field_id not in fields_list:
        fields_list.append(team_field_id)

    page_size = 250
    next_page_token = None
    collected_issues = []
    total_issues = None

    while True:
        payload = {
            'jql': jql,
            'maxResults': page_size,
            'fields': fields_list
        }
        if next_page_token:
            payload['nextPageToken'] = next_page_token

        response = jira_search_request(payload)
        if response.status_code != 200:
            return None, response

        data = response.json()
        total_issues = data.get('total', total_issues)
        issues = data.get('issues', [])
        if not issues:
            break

        collected_issues.extend(issues)
        next_page_token = data.get('nextPageToken')
        if data.get('isLast', not next_page_token) or not next_page_token:
            break

    def normalize_status(value):
        return (value or '').strip().lower()

    def parse_points(value):
        try:
            if value is None:
                return 0.0
            return float(value)
        except Exception:
            return 0.0

    teams = {}
    projects_summary = {}
    totals = {
        'done': 0,
        'incomplete': 0,
        'killed': 0,
        'donePoints': 0.0,
        'incompletePoints': 0.0
    }

    for issue in collected_issues:
        fields = issue.get('fields', {}) or {}
        status_name = (fields.get('status') or {}).get('name', '')
        status_value = normalize_status(status_name)
        is_done = status_value == 'done'
        is_killed = status_value == 'killed'
        priority_name = (fields.get('priority') or {}).get('name', '') or 'Unspecified'

        points = parse_points(fields.get(get_story_points_field_id()))
        project_name = (fields.get('project') or {}).get('name')
        project_key = (fields.get('project') or {}).get('key')
        project_label = project_name or project_key or 'Unknown Project'
        project_bucket = classify_project(project_label, project_key)

        raw_team = None
        if team_field_id and fields.get(team_field_id) is not None:
            raw_team = fields.get(team_field_id)
        team_ids = extract_team_ids(raw_team)
        if stats_team_ids:
            if not team_ids:
                continue
            if not any(team_id in stats_team_ids for team_id in team_ids):
                continue

        team_payload = build_team_value(raw_team) if raw_team is not None else {}
        team_id = None
        team_name = None
        if isinstance(team_payload, dict):
            team_id = team_payload.get('id') or (team_ids[0] if team_ids else None)
            team_name = team_payload.get('name')
        if not team_name:
            team_name = extract_team_name(raw_team)

        team_key = team_id or team_name or 'unknown'
        if team_key not in teams:
            teams[team_key] = {
                'id': team_id,
                'name': team_name or 'Unknown Team',
                'done': 0,
                'incomplete': 0,
                'killed': 0,
                'donePoints': 0.0,
                'incompletePoints': 0.0,
                'projects': {},
                'priorities': {}
            }

        team_entry = teams[team_key]
        if priority_name not in team_entry['priorities']:
            team_entry['priorities'][priority_name] = {'done': 0, 'incomplete': 0, 'killed': 0}
        if project_bucket not in team_entry['projects']:
            team_entry['projects'][project_bucket] = {
                'done': 0,
                'incomplete': 0,
                'killed': 0,
                'donePoints': 0.0,
                'incompletePoints': 0.0,
                'priorities': {}
            }

        if project_bucket not in projects_summary:
            projects_summary[project_bucket] = {
                'done': 0,
                'incomplete': 0,
                'killed': 0,
                'donePoints': 0.0,
                'incompletePoints': 0.0
            }

        if is_killed:
            team_entry['killed'] += 1
            team_entry['priorities'][priority_name]['killed'] += 1
            if priority_name not in team_entry['projects'][project_bucket]['priorities']:
                team_entry['projects'][project_bucket]['priorities'][priority_name] = {'done': 0, 'incomplete': 0, 'killed': 0}
            team_entry['projects'][project_bucket]['priorities'][priority_name]['killed'] += 1
            team_entry['projects'][project_bucket]['killed'] += 1
            projects_summary[project_bucket]['killed'] += 1
            totals['killed'] += 1
            continue

        if is_done:
            team_entry['done'] += 1
            team_entry['donePoints'] += points
            team_entry['priorities'][priority_name]['done'] += 1
            if priority_name not in team_entry['projects'][project_bucket]['priorities']:
                team_entry['projects'][project_bucket]['priorities'][priority_name] = {'done': 0, 'incomplete': 0, 'killed': 0}
            team_entry['projects'][project_bucket]['priorities'][priority_name]['done'] += 1
            team_entry['projects'][project_bucket]['done'] += 1
            team_entry['projects'][project_bucket]['donePoints'] += points
            projects_summary[project_bucket]['done'] += 1
            projects_summary[project_bucket]['donePoints'] += points
            totals['done'] += 1
            totals['donePoints'] += points
        else:
            team_entry['incomplete'] += 1
            team_entry['incompletePoints'] += points
            team_entry['priorities'][priority_name]['incomplete'] += 1
            if priority_name not in team_entry['projects'][project_bucket]['priorities']:
                team_entry['projects'][project_bucket]['priorities'][priority_name] = {'done': 0, 'incomplete': 0, 'killed': 0}
            team_entry['projects'][project_bucket]['priorities'][priority_name]['incomplete'] += 1
            team_entry['projects'][project_bucket]['incomplete'] += 1
            team_entry['projects'][project_bucket]['incompletePoints'] += points
            projects_summary[project_bucket]['incomplete'] += 1
            projects_summary[project_bucket]['incompletePoints'] += points
            totals['incomplete'] += 1
            totals['incompletePoints'] += points

    sorted_teams = sorted(
        teams.values(),
        key=lambda t: (t['name'] or '').lower()
    )

    stats_payload = {
        'sprint': sprint_name,
        'totals': totals,
        'projects': projects_summary,
        'teams': sorted_teams
    }
    return stats_payload, None


def get_stats_burnout_timezone():
    if ZoneInfo is None:
        return None
    try:
        return ZoneInfo(STATS_BURNOUT_TIMEZONE)
    except Exception:
        return None


def parse_jira_datetime(value):
    if not value:
        return None
    raw = str(value).strip()
    if not raw:
        return None
    if raw.endswith('Z'):
        raw = raw[:-1] + '+00:00'
    try:
        return datetime.fromisoformat(raw)
    except Exception:
        pass
    for fmt in ('%Y-%m-%dT%H:%M:%S.%f%z', '%Y-%m-%dT%H:%M:%S%z'):
        try:
            return datetime.strptime(raw, fmt)
        except Exception:
            continue
    return None


def normalize_assignee_value(raw_assignee):
    if not raw_assignee:
        return {'id': None, 'name': 'Unassigned'}
    if isinstance(raw_assignee, dict):
        return {
            'id': raw_assignee.get('accountId') or raw_assignee.get('id'),
            'name': raw_assignee.get('displayName') or raw_assignee.get('name') or raw_assignee.get('emailAddress') or 'Unassigned'
        }
    return {'id': None, 'name': str(raw_assignee)}


def normalize_team_value_for_burnout(raw_team):
    if not raw_team:
        return {'id': None, 'name': 'Unknown Team'}
    if isinstance(raw_team, list):
        first = raw_team[0] if raw_team else None
        return normalize_team_value_for_burnout(first)
    payload = build_team_value(raw_team)
    if isinstance(payload, dict):
        return {
            'id': payload.get('id'),
            'name': payload.get('name') or extract_team_name(raw_team) or 'Unknown Team'
        }
    return {'id': None, 'name': extract_team_name(raw_team) or str(raw_team)}


def normalize_team_change_value(raw_id, raw_name):
    team_id = str(raw_id or '').strip() or None
    team_name = str(raw_name or '').strip() or None
    if not team_id and not team_name:
        return {'id': None, 'name': 'Unknown Team'}
    return {'id': team_id, 'name': team_name or team_id}


def resolve_sprint_date_bounds(sprint_label, cache_enabled=None):
    if cache_enabled is None:
        cache_enabled = jira_home_process_cache_enabled(_cache_policy_context())
    if cache_enabled:
        cache = load_sprints_cache() or {}
        for sprint in cache.get('sprints', []) or []:
            if str(sprint.get('name') or '').strip() != str(sprint_label or '').strip():
                continue
            start_value = str(sprint.get('startDate') or '')[:10]
            end_value = str(sprint.get('endDate') or '')[:10]
            start_date = parse_iso_date(start_value)
            end_date = parse_iso_date(end_value)
            if start_date and end_date:
                return start_date, end_date
    return quarter_dates_from_label(sprint_label)


def normalize_burnout_status_bucket(status_name):
    normalized = str(status_name or '').strip().lower()
    if normalized == 'done':
        return 'done'
    if normalized == 'killed':
        return 'killed'
    if normalized == 'incomplete':
        return 'incomplete'
    return None


def is_assignee_history_item(item):
    return str((item or {}).get('field') or '').strip().lower() == 'assignee'


def is_status_history_item(item):
    return str((item or {}).get('field') or '').strip().lower() == 'status'


def is_team_history_item(item, team_field_id):
    if not item:
        return False
    field_id = str(item.get('fieldId') or '').strip()
    field_name = str(item.get('field') or '').strip().lower()
    if team_field_id and field_id == str(team_field_id).strip():
        return True
    return field_name in ('team', 'team[team]')


def rewind_burnout_state_from_item(item, team_field_id, current_team, current_assignee):
    if is_assignee_history_item(item):
        from_id = str(item.get('from') or '').strip() or None
        from_name = str(item.get('fromString') or '').strip() or 'Unassigned'
        current_assignee = {'id': from_id, 'name': from_name}
    elif is_team_history_item(item, team_field_id):
        current_team = normalize_team_change_value(item.get('from'), item.get('fromString'))
    return current_team, current_assignee


def resolve_team_state_at_date(current_team, current_assignee, histories_desc, team_field_id, tz_info, target_date):
    if not target_date:
        return {
            'team': {'id': current_team.get('id'), 'name': current_team.get('name') or 'Unknown Team'},
            'assignee': {'id': current_assignee.get('id'), 'name': current_assignee.get('name') or 'Unassigned'}
        }
    team_state = {'id': current_team.get('id'), 'name': current_team.get('name') or 'Unknown Team'}
    assignee_state = {'id': current_assignee.get('id'), 'name': current_assignee.get('name') or 'Unassigned'}
    for created_dt, history in histories_desc:
        local_dt = created_dt.astimezone(tz_info) if tz_info else created_dt
        local_date = local_dt.date()
        if local_date <= target_date:
            break
        for item in (history.get('items') or []):
            team_state, assignee_state = rewind_burnout_state_from_item(item, team_field_id, team_state, assignee_state)
    return {'team': team_state, 'assignee': assignee_state}


def extract_burnout_events_from_issue(
        issue,
        team_field_id,
        sprint_start,
        sprint_end,
        tz_info,
        include_post_sprint_closures=False
):
    fields = issue.get('fields') or {}
    issue_key = issue.get('key')
    issue_created_dt_raw = parse_jira_datetime(fields.get('created'))
    issue_created_dt = issue_created_dt_raw.astimezone(tz_info) if (issue_created_dt_raw and tz_info) else issue_created_dt_raw
    issue_created_date = issue_created_dt.date() if issue_created_dt else None

    current_assignee = normalize_assignee_value(fields.get('assignee'))
    current_team = normalize_team_value_for_burnout(fields.get(team_field_id)) if team_field_id else {'id': None, 'name': 'Unknown Team'}

    raw_histories = (issue.get('changelog') or {}).get('histories') or []
    histories = []
    for history in raw_histories:
        created_dt = parse_jira_datetime(history.get('created'))
        if not created_dt:
            continue
        histories.append((created_dt, history))
    histories.sort(key=lambda item: item[0], reverse=True)

    snapshot_at_start = resolve_team_state_at_date(
        current_team,
        current_assignee,
        histories,
        team_field_id,
        tz_info,
        sprint_start
    ) if sprint_start else {'team': current_team, 'assignee': current_assignee}
    snapshot_at_created = resolve_team_state_at_date(
        current_team,
        current_assignee,
        histories,
        team_field_id,
        tz_info,
        issue_created_date
    ) if issue_created_date else {'team': current_team, 'assignee': current_assignee}

    events = []
    for created_dt, history in histories:
        # Current state at this step represents issue values immediately after this history entry.
        event_dt = created_dt.astimezone(tz_info) if tz_info else created_dt
        event_date = event_dt.date()

        items = history.get('items') or []
        for item in items:
            if not is_status_history_item(item):
                continue
            bucket = normalize_burnout_status_bucket(item.get('toString'))
            if not bucket:
                continue
            if sprint_start and event_date < sprint_start:
                continue
            if sprint_end and event_date > sprint_end and not include_post_sprint_closures:
                continue
            events.append({
                'issueKey': issue_key,
                'date': event_date.isoformat(),
                'status': str(item.get('toString') or '').strip(),
                'bucket': bucket,
                'teamId': current_team.get('id'),
                'teamName': current_team.get('name') or 'Unknown Team',
                'assigneeId': current_assignee.get('id'),
                'assigneeName': current_assignee.get('name') or 'Unassigned'
            })

        # Roll state back to "before history" so older events are attributed with the correct values.
        for item in items:
            current_team, current_assignee = rewind_burnout_state_from_item(
                item,
                team_field_id,
                current_team,
                current_assignee
            )

    issue_meta = {
        'issueKey': issue_key,
        'createdDate': issue_created_date.isoformat() if issue_created_date else None,
        'teamAtStart': snapshot_at_start.get('team') or {'id': None, 'name': 'Unknown Team'},
        'teamAtCreated': snapshot_at_created.get('team') or {'id': None, 'name': 'Unknown Team'},
        'assignee': normalize_assignee_value(fields.get('assignee'))
    }

    return {
        'events': events,
        'issue': issue_meta
    }


def fetch_burnout_events_for_sprint(
        sprint_name,
        headers,
        team_field_id,
        team_ids=None,
        issue_keys=None,
        include_post_sprint_closures=False,
        cache_enabled=None
):
    base_jql = STATS_JQL_BASE or f'project in ("{JIRA_PRODUCT_PROJECT}","{JIRA_TECH_PROJECT}")'
    base_jql = strip_sprint_clause(base_jql)

    scoped_team_ids = normalize_team_ids(team_ids or [])
    if scoped_team_ids and not re.search(r'"Team\[Team\]"\s+in\s*\(', base_jql, flags=re.IGNORECASE) and \
            not re.search(r'"Team\[Team\]"\s*=\s*', base_jql, flags=re.IGNORECASE):
        if len(scoped_team_ids) == 1:
            base_jql = add_clause_to_jql(base_jql, f'"Team[Team]" = "{scoped_team_ids[0]}"')
        else:
            quoted = ', '.join(f'"{team_id}"' for team_id in scoped_team_ids)
            base_jql = add_clause_to_jql(base_jql, f'"Team[Team]" in ({quoted})')

    configured_issue_types = [item for item in get_configured_issue_types() if item]
    if configured_issue_types:
        escaped = ', '.join('"{}"'.format(str(name).replace('"', '\\"')) for name in configured_issue_types)
        base_jql = add_clause_to_jql(base_jql, f'issuetype in ({escaped})')

    jql = add_clause_to_jql(base_jql, f'Sprint in ("{sprint_name}")')
    jql = add_clause_to_jql(jql, 'status CHANGED TO ("Done","Killed","Incomplete")')
    if not re.search(r'order\s+by', jql, flags=re.IGNORECASE):
        jql = f'{jql} ORDER BY updated DESC'

    fields_list = ['status', 'assignee', 'created']
    if team_field_id and team_field_id not in fields_list:
        fields_list.append(team_field_id)

    sprint_start, sprint_end = resolve_sprint_date_bounds(sprint_name, cache_enabled=cache_enabled)
    timezone_info = get_stats_burnout_timezone()
    collected_issues = []
    normalized_issue_keys = []
    seen_issue_keys = set()
    for key in issue_keys or []:
        value = str(key or '').strip().upper()
        if not value or value in seen_issue_keys:
            continue
        seen_issue_keys.add(value)
        normalized_issue_keys.append(value)

    debug_payload = {
        'jql': jql,
        'fields': fields_list,
        'scopedTeamIds': scoped_team_ids
    }

    if normalized_issue_keys:
        # UI already scopes keys to the selected sprint/team filter.
        # Phase 1: fetch base fields only for all keys (fast, no changelog expansion).
        # Phase 2: fetch changelog only for keys that changed to closed buckets in sprint window.
        chunk_size = 100
        debug_payload['issueKeys'] = len(normalized_issue_keys)
        debug_payload['mode'] = 'keys-two-phase'
        issue_map = {}

        for start in range(0, len(normalized_issue_keys), chunk_size):
            chunk = normalized_issue_keys[start:start + chunk_size]
            quoted_keys = ','.join(chunk)
            chunk_jql = f'issueKey in ({quoted_keys})'
            payload = {
                'jql': chunk_jql,
                'maxResults': len(chunk),
                'fields': fields_list
            }
            response = jira_search_request(payload)
            if response.status_code != 200:
                return None, response, payload
            data = response.json() or {}
            issues = data.get('issues') or []
            if issues:
                for issue in issues:
                    key = str(issue.get('key') or '').strip().upper()
                    if key:
                        issue_map[key] = issue

        closure_clause = 'status CHANGED TO ("Done","Killed","Incomplete")'
        if sprint_start:
            closure_clause += f' AFTER "{sprint_start.isoformat()}"'
        if sprint_end and not include_post_sprint_closures:
            closure_clause += f' BEFORE "{(sprint_end + timedelta(days=1)).isoformat()}"'

        changelog_hits = 0
        for start in range(0, len(normalized_issue_keys), chunk_size):
            chunk = normalized_issue_keys[start:start + chunk_size]
            quoted_keys = ','.join(chunk)
            chunk_scope_jql = f'issueKey in ({quoted_keys})'
            chunk_jql = add_clause_to_jql(chunk_scope_jql, closure_clause)
            payload = {
                'jql': chunk_jql,
                'maxResults': len(chunk),
                'fields': fields_list,
                'expand': ['changelog']
            }
            response = jira_search_request(payload)
            if response.status_code != 200:
                return None, response, payload
            data = response.json() or {}
            issues = data.get('issues') or []
            if issues:
                changelog_hits += len(issues)
                for issue in issues:
                    key = str(issue.get('key') or '').strip().upper()
                    if key:
                        issue_map[key] = issue

        debug_payload['changelogCandidates'] = changelog_hits
        collected_issues = list(issue_map.values())
    else:
        page_size = 100
        next_page_token = None
        total_issues = None
        debug_payload['mode'] = 'jql'
        while True:
            payload = {
                'jql': jql,
                'maxResults': page_size,
                'fields': fields_list,
                'expand': ['changelog']
            }
            if next_page_token:
                payload['nextPageToken'] = next_page_token
            response = jira_search_request(payload)
            if response.status_code != 200:
                return None, response, payload
            data = response.json() or {}
            total_issues = data.get('total', total_issues)
            issues = data.get('issues') or []
            if not issues:
                break
            collected_issues.extend(issues)
            next_page_token = data.get('nextPageToken')
            if data.get('isLast', not next_page_token) or not next_page_token:
                break

    events = []
    issues_meta = []
    for issue in collected_issues:
        parsed = extract_burnout_events_from_issue(
            issue,
            team_field_id,
            sprint_start,
            sprint_end,
            timezone_info,
            include_post_sprint_closures=include_post_sprint_closures
        )
        issue_meta = parsed.get('issue') if isinstance(parsed, dict) else None
        issue_events = parsed.get('events') if isinstance(parsed, dict) else []
        if issue_meta:
            issues_meta.append(issue_meta)
        if scoped_team_ids:
            issue_events = [event for event in issue_events if event.get('teamId') in scoped_team_ids]
            if issue_meta and issue_meta.get('teamAtStart', {}).get('id') not in scoped_team_ids:
                issue_meta['outOfScopeStartTeam'] = True
        events.extend(issue_events)

    if (sprint_start is None or sprint_end is None) and events:
        parsed_dates = [parse_iso_date(event.get('date')) for event in events]
        parsed_dates = [d for d in parsed_dates if d]
        if parsed_dates:
            if sprint_start is None:
                sprint_start = min(parsed_dates)
            if sprint_end is None:
                sprint_end = max(parsed_dates)

    assignee_counts = {}
    for issue in issues_meta:
        assignee = issue.get('assignee') or {}
        key = assignee.get('id') or assignee.get('name') or 'unassigned'
        if key not in assignee_counts:
            assignee_counts[key] = {
                'id': assignee.get('id'),
                'name': assignee.get('name') or 'Unassigned',
                'events': 0
            }
        assignee_counts[key]['events'] += 1

    assignees = sorted(
        assignee_counts.values(),
        key=lambda item: (str(item.get('name') or '').lower(), str(item.get('id') or ''))
    )

    payload = {
        'sprint': sprint_name,
        'timezone': STATS_BURNOUT_TIMEZONE,
        'range': {
            'startDate': sprint_start.isoformat() if sprint_start else None,
            'endDate': sprint_end.isoformat() if sprint_end else None
        },
        'issues': len(collected_issues),
        'events': events,
        'issuesMeta': issues_meta,
        'assignees': assignees
    }
    return payload, None, debug_payload


def _cohort_project_scope():
    projects = [str(item or '').strip() for item in get_selected_projects() if str(item or '').strip()]
    if not projects:
        projects = [str(item or '').strip() for item in (STATS_PRODUCT_PROJECTS + STATS_TECH_PROJECTS) if str(item or '').strip()]
    deduped = []
    seen = set()
    for project in projects:
        if project in seen:
            continue
        seen.add(project)
        deduped.append(project)
    return deduped


def _cohort_parse_bool(value):
    if isinstance(value, bool):
        return value
    return str(value or '').strip().lower() in ('1', 'true', 'yes', 'on')


def _escape_jql_literal(value):
    return str(value or '').replace('\\', '\\\\').replace('"', '\\"')


def _cohort_fetch_terminal_date_from_changelog(issue_key, target_status, headers, context=None):
    response = current_jira_get(
        f'/rest/api/3/issue/{issue_key}',
        params={'fields': 'status', 'expand': 'changelog'},
        timeout=20,
        context=context,
    )
    if response.status_code != 200:
        return issue_key, None, f'changelog fetch failed ({response.status_code})'
    data = response.json() or {}
    changelog = (data.get('changelog') or {}).get('histories') or []
    resolved = resolve_terminal_date_from_history(changelog, target_status)
    if not resolved:
        return issue_key, None, 'terminal transition not found in changelog'
    return issue_key, resolved, None


def fetch_epic_cohort_data(start_quarter, headers, team_field_id, team_ids=None, component_names=None, context=None):
    start_date, _ = quarter_dates_from_label(start_quarter)
    if not start_date:
        return {
            'range': {'startDate': None, 'endDate': None},
            'issues': [],
            'meta': {
                'warnings': ['invalid startQuarter'],
                'truncated': False,
                'paginationMode': 'nextPageToken/isLast'
            }
        }, None

    scoped_projects = _cohort_project_scope()
    if not scoped_projects:
        return {
            'range': {'startDate': start_date.isoformat(), 'endDate': start_date.isoformat()},
            'issues': [],
            'meta': {
                'warnings': ['no projects configured for cohort query'],
                'truncated': False,
                'paginationMode': 'nextPageToken/isLast'
            }
        }, None

    escaped_projects = ', '.join(f'"{_escape_jql_literal(project)}"' for project in scoped_projects)
    jql = f'issuetype = Epic AND project in ({escaped_projects}) AND created >= "{start_date.isoformat()}"'

    scoped_team_ids = normalize_team_ids(team_ids or [])
    scoped_components = [str(name or '').strip() for name in (component_names or []) if str(name or '').strip()]
    scope_clause = build_missing_info_scope_clause(scoped_team_ids, scoped_components, team_field_name='Team[Team]')
    if scope_clause:
        jql = add_clause_to_jql(jql, scope_clause)

    fields = ['summary', 'created', 'status', 'resolutiondate', 'assignee', 'project']
    if team_field_id and team_field_id not in fields:
        fields.append(team_field_id)

    warnings = []
    all_issues = []
    next_page_token = None
    page_count = 0
    max_pages = 80
    truncated = False

    while True:
        payload = {
            'jql': jql,
            'fields': fields,
            'maxResults': 100
        }
        if next_page_token:
            payload['nextPageToken'] = next_page_token

        response = jira_search_request(payload)
        if response.status_code != 200:
            return None, response
        data = response.json() or {}
        issues = data.get('issues') or []
        if issues:
            all_issues.extend(issues)

        page_count += 1
        is_last = bool(data.get('isLast', True))
        next_page_token = data.get('nextPageToken')
        if is_last or not next_page_token:
            break
        if page_count >= max_pages:
            truncated = True
            warnings.append(f'result truncated at {max_pages} pages')
            break

    terminal_candidates = []
    for issue in all_issues:
        fields_data = issue.get('fields') or {}
        raw_status_name = str((fields_data.get('status') or {}).get('name') or '').strip()
        status_name = normalize_epic_status(raw_status_name)
        if not is_terminal_epic_status(status_name):
            continue
        if fields_data.get('resolutiondate'):
            continue
        terminal_candidates.append((str(issue.get('key') or '').strip(), status_name))

    resolved_terminal_dates = {}
    if terminal_candidates:
        max_targets = max(1, int(EPIC_COHORT_ENRICH_MAX_ISSUES))
        workers = max(1, int(EPIC_COHORT_ENRICH_WORKERS))
        timeout_budget = max(1.0, float(EPIC_COHORT_ENRICH_TIMEOUT_SECONDS))
        if len(terminal_candidates) > max_targets:
            warnings.append(f'changelog enrichment capped at {max_targets} issues')
            terminal_candidates = terminal_candidates[:max_targets]
            truncated = True

        future_map = {}
        timed_out = []
        with ThreadPoolExecutor(max_workers=workers) as pool:
            for issue_key, status_name in terminal_candidates:
                if not issue_key:
                    continue
                future = pool.submit(_cohort_fetch_terminal_date_from_changelog, issue_key, status_name, headers, context)
                future_map[future] = issue_key
            try:
                for future in as_completed(future_map, timeout=timeout_budget):
                    issue_key, resolved_date, warning = future.result()
                    if resolved_date:
                        resolved_terminal_dates[issue_key] = resolved_date
                    if warning:
                        warnings.append(f'{issue_key}: {warning}')
            except FuturesTimeoutError:
                warnings.append('changelog enrichment timeout budget exceeded')
                truncated = True
            for future, issue_key in future_map.items():
                if future.done():
                    continue
                future.cancel()
                timed_out.append(issue_key)
        if timed_out:
            warnings.append(f'changelog enrichment timed out for {len(timed_out)} issues')

    today = date.today()
    current_quarter = ((today.month - 1) // 3) + 1
    _, current_quarter_end = quarter_dates_from_label(f'{today.year}Q{current_quarter}')
    range_end = current_quarter_end or today
    latest_terminal_date = None

    normalized_issues = []
    for issue in all_issues:
        fields_data = issue.get('fields') or {}
        issue_key = str(issue.get('key') or '').strip()
        if not issue_key:
            continue

        created_dt = parse_jira_datetime(fields_data.get('created'))
        created_date = created_dt.date() if created_dt else None
        if not created_date:
            continue

        status_name = normalize_epic_status((fields_data.get('status') or {}).get('name'))
        resolution_dt = parse_jira_datetime(fields_data.get('resolutiondate'))
        terminal_date = resolution_dt.date() if resolution_dt else None
        if not terminal_date and is_terminal_epic_status(status_name):
            terminal_date = resolved_terminal_dates.get(issue_key)

        lead_time_days = None
        if terminal_date:
            lead_time_days = max(0, (terminal_date - created_date).days)
            if latest_terminal_date is None or terminal_date > latest_terminal_date:
                latest_terminal_date = terminal_date
        elif status_name == 'open':
            lead_time_days = max(0, (today - created_date).days)

        project = fields_data.get('project') or {}
        raw_team = fields_data.get(team_field_id) if team_field_id else None
        team_payload = normalize_team_value_for_burnout(raw_team)
        assignee_payload = normalize_assignee_value(fields_data.get('assignee'))

        normalized_issues.append({
            'key': issue_key,
            'summary': fields_data.get('summary') or '',
            'projectKey': project.get('key') or '',
            'team': team_payload,
            'assignee': assignee_payload,
            'createdDate': created_date.isoformat(),
            'terminalDate': terminal_date.isoformat() if terminal_date else None,
            'status': status_name,
            'jiraStatus': raw_status_name or status_name,
            'leadTimeDays': lead_time_days,
            'createdQuarter': assign_to_period(created_date, 'quarter'),
            'createdMonth': assign_to_period(created_date, 'month'),
            'terminalDateSource': 'resolutiondate' if resolution_dt else ('changelog' if terminal_date else None)
        })

    if latest_terminal_date and latest_terminal_date > range_end:
        range_end = latest_terminal_date

    payload = {
        'range': {
            'startDate': start_date.isoformat(),
            'endDate': range_end.isoformat()
        },
        'issues': normalized_issues,
        'meta': {
            'warnings': warnings,
            'truncated': bool(truncated),
            'paginationMode': 'nextPageToken/isLast'
        }
    }
    return payload, None


def build_missing_info_scope_clause(team_ids, component_names, team_field_name='Team[Team]'):
    clauses = []
    if isinstance(component_names, str):
        component_names = [component_names] if component_names.strip() else []
    component_names = [str(c).strip() for c in (component_names or []) if c and str(c).strip()]
    team_ids = [str(t).strip() for t in (team_ids or []) if t and str(t).strip()]
    team_field = str(team_field_name or '').strip() or 'Team[Team]'

    if len(component_names) == 1:
        clauses.append(f'component = "{_escape_jql_literal(component_names[0])}"')
    elif len(component_names) > 1:
        quoted = ', '.join(f'"{_escape_jql_literal(c)}"' for c in component_names)
        clauses.append(f'component in ({quoted})')
    if team_ids:
        if len(team_ids) == 1:
            clauses.append(f'"{_escape_jql_literal(team_field)}" = "{_escape_jql_literal(team_ids[0])}"')
        else:
            quoted = ', '.join(f'"{_escape_jql_literal(t)}"' for t in team_ids)
            clauses.append(f'"{_escape_jql_literal(team_field)}" in ({quoted})')

    if not clauses:
        return ''
    if len(clauses) == 1:
        return clauses[0]
    return f"({ ' OR '.join(clauses) })"


def build_excluded_capacity_stats_jql(sprint_ids, team_ids=None):
    base_jql = strip_sprint_clause(build_base_jql())
    scoped_team_ids = normalize_team_ids(team_ids or [])
    if scoped_team_ids and JQL_QUERY_TEMPLATE:
        templated = apply_team_ids_to_template(scoped_team_ids)
        if templated:
            base_jql = strip_sprint_clause(templated)
    elif scoped_team_ids:
        base_jql = remove_team_filter_from_jql(base_jql)
        if len(scoped_team_ids) == 1:
            base_jql = add_clause_to_jql(base_jql, f'"Team[Team]" = "{_escape_jql_literal(scoped_team_ids[0])}"')
        else:
            quoted_teams = ', '.join(f'"{_escape_jql_literal(team_id)}"' for team_id in scoped_team_ids)
            base_jql = add_clause_to_jql(base_jql, f'"Team[Team]" in ({quoted_teams})')

    sprint_values = []
    for sprint_id in sprint_ids:
        text = str(sprint_id or '').strip()
        if not text:
            continue
        sprint_values.append(text if text.isdigit() else f'"{_escape_jql_literal(text)}"')
    base_jql = add_clause_to_jql(base_jql, f'Sprint in ({", ".join(sprint_values)})')

    issue_types = get_configured_issue_types()
    if issue_types:
        if len(issue_types) == 1:
            base_jql = add_clause_to_jql(base_jql, f'type = "{_escape_jql_literal(issue_types[0])}"')
        else:
            quoted_types = ', '.join(f'"{_escape_jql_literal(issue_type)}"' for issue_type in issue_types)
            base_jql = add_clause_to_jql(base_jql, f'type in ({quoted_types})')
    return base_jql


def build_excluded_capacity_issue_payload(issue, team_field_id, epic_link_field_id, sprint_field_id, epic_summary_by_key=None, team_name_by_id=None):
    fields = issue.get('fields', {}) or {}
    raw_team = fields.get(team_field_id) if team_field_id and fields.get(team_field_id) is not None else None
    team_payload = build_team_value(raw_team) if raw_team is not None else {}
    team_id = team_payload.get('id') if isinstance(team_payload, dict) else None
    team_name = team_payload.get('name') if isinstance(team_payload, dict) else None
    if not team_name:
        team_name = extract_team_name(raw_team)
    if (not team_name) and team_id and team_name_by_id:
        resolved = team_name_by_id.get(str(team_id).strip())
        if resolved:
            team_name = resolved
    if isinstance(team_payload, dict) and team_name and not team_payload.get('name'):
        team_payload['name'] = team_name

    epic_key = None
    parent_field = fields.get('parent') or {}
    parent_summary = (parent_field.get('fields') or {}).get('summary')
    if epic_link_field_id and fields.get(epic_link_field_id):
        epic_key = fields.get(epic_link_field_id)
    elif parent_field.get('key') and \
            (parent_field.get('fields') or {}).get('issuetype', {}).get('name', '').lower() == 'epic':
        epic_key = parent_field.get('key')

    epic_summary = ''
    if epic_summary_by_key and epic_key:
        epic_summary = str(epic_summary_by_key.get(epic_key) or '').strip()
    if not epic_summary and epic_key and parent_field.get('key') == epic_key and parent_summary:
        epic_summary = str(parent_summary or '').strip()

    status = fields.get('status') or {}
    priority = fields.get('priority') or {}
    issuetype = fields.get('issuetype') or {}
    assignee = fields.get('assignee') or {}
    project_field = fields.get('project') or {}
    story_points_field = get_story_points_field_id()
    normalized_sprints = normalize_epm_sprint_field(fields.get(sprint_field_id)) if sprint_field_id else []

    return {
        'id': issue.get('id'),
        'key': issue.get('key'),
        'fields': {
            'summary': fields.get('summary') or '',
            'status': {'name': status.get('name')} if status else None,
            'priority': {'name': priority.get('name')} if priority else None,
            'issuetype': {'name': issuetype.get('name')} if issuetype else None,
            'assignee': {'displayName': assignee.get('displayName')} if assignee else None,
            'updated': fields.get('updated'),
            'customfield_10004': fields.get(story_points_field),
            'team': team_payload,
            'teamName': team_name,
            'teamId': team_id,
            'epicKey': epic_key,
            'epicSummary': epic_summary,
            'customfield_10101': normalized_sprints,
            'parentSummary': parent_summary,
            'projectKey': project_field.get('key', ''),
            'projectName': project_field.get('name', '')
        }
    }


def excluded_capacity_epic_summary_cache_key(epic_key, context=None):
    normalized_key = str(epic_key or '').strip().upper()
    if context is not None:
        return build_auth_cache_key(context, 'excluded-capacity-epic-summary', normalized_key)
    return ('excluded-capacity-epic-summary', 'basic', normalized_key)


def fetch_cached_excluded_capacity_epic_summaries(epic_keys, context=None):
    normalized_keys = []
    original_by_normalized = {}
    for key in epic_keys or []:
        original = str(key or '').strip()
        normalized = original.upper()
        if not normalized or normalized in original_by_normalized:
            continue
        original_by_normalized[normalized] = original
        normalized_keys.append(normalized)

    if not normalized_keys:
        return {}

    now = time.time()
    summaries_by_normalized = {}
    missing_keys = []
    with _cache_lock:
        for normalized in normalized_keys:
            cache_key = excluded_capacity_epic_summary_cache_key(normalized, context=context)
            entry = EXCLUDED_CAPACITY_EPIC_SUMMARY_CACHE.get(cache_key)
            if entry and now - entry.get('timestamp', 0) < EXCLUDED_CAPACITY_EPIC_SUMMARY_CACHE_TTL_SECONDS:
                summaries_by_normalized[normalized] = entry.get('summary', '')
            else:
                missing_keys.append(normalized)

    batch_size = max(1, EXCLUDED_CAPACITY_EPIC_SUMMARY_BATCH_SIZE)
    for index in range(0, len(missing_keys), batch_size):
        batch = missing_keys[index:index + batch_size]
        fetched = {}
        epic_records = fetch_issues_by_keys(batch, ['summary'], context=context)
        for epic in epic_records:
            key = str(epic.get('key') or '').strip().upper()
            summary = str((epic.get('fields') or {}).get('summary') or '').strip()
            if key:
                fetched[key] = summary
        with _cache_lock:
            for normalized in batch:
                summary = fetched.get(normalized, '')
                EXCLUDED_CAPACITY_EPIC_SUMMARY_CACHE[excluded_capacity_epic_summary_cache_key(normalized, context=context)] = {
                    'summary': summary,
                    'timestamp': time.time()
                }
                summaries_by_normalized[normalized] = summary

    return {
        original_by_normalized[normalized]: summaries_by_normalized.get(normalized, '')
        for normalized in normalized_keys
    }


def _record_excluded_capacity_timing(timings_ms, key, started):
    if timings_ms is not None:
        timings_ms[key] = round((time.perf_counter() - started) * 1000, 2)


def _excluded_capacity_server_timing_header(timings_ms):
    if not timings_ms:
        return ''
    if 'cache' in timings_ms:
        return 'cache;dur=1'
    token_names = {
        'field_config': 'field-config',
        'catalog': 'catalog',
        'jira_search': 'jira-search',
        'epic_summaries': 'epic-summaries',
        'build_payload': 'build-payload',
        'cache_store': 'cache-store',
    }
    parts = []
    for key, token in token_names.items():
        value = timings_ms.get(key)
        if value is not None:
            parts.append(f'{token};dur={value}')
    return ', '.join(parts)


def _build_excluded_capacity_stats_source_fields(story_points_field, sprint_field_id, epic_link_field_id, team_field_id):
    fields = []
    for field_id in (story_points_field, 'parent', 'project', sprint_field_id, epic_link_field_id, team_field_id):
        if field_id and field_id not in fields:
            fields.append(field_id)
    return fields


def _excluded_capacity_stats_source_cache_key(context, sprint_ids, team_ids, jql, fields_list):
    sprint_signature = ','.join(str(item or '').strip() for item in (sprint_ids or []) if str(item or '').strip())
    team_signature = ','.join(normalize_team_ids(team_ids or []))
    source_signature = hashlib.sha256(
        json.dumps({
            'jql': jql,
            'fields': fields_list,
        }, sort_keys=True).encode('utf-8')
    ).hexdigest()
    parts = (
        'excluded-capacity-stats-source',
        sprint_signature,
        team_signature,
        source_signature,
    )
    if context is not None:
        return build_jira_home_process_cache_key(context, *parts)
    return parts


def fetch_excluded_capacity_stats_source(sprint_ids, context=None, team_ids=None, refresh=False, timings_ms=None):
    field_started = time.perf_counter()
    team_field_id = resolve_team_field_id(None, context=context)
    epic_link_field_id = resolve_epic_link_field_id(None, context=context)
    sprint_field_id = get_sprint_field_id()
    story_points_field = get_story_points_field_id()
    _record_excluded_capacity_timing(timings_ms, 'field_config', field_started)

    catalog_started = time.perf_counter()
    try:
        catalog = load_team_catalog() or {}
    except Exception:
        catalog = {}
    team_name_by_id = {}
    for cid, entry in (catalog or {}).items():
        if isinstance(entry, dict):
            name = str(entry.get('name') or '').strip()
            if cid and name:
                team_name_by_id[str(cid).strip()] = name
    _record_excluded_capacity_timing(timings_ms, 'catalog', catalog_started)

    jql = build_excluded_capacity_stats_jql(sprint_ids, team_ids=team_ids)
    fields_list = _build_excluded_capacity_stats_source_fields(
        story_points_field,
        sprint_field_id,
        epic_link_field_id,
        team_field_id,
    )

    cache_enabled = context is not None and jira_home_partitioned_process_cache_enabled(context)
    cache_key = _excluded_capacity_stats_source_cache_key(context, sprint_ids, team_ids, jql, fields_list)
    if cache_enabled and not refresh:
        now = time.time()
        with _cache_lock:
            cached = EXCLUDED_CAPACITY_STATS_SOURCE_CACHE.get(cache_key)
            if cached and now - cached.get('timestamp', 0) < EXCLUDED_CAPACITY_STATS_SOURCE_CACHE_TTL_SECONDS:
                if timings_ms is not None:
                    timings_ms['cache'] = 1
                return copy.deepcopy(cached.get('data') or {}), None

    warnings = []
    collected_issues = []
    next_page_token = None
    page_count = 0
    page_size = 100

    jira_search_started = time.perf_counter()
    while len(collected_issues) < EXCLUDED_CAPACITY_STATS_MAX_ISSUES:
        payload = {
            'jql': jql,
            'maxResults': min(page_size, EXCLUDED_CAPACITY_STATS_MAX_ISSUES - len(collected_issues)),
            'fields': fields_list
        }
        if next_page_token:
            payload['nextPageToken'] = next_page_token
        response = jira_search_request(payload, context=context)
        if response.status_code != 200:
            return None, response
        data = response.json() or {}
        issues = data.get('issues', []) or []
        collected_issues.extend(issues)
        page_count += 1
        next_page_token = data.get('nextPageToken')
        if data.get('isLast', True) or not next_page_token or not issues:
            break
    _record_excluded_capacity_timing(timings_ms, 'jira_search', jira_search_started)

    if len(collected_issues) >= EXCLUDED_CAPACITY_STATS_MAX_ISSUES:
        warnings.append(f'issue fetch capped at {EXCLUDED_CAPACITY_STATS_MAX_ISSUES} issues')

    epic_keys = []
    seen_epics = set()
    for issue in collected_issues:
        issue_fields = issue.get('fields', {}) or {}
        epic_key = None
        if epic_link_field_id and issue_fields.get(epic_link_field_id):
            epic_key = issue_fields.get(epic_link_field_id)
        else:
            parent_field = issue_fields.get('parent') or {}
            if parent_field.get('key') and \
                    (parent_field.get('fields') or {}).get('issuetype', {}).get('name', '').lower() == 'epic':
                epic_key = parent_field.get('key')
        if epic_key and epic_key not in seen_epics:
            seen_epics.add(epic_key)
            epic_keys.append(epic_key)

    epic_summary_started = time.perf_counter()
    epic_summary_by_key = fetch_cached_excluded_capacity_epic_summaries(epic_keys, context=context)
    _record_excluded_capacity_timing(timings_ms, 'epic_summaries', epic_summary_started)

    build_payload_started = time.perf_counter()
    issues_payload = [
        build_excluded_capacity_issue_payload(issue, team_field_id, epic_link_field_id, sprint_field_id, epic_summary_by_key, team_name_by_id)
        for issue in collected_issues
    ]

    result = {
        'issues': issues_payload,
        'meta': {
            'warnings': warnings,
            'truncated': bool(warnings),
            'paginationMode': 'nextPageToken/isLast',
            'queryPages': page_count,
            'issueLimit': EXCLUDED_CAPACITY_STATS_MAX_ISSUES
        }
    }
    _record_excluded_capacity_timing(timings_ms, 'build_payload', build_payload_started)

    if cache_enabled:
        cache_store_started = time.perf_counter()
        with _cache_lock:
            EXCLUDED_CAPACITY_STATS_SOURCE_CACHE[cache_key] = {
                'timestamp': time.time(),
                'data': copy.deepcopy(result)
            }
        _record_excluded_capacity_timing(timings_ms, 'cache_store', cache_store_started)

    return result, None


def get_excluded_capacity_stats_source():
    payload = request.get_json(silent=True) or {}
    raw_sprint_ids = payload.get('sprintIds') if isinstance(payload, dict) else []
    sprint_ids = [str(item or '').strip() for item in (raw_sprint_ids if isinstance(raw_sprint_ids, list) else []) if str(item or '').strip()]
    if not sprint_ids:
        return jsonify({'error': 'sprintIds is required'}), 400
    if len(sprint_ids) > EXCLUDED_CAPACITY_STATS_MAX_SPRINTS:
        return jsonify({'error': f'sprintIds is limited to {EXCLUDED_CAPACITY_STATS_MAX_SPRINTS} sprints'}), 400

    raw_team_ids = payload.get('teamIds') if isinstance(payload, dict) else []
    team_ids = normalize_team_ids(raw_team_ids if isinstance(raw_team_ids, list) else [])
    refresh = bool(payload.get('refresh')) if isinstance(payload, dict) else False

    try:
        auth_context = current_request_auth_context()
        timings_ms = {}
        stats_payload, error_response = fetch_excluded_capacity_stats_source(
            sprint_ids,
            context=auth_context,
            team_ids=team_ids,
            refresh=refresh,
            timings_ms=timings_ms,
        )
        if error_response is not None:
            return jsonify({
                'error': 'Failed to fetch excluded-capacity stats source',
                'details': error_response.text
            }), error_response.status_code

        response = jsonify({
            'cached': 'cache' in timings_ms,
            'generatedAt': datetime.now().isoformat(),
            'data': stats_payload
        })
        server_timing = _excluded_capacity_server_timing_header(timings_ms)
        if server_timing:
            response.headers['Server-Timing'] = server_timing
        return response
    except AuthError as error:
        if error.code == "auth_required":
            payload, status = oauth_auth_required_payload()
            return jsonify(payload), status
        raise
    except Exception as error:
        logger.exception('Failed to fetch excluded-capacity stats source')
        return jsonify({
            'error': 'Failed to fetch excluded-capacity stats source',
            'message': str(error)
        }), 500


def get_completed_sprint_stats():
    """Fetch cached delivery stats for a completed sprint."""
    sprint_name = request.args.get('sprint', '').strip()
    team_id = request.args.get('team', '').strip()
    team_ids_raw = request.args.get('teamIds', '').strip()
    group_id = request.args.get('groupId', '').strip()
    refresh = request.args.get('refresh', '').lower() == 'true'

    if not sprint_name:
        return jsonify({'error': 'Missing sprint name'}), 400

    base_jql = STATS_JQL_BASE or build_base_jql()
    team_ids = []
    if team_ids_raw:
        team_ids = [t.strip() for t in team_ids_raw.split(',') if t.strip()]
    elif team_id:
        team_ids = [team_id]
    else:
        team_ids = get_stats_team_ids()
    auth_context = current_request_auth_context()
    cache_enabled = jira_home_process_cache_enabled(auth_context)
    cache_key = build_stats_cache_key(sprint_name, base_jql, team_ids, group_id=group_id)
    cache_data = {}
    if cache_enabled:
        cache_data = load_stats_cache()
        if not refresh and cache_key in cache_data:
            cached_payload = cache_data.get(cache_key, {})
            response = {
                'cached': True,
                'generatedAt': cached_payload.get('generatedAt'),
                'data': cached_payload.get('data')
            }
            return jsonify(response)

    try:
        team_field_id = resolve_team_field_id(None, context=auth_context)
        stats_payload, error_response = fetch_stats_for_sprint(sprint_name, None, team_field_id, team_ids=team_ids or None)
        if error_response is not None:
            return jsonify({
                'error': 'Failed to fetch stats',
                'details': error_response.text
            }), error_response.status_code

        generated_at = datetime.now().isoformat()
        if cache_enabled:
            cache_data[cache_key] = {
                'generatedAt': generated_at,
                'data': stats_payload
            }
            save_stats_cache(cache_data)

        return jsonify({
            'cached': False,
            'generatedAt': generated_at,
            'data': stats_payload
        })
    except AuthError:
        payload, status = oauth_auth_required_payload()
        return jsonify(payload), status


def get_burnout_stats():
    """Fetch sprint burnout events from Jira changelog on demand."""
    payload = request.get_json(silent=True) if request.method == 'POST' else None
    def parse_bool(value):
        if isinstance(value, bool):
            return value
        return str(value or '').strip().lower() in ('1', 'true', 'yes', 'on')
    sprint_name = str((payload or {}).get('sprint') or request.args.get('sprint', '')).strip()
    raw_team_ids = (payload or {}).get('teamIds') if isinstance(payload, dict) else None
    if isinstance(raw_team_ids, list):
        team_ids_raw = ','.join(str(item or '').strip() for item in raw_team_ids if str(item or '').strip())
    else:
        team_ids_raw = str(raw_team_ids or request.args.get('teamIds', '')).strip()
    team_id = str((payload or {}).get('team') or request.args.get('team', '')).strip()
    raw_issue_keys = (payload or {}).get('issueKeys') if isinstance(payload, dict) else None
    include_post_sprint_closures = parse_bool(
        (payload or {}).get('includePostSprintClosures')
        if isinstance(payload, dict)
        else request.args.get('includePostSprintClosures', '')
    )
    issue_keys = []
    if isinstance(raw_issue_keys, list):
        issue_keys = [str(key or '').strip() for key in raw_issue_keys if str(key or '').strip()]
    if not sprint_name:
        return jsonify({'error': 'Missing sprint name'}), 400

    scoped_team_ids = []
    if team_ids_raw:
        scoped_team_ids = normalize_team_ids(team_ids_raw.split(','))
    elif team_id and team_id.lower() != 'all':
        scoped_team_ids = normalize_team_ids([team_id])

    try:
        auth_context = current_request_auth_context()
        cache_enabled = jira_home_process_cache_enabled(auth_context)
        team_field_id = resolve_team_field_id(None, context=auth_context)
        burnout_payload, error_response, debug_payload = fetch_burnout_events_for_sprint(
            sprint_name,
            None,
            team_field_id,
            team_ids=scoped_team_ids,
            issue_keys=issue_keys,
            include_post_sprint_closures=include_post_sprint_closures,
            cache_enabled=cache_enabled
        )
        if error_response is not None:
            return jsonify({
                'error': 'Failed to fetch burnout stats',
                'details': error_response.text,
                'query': debug_payload
            }), error_response.status_code

        return jsonify({
            'generatedAt': datetime.now().isoformat(),
            'data': burnout_payload
        })
    except AuthError:
        payload, status = oauth_auth_required_payload()
        return jsonify(payload), status


def get_epic_cohort_stats():
    payload = request.get_json(silent=True) or {}
    start_quarter = str(payload.get('startQuarter') or '').strip()
    if not start_quarter:
        return jsonify({'error': 'startQuarter is required'}), 400

    raw_team_ids = payload.get('teamIds')
    team_ids = normalize_team_ids(raw_team_ids if isinstance(raw_team_ids, list) else [])
    raw_components = payload.get('components')
    component_names = [str(item or '').strip() for item in (raw_components if isinstance(raw_components, list) else []) if str(item or '').strip()]
    refresh = _cohort_parse_bool(payload.get('refresh'))
    scoped_projects = _cohort_project_scope()
    cache_key = _build_epic_cohort_cache_key(start_quarter, team_ids, scoped_projects, component_names)
    auth_context = current_request_auth_context()
    cache_enabled = jira_home_process_cache_enabled(auth_context)

    now_ts = time.time()
    cached = None
    if cache_enabled:
        with _cache_lock:
            cached = EPIC_COHORT_CACHE.get(cache_key)
    if cache_enabled and cached and not refresh and (now_ts - float(cached.get('ts') or 0)) <= EPIC_COHORT_CACHE_TTL_SECONDS:
        return jsonify({
            'cached': True,
            'generatedAt': cached.get('generatedAt'),
            'data': cached.get('data')
        })

    try:
        team_field_id = resolve_team_field_id(None, context=auth_context)
        cohort_payload, error_response = fetch_epic_cohort_data(
            start_quarter,
            None,
            team_field_id,
            team_ids=team_ids,
            component_names=component_names,
            context=auth_context,
        )
        if error_response is not None:
            return jsonify({
                'error': 'Failed to fetch epic cohort stats',
                'details': error_response.text
            }), error_response.status_code
    except AuthError:
        payload, status = oauth_auth_required_payload()
        return jsonify(payload), status

    generated_at = datetime.now().isoformat()
    if cache_enabled:
        with _cache_lock:
            EPIC_COHORT_CACHE[cache_key] = {
                'ts': now_ts,
                'generatedAt': generated_at,
                'data': cohort_payload
            }

    return jsonify({
        'cached': False,
        'generatedAt': generated_at,
        'data': cohort_payload
    })


@app.route('/favicon.ico')
def get_favicon():
    favicon_path = os.path.join(os.path.dirname(__file__), 'favicon.ico')
    if os.path.exists(favicon_path):
        return send_file(favicon_path, mimetype='image/x-icon')
    return '', 404


@app.route('/epm-burst.svg')
def get_epm_burst_icon():
    icon_path = os.path.join(os.path.dirname(__file__), 'epm-burst.svg')
    if os.path.exists(icon_path):
        return send_file(icon_path, mimetype='image/svg+xml')
    return '', 404


@app.route('/', methods=['GET'])
def serve_dashboard():
    dashboard_path = os.path.join(os.path.dirname(__file__), 'jira-dashboard.html')
    if os.path.exists(dashboard_path):
        return send_file(dashboard_path)
    return 'Dashboard not found', 404


@app.route('/jira-dashboard.html', methods=['GET'])
def serve_dashboard_file():
    return serve_dashboard()


@app.route('/frontend/dist/<path:filename>')
def serve_frontend_dist(filename):
    """Serve the bundled frontend assets."""
    dist_dir = os.path.join(os.path.dirname(__file__), 'frontend', 'dist')
    response = send_from_directory(dist_dir, filename)
    if filename.endswith('.map'):
        response.mimetype = 'application/json'
    return response


PROJECTS_CACHE = {'data': None, 'timestamp': 0}
PROJECTS_CACHE_TTL = 60 * 60  # 1 hour

COMPONENTS_CACHE = {'data': None, 'timestamp': 0}
COMPONENTS_CACHE_TTL = 60 * 60  # 1 hour

EPICS_SEARCH_CACHE = {}
EPICS_SEARCH_CACHE_TTL = 60 * 5  # 5 minutes

LABELS_CACHE = {'data': None, 'timestamp': 0}
LABELS_CACHE_TTL = 15 * 60  # 15 minutes


# --- Custom Field Config Endpoints ---

def _save_field_config(config_key, cache_name=None):
    """Generic helper to save a field config (fieldId + fieldName) into dashboard-config.json."""
    payload = request.get_json(silent=True) or {}
    field_id = str(payload.get('fieldId', '')).strip()
    field_name = str(payload.get('fieldName', '')).strip()
    try:
        dashboard_config = load_dashboard_config() or {'version': 1, 'projects': {'selected': []}, 'teamGroups': {}}
        dashboard_config[config_key] = {'fieldId': field_id, 'fieldName': field_name}
        save_dashboard_config(dashboard_config)
        # Invalidate tasks cache so next fetch uses the new field
        global TASKS_CACHE
        TASKS_CACHE = {}
        # Invalidate the specific resolve cache if applicable
        if cache_name:
            g = globals()
            with _cache_lock:
                g[cache_name] = None
    except Exception as e:
        return jsonify({'error': f'Failed to save {config_key} config', 'message': str(e)}), 500
    return jsonify({'fieldId': field_id, 'fieldName': field_name})


# --- Issue Types ---
ISSUE_TYPES_CACHE = {'data': None, 'timestamp': 0}
ISSUE_TYPES_CACHE_TTL = 60 * 60  # 1 hour


@app.route('/health', methods=['GET'])
def health_check():
    """Health check endpoint"""
    return jsonify({
        'status': 'OK',
        'message': 'Jira proxy server is running'
    })


def test_connection():
    """Test Jira connection with simple query"""
    try:
        log_info('Testing Jira connection')

        response = current_jira_get('/rest/api/3/myself', timeout=15)

        log_info(f'Test response status={response.status_code}')

        if response.status_code != 200:
            return jsonify({
                'status': 'error',
                'code': response.status_code,
                'message': response.text
            }), response.status_code

        data = response.json()
        return jsonify({
            'status': 'success',
            'message': f'Connection OK! Authenticated as {data.get("displayName") or data.get("emailAddress") or "Jira user"}',
            'sample_issue': None,
        })
    except AuthError as e:
        if e.code == 'auth_required':
            save_oauth_session({})
            return jsonify({
                'error': 'auth_required',
                'message': 'Your Jira sign-in expired. Sign in again to continue.',
                'loginUrl': '/login?reason=session_expired',
            }), 401
        return jsonify({
            'status': 'error',
            'message': str(e),
        }), 500

    except Exception as e:
        return jsonify({
            'status': 'error',
            'message': str(e)
        }), 500


def debug_fields():
    """Debug endpoint to see all fields of a single task"""
    if not dev_diagnostics_allowed():
        return jsonify({'error': 'not_found'}), 404
    try:
        # Get one issue with ALL fields
        payload = {
            'jql': build_base_jql() or 'ORDER BY created DESC',
            'maxResults': 1,
            'fields': ['*all']
        }

        log_info('Fetching all fields for debugging')

        response = jira_search_request(payload)

        if response.status_code != 200:
            return jsonify({
                'error': f'Jira API error: {response.status_code}',
                'details': response.text
            }), response.status_code

        data = response.json()

        if data.get('issues') and len(data['issues']) > 0:
            issue = data['issues'][0]
            fields = issue.get('fields', {})

            # Look for Story Points in customfields
            customfields = {}
            for key, value in fields.items():
                if key.startswith('customfield_') and value is not None:
                    customfields[key] = value

            return jsonify({
                'issue_key': issue.get('key'),
                'all_customfields': customfields,
                'fields_keys': list(fields.keys())
            })
        else:
            return jsonify({
                'error': 'No issues found',
                'jql': build_base_jql()
            }), 404

    except Exception as e:
        return jsonify({
            'error': str(e)
        }), 500


def get_tasks_fields():
    """Return all available fields and values for issues matching JQL_QUERY."""
    if not dev_diagnostics_allowed():
        return jsonify({'error': 'not_found'}), 404
    try:
        limit = request.args.get('limit', '5')
        try:
            limit_value = max(1, min(int(limit), 50))
        except ValueError:
            limit_value = 5

        payload = {
            'jql': build_base_jql() or 'ORDER BY created DESC',
            'maxResults': limit_value,
            'fields': ['*all']
        }

        log_info(f'Fetching all fields for {limit_value} issues')

        response = jira_search_request(payload)

        if response.status_code != 200:
            return jsonify({
                'error': f'Jira API error: {response.status_code}',
                'details': response.text
            }), response.status_code

        data = response.json()
        issues = data.get('issues', [])

        return jsonify({
            'total': data.get('total'),
            'returned': len(issues),
            'issues': issues
        })

    except Exception as e:
        return jsonify({
            'error': str(e)
        }), 500


def export_excel():
    """Export selected tasks to Excel file"""
    try:
        data = request.get_json()
        tasks = data.get('tasks', [])

        if not tasks:
            return jsonify({'error': 'No tasks provided'}), 400

        log_info(f'Exporting {len(tasks)} tasks to Excel')

        # Create a new workbook
        wb = Workbook()
        ws = wb.active
        ws.title = 'Sprint Tasks'

        # Define header style
        header_fill = PatternFill(start_color='107C41', end_color='107C41', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF', size=12)
        header_alignment = Alignment(horizontal='center', vertical='center')

        # Add headers
        headers = ['ID', 'Subject', 'Story Points']
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

        # Add data
        for row_num, task in enumerate(tasks, 2):
            ws.cell(row=row_num, column=1, value=task.get('key', ''))
            ws.cell(row=row_num, column=2, value=task.get('summary', ''))
            ws.cell(row=row_num, column=3, value=task.get('storyPoints', 0))

        # Auto-adjust column widths
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 60
        ws.column_dimensions['C'].width = 15

        # Align Story Points column to center
        for row in range(2, len(tasks) + 2):
            ws.cell(row=row, column=3).alignment = Alignment(horizontal='center')

        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        log_info('Excel file generated successfully')

        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'sprint_tasks_{datetime.now().strftime("%Y-%m-%d")}.xlsx'
        )

    except Exception as e:
        logger.exception('Export Excel error')
        return jsonify({
            'error': 'Failed to export to Excel',
            'message': str(e)
        }), 500


def main():
    global JIRA_URL, JIRA_EMAIL, JIRA_TOKEN, JQL_QUERY, SERVER_PORT

    args = parse_args()

    # Apply CLI overrides while keeping env defaults as fallbacks
    if args.jira_url:
        JIRA_URL = args.jira_url
        os.environ['JIRA_URL'] = args.jira_url
    if args.jira_email:
        JIRA_EMAIL = args.jira_email
    if args.jira_token:
        JIRA_TOKEN = args.jira_token
    if args.jira_query:
        JQL_QUERY = args.jira_query
    if args.server_port:
        SERVER_PORT = args.server_port

    # Validate configuration
    try:
        validate_startup_auth_config()
    except AuthError as error:
        log_error(str(error))
        log_info('Please copy .env.example to .env and configure either basic auth or Atlassian OAuth')
        return 1

    try:
        bind_host = validate_network_bind(default_bind_host())
    except AuthError as error:
        log_error(str(error))
        return 1

    # Only print on first startup, not on reload
    if not DEBUG_MODE or os.environ.get('WERKZEUG_RUN_MAIN') != 'true':
        # Calculate current quarter for examples
        today = date.today()
        current_quarter = f"{today.year}Q{(today.month - 1) // 3 + 1}"

        display_host = 'localhost' if bind_host in {'127.0.0.1', '::1'} else bind_host
        log_info(f'Jira Proxy Server starting on http://{display_host}:{SERVER_PORT}')
        log_info(f'   Jira: {JIRA_URL}')
        log_info(f'   Auth mode: {JIRA_AUTH_MODE}')
        if JIRA_AUTH_MODE == AUTH_MODE_BASIC:
            log_info(f'   Email: {JIRA_EMAIL}')
        effective_board_id = get_effective_board_id()
        if effective_board_id:
            log_info(f'   Board: {effective_board_id}')
        if GROUPS_CONFIG_PATH and os.path.exists(GROUPS_CONFIG_PATH):
            log_info(f'   Groups: {GROUPS_CONFIG_PATH}')
        log_info('Key Endpoints:')
        log_info(f'   • http://localhost:{SERVER_PORT}/api/tasks?sprint={current_quarter}')
        log_info(f'   • http://localhost:{SERVER_PORT}/api/teams?sprint={current_quarter}&all=true')
        log_info(f'   • http://localhost:{SERVER_PORT}/api/teams/all?sprint={current_quarter}  (all teams with names)')
        log_info(f'   • http://localhost:{SERVER_PORT}/api/sprints')
        log_info(f'   • http://localhost:{SERVER_PORT}/api/groups-config')
        log_info()

    app.run(host=bind_host, port=SERVER_PORT, debug=DEBUG_MODE)
    return 0


if __name__ == '__main__':
    raise SystemExit(main())
